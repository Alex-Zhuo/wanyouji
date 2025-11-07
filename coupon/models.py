# coding=utf-8
from __future__ import unicode_literals

from django.core.validators import FileExtensionValidator, validate_image_file_extension
from django.db import models
from django.utils import timezone
from django.conf import settings
from common.config import FILE_FIELD_PREFIX, IMAGE_FIELD_PREFIX
from mall.models import validate_positive_int_gen
from mp.models import WeiXinPayConfig
from restframework_ext.models import UseNoAbstract, ReceiptAbstract
from ticket.models import ShowProject, ShowType, TicketOrder, ShowContentCategorySecond, SessionInfo
from caches import get_pika_redis, get_redis_name
import logging
from django.db import close_old_connections
from decimal import Decimal
from common.utils import quantize, get_short_no
from django.core.exceptions import ValidationError
from restframework_ext.exceptions import CustomAPIException
from random import sample
import json
from django.db.transaction import atomic

log = logging.getLogger(__name__)
notify_url = '/api/coupon/receipt/notify/'
refund_notify_url = '/api/coupon/receipt/refund_notify/'


def coupon_cancel_minutes_limit(value):
    if value < 5:
        raise ValidationError("必须大于等于5分钟")
    else:
        return value


def coupon_order_no(tail_length=3):
    """
    使用当前时间(datetime)生成随机字符串,可以作为订单号
    :return:
    """
    now = timezone.now()
    return 'CO%s%s' % (now.strftime('%Y%m%d%H%M%S%f'), ''.join(sample(list(map(str, range(0, 10))), tail_length)))


def coupon_refund_no(tail_length=3):
    """
    使用当前时间(datetime)生成随机字符串,可以作为订单号
    :return:
    """
    now = timezone.now()
    return 'CR%s%s' % (now.strftime('%Y%m%d%H%M%S%f'), ''.join(sample(list(map(str, range(0, 10))), tail_length)))


class CouponBasic(models.Model):
    image = models.ImageField('弹窗图片', upload_to=f'{IMAGE_FIELD_PREFIX}/coupon/basic',
                              validators=[validate_image_file_extension])
    auto_cancel_minutes = models.PositiveSmallIntegerField('自动关闭订单分钟数', default=5,
                                                           help_text='订单创建时间开始多少分钟后未支付自动取消订单',
                                                           validators=[coupon_cancel_minutes_limit], editable=False)

    class Meta:
        verbose_name_plural = verbose_name = '消费券配置'

    @classmethod
    def get(cls):
        return cls.objects.first()


class Coupon(UseNoAbstract):
    TYPE_MONEY_OFF = 1
    TYPE_MONEY_DISCOUNT = 2
    TYPE_NUM_DISCOUNT = 3
    COUPON_TYPE_CHOICES = ((TYPE_MONEY_OFF, '满额减券'), (TYPE_MONEY_DISCOUNT, '满额打折券'), (TYPE_NUM_DISCOUNT, '满张(套)数打折券'))
    name = models.CharField('名称', max_length=128)
    type = models.PositiveSmallIntegerField('优惠类型', choices=COUPON_TYPE_CHOICES, default=TYPE_MONEY_OFF)
    SR_FREE = 1
    SR_PAY = 2
    SOURCE_TYPE_CHOICES = ((SR_FREE, '免费领取'), (SR_PAY, '付费购买'))
    source_type = models.PositiveSmallIntegerField('消费券类型', choices=SOURCE_TYPE_CHOICES, default=SR_FREE)
    amount = models.DecimalField('减免金额', max_digits=13, decimal_places=2, default=0)
    pay_amount = models.DecimalField('消费券价格', max_digits=13, decimal_places=2, default=0)
    discount = models.PositiveSmallIntegerField('打折比率', default=0, help_text='80为打8折')
    require_amount = models.DecimalField('使用满足金额', max_digits=13, decimal_places=2, default=0, help_text='满减券、满额打折券必填')
    require_num = models.PositiveSmallIntegerField('使用满足张(套)数', default=0, help_text='满张(套)数打折券必填')
    expire_time = models.DateTimeField('使用截止时间')
    user_tips = models.TextField('使用提示', help_text='提示使用范围等')
    STATUS_ON = 2
    STATUS_OFF = 1
    STATUS_CHOICES = ((STATUS_ON, '上架'), (STATUS_OFF, '下架'))
    status = models.IntegerField('状态', choices=STATUS_CHOICES, default=STATUS_OFF)
    off_use = models.BooleanField('下架后是否可继续使用', default=True)
    stock = models.IntegerField('库存数量', default=0)
    user_obtain_limit = models.IntegerField('用户限领(购)次数', default=0, help_text='0为不限次数')
    shows = models.ManyToManyField(ShowProject, verbose_name='可使用的节目', related_name='shows', blank=True)
    limit_show_types_second = models.ManyToManyField(ShowContentCategorySecond, verbose_name='可使用节目分类',
                                                     related_name='show_types_second', blank=True, editable=False)
    create_at = models.DateTimeField('创建时间', auto_now_add=True)
    update_at = models.DateTimeField('更新时间', auto_now=True)

    class Meta:
        verbose_name_plural = verbose_name = '消费券'

    def __str__(self):
        return self.name

    def clean(self):
        if self.type == self.TYPE_MONEY_OFF:
            if self.amount <= 0:
                raise ValidationError('减免金额必须大于0')
        elif self.type in [self.TYPE_MONEY_DISCOUNT, self.TYPE_NUM_DISCOUNT]:
            if self.discount <= 0:
                raise ValidationError('打折比率必须大于0')
        if self.source_type == self.SR_PAY and self.pay_amount <= 0:
            raise ValidationError('付费购买消费券，价格必须大于0')

    @property
    def need_buy(self):
        return self.pay_amount > 0 and self.source_type == self.SR_PAY

    @classmethod
    def auto_off_task(cls):
        """
        自动下架
        """
        cls.objects.filter(status=cls.STATUS_ON, expire_time__lte=timezone.now()).update(status=cls.STATUS_OFF)

    def check_can_use(self):
        st = True
        if self.status == self.STATUS_OFF and not self.off_use:
            st = False
        return st

    @classmethod
    def amount_type(cls):
        # 满额类型
        return [cls.TYPE_MONEY_OFF, cls.TYPE_MONEY_DISCOUNT]

    @classmethod
    def mul_type(cls):
        # 满张类型
        return [cls.TYPE_NUM_DISCOUNT]

    @classmethod
    def discount_type(cls):
        # 打折类型
        return [cls.TYPE_NUM_DISCOUNT, cls.TYPE_MONEY_DISCOUNT]

    @classmethod
    def pop_up_key(cls, user_id: int):
        key = get_redis_name('cp_pop_up')
        name = get_redis_name(str(user_id))
        return key, name

    @classmethod
    def set_pop_up(cls, user_id: int):
        key, name = cls.pop_up_key(user_id)
        with get_pika_redis() as pika:
            pika.hset(key, name, 1)

    @classmethod
    def get_pop_up(cls, user_id: int):
        key, name = cls.pop_up_key(user_id)
        with get_pika_redis() as pika:
            return pika.hget(key, name)

    @classmethod
    def del_pop_up(cls):
        key = get_redis_name('cp_pop_up')
        with get_pika_redis() as pika:
            pika.delete(key)

    @classmethod
    def coupon_update_stock_from_redis(cls):
        close_old_connections()
        from coupon.stock_updater import csc
        csc.persist()

    def coupon_change_stock(self, mul):
        ret = True
        coupon_upd = []
        coupon_upd.append((self.pk, mul, 0))
        from coupon.stock_updater import csc
        succ1, tfc_result = csc.batch_incr(coupon_upd)
        if succ1:
            csc.batch_record_update_ts(csc.resolve_ids(tfc_result))
        else:
            log.warning(f"coupon incr failed,{self.pk}")
            # 库存不足
            ret = False
        return ret

    def coupon_redis_stock(self, stock=None):
        # 初始化库存
        from coupon.stock_updater import csc, StockModel
        if stock == None:
            stock = self.stock
        csc.append_cache(StockModel(_id=self.id, stock=stock))

    def coupon_del_redis_stock(self):
        from coupon.stock_updater import csc
        csc.remove(self.id)


class UserCouponRecord(UseNoAbstract):
    user = models.ForeignKey(settings.AUTH_USER_MODEL, verbose_name='用户', related_name='coupons',
                             on_delete=models.CASCADE)
    coupon = models.ForeignKey(Coupon, verbose_name='优惠券', on_delete=models.CASCADE)
    coupon_type = models.PositiveSmallIntegerField('优惠卷类型', choices=Coupon.COUPON_TYPE_CHOICES,
                                                   default=Coupon.TYPE_MONEY_OFF)
    STATUS_DEFAULT = 1
    STATUS_USE = 2
    STATUS_EXPIRE = 3
    STATUS_INVALID = 4
    STATUS_CHOICES = ((STATUS_DEFAULT, u'未使用'), (STATUS_USE, u'已使用'), (STATUS_EXPIRE, u'已过期'), (STATUS_INVALID, u'已作废'))
    status = models.IntegerField(u'状态', choices=STATUS_CHOICES, default=STATUS_DEFAULT)
    expire_time = models.DateTimeField('使用截止时间')
    amount = models.DecimalField(u'减免金额', max_digits=13, decimal_places=2, default=0)
    discount = models.PositiveSmallIntegerField('打折比率', default=0, help_text='80为打8折')
    require_amount = models.DecimalField('使用满足金额', max_digits=13, decimal_places=2, default=0, help_text='满减券、满额打折券必填')
    require_num = models.PositiveSmallIntegerField('使用满足张数', default=0, help_text='满张数打折券必填')
    used_time = models.DateTimeField('使用时间', null=True, blank=True)
    create_at = models.DateTimeField('领取时间', auto_now_add=True)
    buy_order = models.OneToOneField('CouponOrder', verbose_name='购买订单', null=True, blank=True,
                                     on_delete=models.SET_NULL,
                                     related_name='buy_order')
    order = models.OneToOneField(TicketOrder, verbose_name='使用订单', null=True, blank=True, on_delete=models.SET_NULL,
                                 related_name='coupon_order')
    snapshot = models.TextField('优惠券快照', null=True, blank=True, help_text='领取时保存的快照', editable=False)

    class Meta:
        verbose_name_plural = verbose_name = '用户消费券记录'

    def __str__(self):
        return '{}:{}'.format(self.user, self.coupon)

    @classmethod
    def create_record(cls, user_id: int, coupon, buy_order: 'CouponOrder' = None):
        obj = cls.objects.create(user_id=user_id, coupon=coupon, expire_time=coupon.expire_time, amount=coupon.amount,
                                 discount=coupon.discount, coupon_type=coupon.type,
                                 require_amount=coupon.require_amount, require_num=coupon.require_num,
                                 buy_order=buy_order)
        obj.save_common()
        # 增加购买数量
        cls.set_user_obtain_cache(coupon.no, user_id, 1)
        try:
            Coupon.set_pop_up(user_id)
        except Exception as e:
            log.error(e)
        return obj

    @property
    def can_use(self):
        return self.status == self.STATUS_DEFAULT and self.expire_time > timezone.now()

    def save_common(self):
        fields = ['snapshot']
        self.snapshot = self.get_snapshot()
        self.save(update_fields=fields)

    @classmethod
    def check_expire_task(cls):
        """
        自动过期任务
        """
        qs = cls.objects.filter(status=cls.STATUS_DEFAULT, expire_time__lte=timezone.now())
        for obj in qs:
            obj.status = cls.STATUS_EXPIRE
            obj.save(update_fields=['status'])
            if obj.buy_order:
                obj.buy_order.do_refund(op_user=None, refund_reason='过期自动退款')

    @classmethod
    def user_obtain_key(cls, coupon_no: str, user_id: int):
        key = get_redis_name('coupon{}'.format(coupon_no))
        name = get_redis_name('user{}'.format(user_id))
        return key, name

    @classmethod
    def get_user_obtain_cache(cls, coupon_no: str, user_id: int) -> int:
        key, name = cls.user_obtain_key(coupon_no, user_id)
        with get_pika_redis() as redis:
            num = redis.hget(key, name) or 0
        return int(num)

    @classmethod
    def set_user_obtain_cache(cls, coupon_no: str, user_id: int, num=1):
        key, name = cls.user_obtain_key(coupon_no, user_id)
        with get_pika_redis() as redis:
            redis.hincrby(key, name, num)

    def get_snapshot(self):
        import json
        """
        消费券快照
        :return:
        """
        coupon = self.coupon
        shows_ids = []
        shows_names = []
        show_types_ids = []
        show_types_names = []
        for show in coupon.shows.all():
            shows_ids.append(show.no)
            shows_names.append(show.title)
        # for f in coupon.limit_show_types_second.all():
        #     show_types_ids.append(f.id)
        #     show_types_names.append(f.name)
        data = dict(name=coupon.name, no=coupon.no, amount=float(coupon.amount),
                    user_obtain_limit=coupon.user_obtain_limit,
                    shows_nos=shows_ids, shows_names=shows_names, show_types_second_ids=show_types_ids,
                    show_types_names=show_types_names, user_tips=coupon.user_tips)
        return json.dumps(data)

    def check_can_show_use(self, show: ShowProject):
        # snapshot = json.loads(self.snapshot)
        # limit_show_types_second_ids = snapshot['show_types_second_ids']
        # limit_shows_nos = snapshot['shows_nos']
        # can_use = False
        # if not (limit_show_types_second_ids and show.cate_second.id not in limit_show_types_second_ids):
        #     can_use = True
        # if not can_use and not (limit_shows_nos and show.no not in limit_shows_nos):
        #     can_use = True
        can_use = True
        shows = self.coupon.shows.all()
        # show_types = self.coupon.limit_show_types_second.all()
        # if show_types and not show_types.filter(id=show.cate_second.id).exists():
        #     can_use = False
        if shows and not shows.filter(id=show.id).exists():
            can_use = False
        return can_use

    def check_type_use(self, amount: Decimal = 0, multiply: int = 0):
        can_use = True
        if self.coupon_type in Coupon.amount_type():
            if amount < self.require_amount:
                can_use = False
        elif self.coupon_type in Coupon.mul_type():
            if multiply < self.require_num:
                can_use = False
        return can_use

    def coupon_check_can_use(self, show: ShowProject, amount: Decimal = 0, multiply: int = 0):
        # snapshot = json.loads(self.snapshot)
        # limit_show_types_second_ids = snapshot['show_types_second_ids']
        # limit_shows_nos = snapshot['shows_nos']
        # can_use = False
        # if not (limit_show_types_second_ids and show.cate_second.id not in limit_show_types_second_ids):
        #     can_use = True
        # if not can_use and not (limit_shows_nos and show.no not in limit_shows_nos):
        #     can_use = True
        can_use = self.check_type_use(amount, multiply)
        if not can_use:
            return can_use
        shows = self.coupon.shows.all()
        # show_types = self.coupon.limit_show_types_second.all()
        # if show_types and not show_types.filter(id=show.cate_second.id).exists():
        #     can_use = False
        if shows and not shows.filter(id=show.id).exists():
            can_use = False
        return can_use

    def set_use(self, order):
        self.order = order
        self.status = self.STATUS_USE
        self.used_time = timezone.now()
        self.save(update_fields=['used_time', 'order', 'status'])

    def cancel_use(self):
        self.order = None
        self.status = self.STATUS_DEFAULT if self.expire_time > timezone.now() else self.STATUS_EXPIRE
        self.used_time = None
        self.save(update_fields=['used_time', 'order', 'status'])

    def get_promote_amount(self, actual_amount, multiply):
        can_use = self.check_type_use(actual_amount, multiply)
        promote_amount = Decimal('0')
        if can_use:
            if self.coupon_type in Coupon.discount_type():
                promote_amount = actual_amount * (100 - self.discount) / 100
            elif self.coupon_type == Coupon.TYPE_MONEY_OFF:
                promote_amount = self.amount
        return quantize(Decimal(promote_amount), 2) if promote_amount > 0 else promote_amount


class UserCouponImport(models.Model):
    file = models.FileField(u'文件', upload_to=f'{FILE_FIELD_PREFIX}/coupon',
                            validators=[FileExtensionValidator(allowed_extensions=['xlsx'])], help_text='只支持xlsx')
    remark = models.CharField('备注说明', max_length=100, null=True, blank=True)
    ST_NEED = 1
    ST_DOING = 2
    ST_FINISH = 3
    ST_FAIL = 4
    status = models.PositiveIntegerField('状态', default=1,
                                         choices=[(ST_NEED, '未执行'), (ST_DOING, '执行中'), (ST_FINISH, '已完成'),
                                                  (ST_FAIL, '异常')])
    create_at = models.DateTimeField(u'创建时间', auto_now_add=True)
    exec_at = models.DateTimeField(u'执行发放时间', null=True, blank=True)
    user = models.ForeignKey(settings.AUTH_USER_MODEL, verbose_name='操作人员', on_delete=models.CASCADE)
    total_num = models.IntegerField('总行数', default=0)
    success_num = models.IntegerField('成功行数', default=0)
    fail_num = models.IntegerField('失败行数', default=0)
    fail_msg = models.TextField('失败行信息', null=True, blank=True)

    class Meta:
        verbose_name_plural = verbose_name = u'批量发放记录'
        ordering = ['-pk']

    def __str__(self):
        return self.file.name

    @classmethod
    def do_coupon_import_task(cls, pk):
        close_old_connections()
        inst = cls.objects.get(pk=pk)
        inst.do_import()
        log.info(f'批量发放记录导入完成,{pk}')

    def do_import(self):
        def format_str(content: str):
            return content.replace('_x000D_', ' ')

        self.exec_at = timezone.now()
        self.status = self.ST_DOING
        self.save(update_fields=['status', 'exec_at'])
        fail_num = 0
        success_num = 0
        fail_msg = ''
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.file.path)
            # wb = load_workbook('d:/11.xlsx')
            ws = wb.active
            update_list = []
            coupon_dict = dict()
            user_dict = dict()
            for line, row in enumerate(ws.rows, 1):
                if line == 1:
                    title_list = [t.value for t in row]
                    mobile_index = title_list.index('手机号')
                    coupon_index = title_list.index('消费券编号')
                    num_index = title_list.index('发放数量')
                    continue
                try:
                    dd = list(map(lambda cell: cell.value, row))
                    mobile = format_str(str(dd[mobile_index]))
                    from common.utils import validate_mobile
                    vm = validate_mobile(mobile)
                    if not vm:
                        raise ValueError('手机号格式错误')
                    if not user_dict.get(mobile):
                        from mall.models import User
                        user = User.objects.filter(mobile=mobile).first()
                        user_dict[mobile] = user
                    else:
                        user = user_dict[mobile]
                    coupon_no = format_str(str(dd[coupon_index]))
                    num = int(dd[num_index])
                    if not coupon_dict.get(coupon_no):
                        coupon = Coupon.objects.get(no=coupon_no)
                        coupon_dict[coupon_no] = coupon
                    else:
                        coupon = coupon_dict[coupon_no]
                    if not user:
                        obj = UserCouponCacheRecord.get_obj(self.id, mobile, coupon.id)
                        obj.num = num
                        update_list.append(obj)
                    else:
                        for i in list(range(0, num)):
                            UserCouponRecord.create_record(user.id, coupon)
                    success_num += 1
                except Exception as e:
                    fail_num += 1
                    fail_msg = fail_msg + '{}行,'.format(line)
                    print(e)
            if update_list:
                UserCouponCacheRecord.objects.bulk_update(update_list, ['num'])
        except Exception as e:
            log.error(e)
            fail_msg = str(e)
        self.status = self.ST_FINISH if fail_num == 0 and not fail_msg else self.ST_FAIL
        self.fail_num = fail_num
        self.success_num = success_num
        self.fail_msg = fail_msg
        self.save(update_fields=['status', 'fail_num', 'success_num', 'fail_msg'])


class UserCouponCacheRecord(models.Model):
    record = models.ForeignKey(UserCouponImport, verbose_name='批量发放记录', on_delete=models.SET_NULL, null=True)
    mobile = models.CharField('手机号', max_length=20, db_index=True)
    coupon = models.ForeignKey(Coupon, verbose_name='优惠券', on_delete=models.CASCADE)
    num = models.PositiveSmallIntegerField('发放数量', default=0)

    class Meta:
        verbose_name_plural = verbose_name = '未领取记录'
        ordering = ['-pk']

    @classmethod
    def get_obj(cls, record_id: int, mobile: str, coupon_id: int):
        obj, _ = cls.objects.get_or_create(record_id=record_id, mobile=mobile, coupon_id=coupon_id)
        return obj

    @classmethod
    def do_bind_user_task(cls, mobile: str, user_id: int):
        qs = cls.objects.filter(mobile=mobile)
        for obj in qs:
            coupon = obj.coupon
            for i in list(range(0, obj.num)):
                UserCouponRecord.create_record(user_id, coupon)
        qs.delete()


class CouponActivity(models.Model):
    no = models.CharField('编号', max_length=64, unique=True, db_index=True, null=True, default=get_short_no)
    title = models.CharField('活动名称', max_length=50)
    coupons = models.ManyToManyField(Coupon, verbose_name='关联消费券')
    url_link = models.CharField('领取链接', max_length=100, null=True, blank=True, editable=False, help_text='小程序加密URLLink')
    share_img = models.ImageField('分享图片', upload_to=f'{IMAGE_FIELD_PREFIX}/coupon/act',
                                  validators=[validate_image_file_extension])
    ST_ON = 1
    ST_OFF = 2
    status = models.PositiveIntegerField('活动状态', default=ST_ON, choices=[(ST_ON, '上架'), (ST_OFF, '下架')])
    create_at = models.DateTimeField('创建时间', auto_now_add=True)
    update_at = models.DateTimeField('更新时间', null=True, blank=True)

    class Meta:
        verbose_name_plural = verbose_name = '专属活动'
        ordering = ['-pk']

    def coupons_desc(self):
        return ','.join(list(self.coupons.all().values_list('name', flat=True)))

    def get_url_link(self, is_refresh=False):
        if is_refresh or not self.url_link:
            try:
                from mp.wechat_client import get_wxa_client
                wxa = get_wxa_client()
                # url = 'pages/index/index'
                url = 'pages/pagesKageB/couponActivity/couponActivity'
                data = wxa.generate_urllink(url, f'act_no={self.no}')
                if data['errcode'] == 0:
                    url = data['url_link']
                    self.url_link = url
                    self.save(update_fields=['url_link'])
            except Exception as e:
                log.error(e)
        return self.url_link


class CouponReceipt(ReceiptAbstract):
    user = models.ForeignKey(settings.AUTH_USER_MODEL, verbose_name=u'用户', null=True, on_delete=models.SET_NULL)
    BIZ_ACT = 1
    BIZ_CHOICES = [(BIZ_ACT, '消费卷订单')]
    biz = models.SmallIntegerField('业务类型', default=BIZ_ACT, choices=BIZ_CHOICES)
    attachment = models.TextField('附加数据', null=True, blank=True, help_text='请勿修改此字段')
    wx_pay_config = models.ForeignKey(WeiXinPayConfig, verbose_name='微信支付', blank=True, null=True,
                                      on_delete=models.SET_NULL)

    class Meta:
        ordering = ['-pk']
        verbose_name = verbose_name_plural = '消费卷支付记录'

    def __str__(self):
        return f"{str(self.user)} - {self.amount}元"

    @classmethod
    def create_record(cls, amount, user, pay_type, biz, wx_pay_config=None):
        return cls.objects.create(amount=amount, user=user, pay_type=pay_type, biz=biz, wx_pay_config=wx_pay_config)

    def get_pay_order_info(self):
        return notify_url

    @property
    def paid(self):
        return self.status == self.STATUS_FINISHED

    def get_notify_url(self):
        return dict(body='消费卷支付记录订单', user_id=self.get_user_id())

    def get_user_id(self):
        if self.pay_type == self.PAY_WeiXin_LP:
            return self.user.lp_openid
        else:
            log.error('unsupported pay_type: %s' % self.pay_type)
            raise CustomAPIException('不支持的支付类型')

    def update_info(self):
        from mall.pay_service import get_mp_pay_client
        from wechatpy import WeChatPayException
        c = get_mp_pay_client(self.pay_type, self.wx_pay_config)
        try:
            res = c.query_order(self)
            if res:
                self.transaction_id = res.get('transaction_id')
                self.save(update_fields=['transaction_id'])
        except WeChatPayException as e:
            log.error(e)

    @classmethod
    def available_pay_types(self):
        return [self.PAY_WeiXin_LP]

    def biz_paid(self):
        """
        付款成功时,根据业务类型决定处理方法
        :return:
        """
        if self.biz == self.BIZ_ACT:
            self.coupon_receipt.set_paid()
        else:
            log.fatal('unkonw biz: %s, of receipt: %s' % (self.biz, self.pk))

    def set_paid(self, pay_type=ReceiptAbstract.PAY_WeiXin_LP, transaction_id=None):
        if self.paid:
            return
        self.status = self.STATUS_FINISHED
        self.pay_type = pay_type
        self.transaction_id = transaction_id if transaction_id else self.transaction_id
        self.save(update_fields=['status', 'transaction_id'])
        self.biz_paid()


class CouponOrder(models.Model):
    ST_DEFAULT = 1
    ST_PAID = 2
    ST_CANCEL = 3
    ST_REFUNDING = 4
    ST_REFUNDED = 5
    PAYMENT_STATUS = (
        (ST_DEFAULT, '未付款'),
        (ST_PAID, '已支付'),
        (ST_CANCEL, '已取消'),
        (ST_REFUNDING, '退款中'),
        (ST_REFUNDED, '已退款'),
    )
    order_no = models.CharField(u'订单号', max_length=128, unique=True, default=coupon_order_no, db_index=True)
    user = models.ForeignKey(settings.AUTH_USER_MODEL, on_delete=models.SET_NULL, verbose_name='用户', null=True)
    mobile = models.CharField('手机号', max_length=20)
    coupon = models.ForeignKey(Coupon, on_delete=models.CASCADE, related_name='cp_order', verbose_name='消费卷')
    coupon_name = models.CharField('消费卷名称', max_length=128, null=True)
    status = models.PositiveSmallIntegerField('状态', choices=PAYMENT_STATUS, default=ST_DEFAULT)
    multiply = models.PositiveSmallIntegerField('数量')
    amount = models.DecimalField('实付金额', max_digits=10, decimal_places=2)
    refund_amount = models.DecimalField('已退款金额', max_digits=10, decimal_places=2, default=0)
    receipt = models.OneToOneField(CouponReceipt, verbose_name='支付记录', on_delete=models.CASCADE,
                                   related_name='coupon_receipt')
    wx_pay_config = models.ForeignKey(WeiXinPayConfig, verbose_name='微信支付', blank=True, null=True,
                                      on_delete=models.SET_NULL)
    pay_type = models.SmallIntegerField('付款类型', choices=CouponReceipt.PAY_CHOICES, default=CouponReceipt.PAY_NOT_SET)
    create_at = models.DateTimeField('创建时间', auto_now_add=True)
    pay_at = models.DateTimeField('支付时间', null=True, blank=True)
    transaction_id = models.CharField('交易号', max_length=100, null=True, blank=True)
    snapshot = models.TextField('消费卷快照', null=True, blank=True, help_text='下单时保存的快照', editable=False, max_length=2048)

    class Meta:
        verbose_name = verbose_name_plural = '购买订单'
        ordering = ['-pk']

    def __str__(self):
        return self.order_no

    @classmethod
    def can_refund_status(cls):
        return [cls.ST_PAID]

    @classmethod
    def get_snapshot(cls, coupon: Coupon):
        return None
        data = dict(coupon=dict(amount=float(coupon.pay_amount)))
        return json.dumps(data)

    def push_refund(self, amount):
        self.refund_amount += amount
        self.status = self.ST_REFUNDING
        self.save(update_fields=['refund_amount', 'status'])
        UserCouponRecord.objects.filter(buy_order=self).update(status=UserCouponRecord.STATUS_INVALID)

    def set_refunded(self):
        self.refund_at = timezone.now()
        self.status = self.ST_REFUNDED
        self.save(update_fields=['status', 'refund_at'])

    def set_paid(self):
        if self.status in [self.ST_DEFAULT, self.ST_CANCEL]:
            self.status = self.ST_PAID
            self.pay_at = timezone.now()
            self.save(update_fields=['status', 'pay_at'])
            self.send_coupon()

    def send_coupon(self):
        try:
            UserCouponRecord.create_record(self.user.id, self.coupon, buy_order=self)
        except Exception as e:
            log.error('购买消费券，获取失败')
            log.error(e)

    def set_cancel(self):
        if self.status == self.ST_DEFAULT:
            receipt = self.receipt
            if receipt.pay_type == receipt.PAY_WeiXin_LP:
                receipt.query_status(self.order_no)
            if receipt.paid:
                receipt.biz_paid()
                return False, '取消失败订单已付款'
            self.status = self.ST_CANCEL
            self.cancel_at = timezone.now()
            self.save(update_fields=['status', 'cancel_at'])
            # 归还库存
            coupon = self.coupon
            coupon.coupon_change_stock(self.multiply)
            # 减去购买数量
            UserCouponRecord.set_user_obtain_cache(coupon.no, self.user.id, -1)
        return True, ''

    def do_refund(self, op_user=None, refund_reason=None):
        refund_amount = self.amount - self.refund_amount
        st, msg, obj = CouponOrderRefund.create_record(self, amount=refund_amount, refund_reason=refund_reason)
        if obj:
            st, msg = obj.set_confirm(op_user)
        return st, msg


class CommonRefundAbstract(models.Model):
    STATUS_DEFAULT = 1
    STATUS_PAYING = 2
    STATUS_PAY_FAILED = 3
    STATUS_FINISHED = 4
    STATUS_CANCELED = 5
    STATUS_CHOICES = (
        (STATUS_DEFAULT, '待退款'), (STATUS_PAYING, '退款支付中'), (STATUS_PAY_FAILED, '退款支付失败'), (STATUS_FINISHED, '已完成'),
        (STATUS_CANCELED, '已拒绝'))
    status = models.IntegerField('状态', choices=STATUS_CHOICES, default=STATUS_DEFAULT)
    user = models.ForeignKey(settings.AUTH_USER_MODEL, verbose_name='用户', on_delete=models.CASCADE)
    refund_amount = models.DecimalField(u'退款金额', max_digits=13, decimal_places=2, default=0)
    refund_reason = models.CharField('退款原因', max_length=200, null=True, blank=True)
    amount = models.DecimalField(u'实退金额', max_digits=13, decimal_places=2, default=0)
    error_msg = models.CharField('退款返回信息', max_length=1000, null=True, blank=True)
    transaction_id = models.CharField('交易号', max_length=100, null=True, blank=True)
    refund_id = models.CharField('退款方退款单号', max_length=32, null=True, blank=True)
    return_code = models.CharField('微信通信结果', max_length=20, null=True, blank=True)
    result_code = models.CharField('微信返回结果', max_length=20, null=True, blank=True)
    create_at = models.DateTimeField('创建时间', auto_now_add=True)
    confirm_at = models.DateTimeField('确认时间', null=True, blank=True)
    finish_at = models.DateTimeField('完成时间', null=True, blank=True)
    op_user = models.ForeignKey(settings.AUTH_USER_MODEL, verbose_name='操作退款用户', on_delete=models.CASCADE, null=True,
                                blank=True, related_name='+')

    class Meta:
        abstract = True

    def wx_refund(self, *args):
        raise NotImplementedError()

    def get_refund_notify_url(self):
        return refund_notify_url

    @classmethod
    def create_record(cls, order, amount=Decimal('0'), refund_reason=None):
        receipt = order.receipt
        refund_amount = amount if amount > 0 else order.amount
        if receipt.status == CouponReceipt.STATUS_FINISHED and receipt.transaction_id:
            inst = cls.objects.create(user=order.user, order=order, refund_amount=refund_amount,
                                      order_no=order.order_no,
                                      refund_reason=refund_reason, transaction_id=receipt.transaction_id)
            order.push_refund(refund_amount)
            return True, '', inst
        else:
            return False, '该订单未付款不能退款', None

    @atomic
    def set_confirm(self, op_user=None):
        try:
            st = self.wx_refund()
            msg = self.error_msg
            if st:
                self.status = self.STATUS_PAYING
                self.confirm_at = timezone.now()
                self.op_user = op_user
                self.save(update_fields=['status', 'confirm_at', 'op_user'])
            return st, msg
        except Exception as e:
            self.status = self.STATUS_PAY_FAILED
            self.confirm_at = timezone.now()
            self.error_msg = str(e)
            self.op_user = op_user
            self.save(update_fields=['status', 'confirm_at', 'op_user'])
            log.error(e)
            return False, str(e)

    @classmethod
    def can_confirm_status(cls):
        return [cls.STATUS_DEFAULT, cls.STATUS_PAY_FAILED]


class CouponOrderRefund(CommonRefundAbstract):
    order = models.ForeignKey(CouponOrder, related_name='coupon_refund', verbose_name='退款订单', on_delete=models.CASCADE)
    order_no = models.CharField(u'订单号', max_length=128)
    out_refund_no = models.CharField(u'退款单号', max_length=64, default=coupon_refund_no, unique=True, db_index=True)

    class Meta:
        verbose_name_plural = verbose_name = '消费卷退款记录'
        ordering = ['-pk']

    @classmethod
    def user_refund(cls, order, op_user=None, reason=None):
        st, msg, inst = cls.create_record(order, reason)
        if st:
            st, msg = inst.set_confirm(op_user)
        if not st:
            raise CustomAPIException(msg)

    def wx_refund(self, request):
        receipt = self.order.receipt
        if not receipt.transaction_id:
            receipt.update_info()
        from mall.pay_service import get_mp_pay_client
        mp_pay_client = get_mp_pay_client(receipt.pay_type, receipt.wx_pay_config)
        refund_notify_url = request.build_absolute_uri(self.get_refund_notify_url())
        result = mp_pay_client.new_refund(self, notify_url=refund_notify_url)
        self.return_code = result.get('return_code')
        self.result_code = result.get('result_code')
        self.error_msg = '{}, {}'.format(result.get('return_msg'), result.get('err_code_des'))
        self.refund_id = result.get('refund_id', None)
        self.save(update_fields=['return_code', 'result_code', 'error_msg', 'refund_id'])
        if self.return_code == self.result_code == 'SUCCESS':
            return True
        return False

    def set_cancel(self, op_user):
        self.status = self.STATUS_CANCELED
        self.confirm_at = timezone.now()
        self.op_user = op_user
        self.save(update_fields=['status', 'confirm_at', 'op_user'])
        self.return_order_status()
        UserCouponRecord.objects.filter(buy_order=self).update(status=UserCouponRecord.STATUS_DEFAULT)

    def return_order_status(self):
        if self.order.refund_amount > 0:
            self.order.status = self.order.ST_PAID
            self.order.refund_amount -= self.refund_amount
            self.order.save(update_fields=['status', 'refund_amount'])

    def set_fail(self, msg=None):
        self.status = self.STATUS_PAY_FAILED
        self.finish_at = timezone.now()
        self.error_msg = msg
        self.save(update_fields=['status', 'error_msg', 'finish_at'])
        # self.return_order_status()

    def set_finished(self, amount):
        """
        设置为完成:
        1.在退款支付通知回调中
        2.后台手动
        :return:
        """
        from decimal import Decimal
        self.status = self.STATUS_FINISHED
        self.finish_at = timezone.now()
        self.amount = Decimal(amount) / 100
        self.save(update_fields=['status', 'finish_at', 'amount'])
        self.order.set_refunded()
