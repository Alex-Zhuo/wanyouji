# coding=utf-8
from django.contrib import admin
from django.http import HttpResponse
from django.utils import timezone
from django.utils.safestring import mark_safe
from django.contrib import messages
from openpyxl import Workbook
from simpleui.admin import AjaxAdmin
from django.http import JsonResponse
from blind_box.models import (
    Prize, BlindBox, WheelWinningRecord, BlindBoxWinningRecord, WinningRecordShipmentReceipt,
    WheelActivity, WheelSection, LotteryPurchaseRecord,
    PrizeDetailImage, BlindBoxCarouselImage, BlindBoxDetailImage, UserLotteryTimes, BlindBasic, UserLotteryRecord,
    BlindBoxOrder, WinningRecordAbstract
)
from dj import technology_admin
from dj_ext.permissions import RemoveDeleteModelAdmin, OnlyViewAdmin, ChangeAndViewAdmin, RemoveDeleteStackedInline
import logging
from caches import get_redis_name, run_with_lock
from decimal import Decimal
from blind_box.stock_updater import prsc, bdbc

from dj_ext import AdminException

log = logging.getLogger(__name__)


class BlindBasicAdmin(RemoveDeleteModelAdmin):
    def changelist_view(self, request, extra_context=None):
        obj = BlindBasic.get()
        if obj:
            return self.change_view(request, str(obj.id))
        return self.add_view(request, extra_context={'show_save_and_add_another': False})


# ========== 奖品相关 ==========
def set_on(modeladmin, request, queryset):
    qs = queryset.filter(status=Prize.STATUS_OFF)
    for obj in qs:
        obj.prize_redis_stock()
    qs.update(status=Prize.STATUS_ON)
    messages.success(request, '执行成功')


set_on.short_description = '批量上架'


def set_off(modeladmin, request, queryset):
    qs = queryset.filter(status=Prize.STATUS_ON)
    # for obj in qs:
    #     obj.prize_del_redis_stock()
    qs.update(status=Prize.STATUS_OFF)
    messages.success(request, '执行成功')


set_off.short_description = '批量下架'


class PrizeDetailImageInline(admin.TabularInline):
    model = PrizeDetailImage
    extra = 0


class PrizeAdmin(RemoveDeleteModelAdmin, AjaxAdmin):
    list_display = ['no', 'title', 'display_order', 'status', 'source_type', 'rare_type', 'amount', 'stock',
                    'weight', 'create_at']
    list_filter = ['status', 'source_type', 'rare_type']
    list_editable = ['display_order']
    search_fields = ['no', 'title']
    actions = [set_on, set_off, 'add_stock']
    inlines = [PrizeDetailImageInline]

    def get_readonly_fields(self, request, obj=None):
        if obj:
            return ['no', 'stock']
        else:
            return ['no']

    def response_post_save_add(self, request, obj):
        if obj.status == Prize.STATUS_ON:
            obj.prize_redis_stock()
        return super(PrizeAdmin, self).response_post_save_add(request, obj)

    def add_stock(self, request, queryset):
        obj = queryset.first()
        key = get_redis_name('prize_stock_lock{}'.format(obj.id))
        with run_with_lock(key, 3) as got:
            if not got:
                return JsonResponse(data={
                    'status': 'error',
                    'msg': '请勿点击多次！'
                })
        post = request.POST
        if not post.get('_selected'):
            return JsonResponse(data={
                'status': 'error',
                'msg': '请先勾选一条记录！'
            })
        else:
            if queryset.count() > 1:
                return JsonResponse(data={
                    'status': 'error',
                    'msg': '该功能功能只能单选！'
                })
            num = int(post.get('add_stock'))
            st = obj.prize_change_stock(int(num))
            if st:
                prsc.instant_persist(obj.id)
                return JsonResponse(data={
                    'status': 'success',
                    'msg': '执行成功！'
                })
            return JsonResponse(data={
                'status': 'error',
                'msg': '执行成功，请稍后再试！'
            })

    add_stock.short_description = '增加库存数量'
    add_stock.type = 'success'
    add_stock.icon = 'el-icon-s-promotion'
    # 指定为弹出层，这个参数最关键
    add_stock.layer = {
        # 弹出层中的输入框配置
        # 这里指定对话框的标题
        'title': '增加库存数量',
        # 提示信息
        'tips': '输入增加的库存数量',
        # 确认按钮显示文本
        'confirm_button': '确认提交',
        # 取消按钮显示文本
        'cancel_button': '取消',
        # 弹出层对话框的宽度，默认50%
        'width': '50%',
        # 表单中 label的宽度，对应element-ui的 label-width，默认80px
        'labelWidth': "100px",
        'params': [
            {
                'require': True,
                'type': 'number',
                'width': '300px',
                'key': 'add_stock',
                'label': '数量',
                'value': 0
            },
        ]
    }


# ========== 盲盒相关 ==========
class BlindBoxCarouselImageInline(admin.TabularInline):
    model = BlindBoxCarouselImage
    extra = 0


class BlindBoxDetailImageInline(admin.TabularInline):
    model = BlindBoxDetailImage
    extra = 0


def blind_set_on(modeladmin, request, queryset):
    qs = queryset.filter(status=BlindBox.STATUS_OFF)
    for obj in qs:
        obj.blind_box_redis_stock()
    qs.update(status=BlindBox.STATUS_ON)
    messages.success(request, '执行成功')


blind_set_on.short_description = '上架'


def blind_set_off(modeladmin, request, queryset):
    qs = queryset.filter(status=BlindBox.STATUS_ON)
    # for obj in qs:
    #     obj.blind_box_del_redis_stock()
    qs.update(status=BlindBox.STATUS_OFF)
    messages.success(request, '执行成功')


blind_set_off.short_description = '下架'


class BlindBoxAdmin(RemoveDeleteModelAdmin, AjaxAdmin):
    list_display = ['no', 'title', 'display_order', 'status', 'type', 'grids_num', 'price', 'original_price', 'stock',
                    'rare_weight_multiple', 'hidden_weight_multiple', 'create_at']
    list_filter = ['status', 'type', 'grids_num']
    list_editable = ['display_order']
    search_fields = ['no', 'title']
    inlines = [BlindBoxCarouselImageInline, BlindBoxDetailImageInline]
    actions = [blind_set_on, blind_set_off, 'add_stock']

    def get_readonly_fields(self, request, obj=None):
        if obj:
            return ['no', 'stock']
        else:
            return ['no']

    def response_post_save_add(self, request, obj):
        if obj.status == BlindBox.STATUS_ON:
            obj.blind_box_redis_stock()
        return super(BlindBoxAdmin, self).response_post_save_add(request, obj)

    def add_stock(self, request, queryset):
        obj = queryset.first()
        key = get_redis_name('blind_stock_lock{}'.format(obj.id))
        with run_with_lock(key, 3) as got:
            if not got:
                return JsonResponse(data={
                    'status': 'error',
                    'msg': '请勿点击多次！'
                })
        post = request.POST
        if not post.get('_selected'):
            return JsonResponse(data={
                'status': 'error',
                'msg': '请先勾选一条记录！'
            })
        else:
            if queryset.count() > 1:
                return JsonResponse(data={
                    'status': 'error',
                    'msg': '该功能功能只能单选！'
                })
            num = int(post.get('add_stock'))
            st = obj.blind_box_change_stock(int(num))
            if st:
                bdbc.instant_persist(obj.id)
                return JsonResponse(data={
                    'status': 'success',
                    'msg': '执行成功！'
                })
            return JsonResponse(data={
                'status': 'error',
                'msg': '执行成功，请稍后再试！'
            })

    add_stock.short_description = '增加库存数量'
    add_stock.type = 'success'
    add_stock.icon = 'el-icon-s-promotion'
    # 指定为弹出层，这个参数最关键
    add_stock.layer = {
        # 弹出层中的输入框配置
        # 这里指定对话框的标题
        'title': '增加库存数量',
        # 提示信息
        'tips': '输入增加的库存数量',
        # 确认按钮显示文本
        'confirm_button': '确认提交',
        # 取消按钮显示文本
        'cancel_button': '取消',
        # 弹出层对话框的宽度，默认50%
        'width': '50%',
        # 表单中 label的宽度，对应element-ui的 label-width，默认80px
        'labelWidth': "100px",
        'params': [
            {
                'require': True,
                'type': 'number',
                'width': '300px',
                'key': 'add_stock',
                'label': '数量',
                'value': 0
            },
        ]
    }


def export_blind_box_order(modeladmin, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format(
        timezone.now().strftime('%Y%m%d%H%M'))
    wb = Workbook()
    ws = wb.active
    ws.append(['订单号', '用户', '手机号', '盲盒', '实付金额', '状态', '创建时间', '付款时间', '微信支付单号', '已退款金额', '退款时间'])
    for record in queryset:
        create_at = record.create_at.strftime('%Y-%m-%d %H:%M')
        pay_at = record.pay_at.strftime('%Y-%m-%d %H:%M') if record.pay_at else None
        refund_at = record.refund_at.strftime('%Y-%m-%d %H:%M') if record.refund_at else None
        data = [record.order_no, str(record.user), record.mobile, record.blind_box.title if record.blind_box else None,
                record.amount,
                record.get_status_display(), create_at, pay_at, record.transaction_id, record.refund_amount, refund_at]
        ws.append(data)
    wb.save(response)
    return response


export_blind_box_order.short_description = '导出选中记录'


def blind_box_set_paid(modeladmin, request, queryset):
    queryset = queryset.filter(status=BlindBoxOrder.ST_DEFAULT)
    for order in queryset:
        order.set_paid()
    messages.success(request, '执行成功')


blind_box_set_paid.short_description = u'后台付款'


class BlindBoxOrderAdmin(AjaxAdmin, OnlyViewAdmin):
    list_display = ['order_no', 'user', 'mobile', 'blind_box', 'amount', 'refund_amount', 'status', 'create_at',
                    'pay_at',
                    'transaction_id']
    search_fields = ['=transaction_id', '=mobile', '=order_no']
    list_filter = ['status', 'pay_at', 'blind_box']
    readonly_fields = ['user', 'receipt', 'blind_box', 'order_no']
    actions = ['blind_box_refund', export_blind_box_order, blind_box_set_paid]

    def blind_box_refund(self, request, queryset):
        qs = queryset.filter(status__in=BlindBoxOrder.can_refund_status())
        if not qs:
            return JsonResponse(data={
                'status': 'error',
                'msg': '只有已付款可以退款！'
            })
        if qs.count() > 1:
            return JsonResponse(data={
                'status': 'error',
                'msg': '每次最多执行一条记录！'
            })
        order = queryset.first()
        key = get_redis_name('blind_box_refund{}'.format(order.id))
        with run_with_lock(key, 3) as got:
            if not got:
                return JsonResponse(data={
                    'status': 'error',
                    'msg': '请勿点击多次！'
                })
        post = request.POST
        refund_reason = post.get('reason')
        amount = post.get('amount')
        amount = Decimal(amount)
        refund_amount = order.amount - order.refund_amount
        if amount <= 0:
            return JsonResponse(data={
                'status': 'error',
                'msg': '退款金额需要大于0！'
            })
        elif amount > refund_amount:
            return JsonResponse(data={
                'status': 'error',
                'msg': '退款金额不能大于实付金额！'
            })
        st, msg = order.do_refund(amount, refund_reason)
        if not st:
            return JsonResponse(data={
                'status': 'error',
                'msg': msg
            })
        return JsonResponse(data={
            'status': 'success',
            'msg': '操作成功！'
        })

    blind_box_refund.short_description = '申请退款'
    blind_box_refund.type = 'success'
    blind_box_refund.icon = 'el-icon-s-promotion'
    # 指定为弹出层，这个参数最关键
    blind_box_refund.layer = {
        # 弹出层中的输入框配置
        # 这里指定对话框的标题
        'title': '申请退款',
        # 提示信息
        'tips': '',
        # 确认按钮显示文本
        'confirm_button': '确认提交',
        # 取消按钮显示文本
        'cancel_button': '取消',
        # 弹出层对话框的宽度，默认50%
        'width': '50%',
        # 表单中 label的宽度，对应element-ui的 label-width，默认80px
        'labelWidth': "100px",
        'params': [
            {
                'require': True,
                'type': 'number',
                'width': '300px',
                'key': 'amount',
                'label': '退款金额',
                'value': 0
            },
            {
                # 这里的type 对应el-input的原生input属性，默认为input
                'require': True,
                'type': 'input',
                # key 对应post参数中的key
                'key': 'reason',
                # 显示的文本
                'label': '退款原因',
                'width': '70%',
                # 表单中 label的宽度，对应element-ui的 label-width，默认80px
                'labelWidth': "120px",
            }
        ]
    }


# ========== 中奖记录相关 ==========
def export_shipment_list(modeladmin, request, queryset):
    """批量导出发货单"""
    wb = Workbook()
    ws = wb.active
    ws.title = '发货单'
    # 表头
    headers = ['中奖序号', '中奖用户', '手机号', '奖品', '奖品类型', '收货地址', '收货人', '收货联系人电话', '快递公司', '快递单号', '备注']
    ws.append(headers)
    # 数据行
    for record in queryset.filter(status=WinningRecordAbstract.ST_PENDING_SHIP):
        row = [
            record.no,
            str(record.user) if record.user else '',
            record.mobile,
            record.prize.title if record.prize else '',
            record.get_source_type_display(),
            record.express_address,
            record.express_user_name,
            record.express_phone,
            record.express_company_name or '',
            record.express_no or '',
            record.remark
        ]
        ws.append(row)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="发货单_{timezone.now().strftime("%Y%m%d%H%M%S")}.xlsx"'
    wb.save(response)
    return response


export_shipment_list.short_description = '批量导出发货单'


def set_completed(modeladmin, request, queryset):
    qs = queryset.filter(
        status__in=[WinningRecordAbstract.ST_PENDING_RECEIVE, WinningRecordAbstract.ST_PENDING_RECEIPT])
    for obj in qs:
        obj.set_completed()
    messages.success(request, '执行成功')


set_completed.short_description = '设为已完成'


class WinningRecordAbstractAdmin(AjaxAdmin, ChangeAndViewAdmin):
    list_display = [
        'no', 'user', 'mobile', 'prize', 'status', 'source_type',
        'express_company_name', 'express_no', 'winning_at', 'receive_at', 'ship_at', 'complete_at']
    list_filter = ['status', 'source_type', 'winning_at']
    search_fields = ['=mobile', '=no']
    readonly_fields = ['no', 'user', 'prize', 'winning_at', 'receive_at', 'ship_at', 'complete_at']
    actions = [export_shipment_list, 'shipping_good', set_completed]

    # def source_type_display(self, obj):
    #     return obj.get_source_type_display()
    #
    # source_type_display.short_description = '奖品类型'
    #
    # def status_display(self, obj):
    #     return obj.get_status_display()
    #
    # status_display.short_description = '状态'

    def shipping_good(self, request, queryset):
        qs = queryset.filter(status__in=[WinningRecordAbstract.ST_PENDING_SHIP, WinningRecordAbstract.ST_COMPLETED])
        if not qs:
            return JsonResponse(data={
                'status': 'error',
                'msg': '只有待发货状态才能执行！'
            })
        if qs.count() > 1:
            return JsonResponse(data={
                'status': 'error',
                'msg': '每次最多执行一条记录！'
            })
        order = queryset.first()
        key = get_redis_name('win_o{}'.format(order.id))
        with run_with_lock(key, 3) as got:
            if not got:
                return JsonResponse(data={
                    'status': 'error',
                    'msg': '请勿点击多次！'
                })
        post = request.POST
        express_no = post.get('express_no')
        express_company_code = post.get('express_company_code')
        express_company_name = post.get('express_company_name')
        order.set_shipped(express_no, express_company_code, express_company_name)
        return JsonResponse(data={
            'status': 'success',
            'msg': '操作成功！'
        })

    shipping_good.short_description = '发货'
    shipping_good.type = 'success'
    shipping_good.icon = 'el-icon-s-promotion'
    # 指定为弹出层，这个参数最关键
    shipping_good.layer = {
        # 弹出层中的输入框配置
        # 这里指定对话框的标题
        'title': '录入发货',
        # 提示信息
        'tips': '',
        # 确认按钮显示文本
        'confirm_button': '确认提交',
        # 取消按钮显示文本
        'cancel_button': '取消',
        # 弹出层对话框的宽度，默认50%
        'width': '50%',
        # 表单中 label的宽度，对应element-ui的 label-width，默认80px
        'labelWidth': "100px",
        'params': [
            {
                # 这里的type 对应el-input的原生input属性，默认为input
                'require': True,
                'type': 'input',
                # key 对应post参数中的key
                'key': 'express_no',
                # 显示的文本
                'label': '快递单号',
                'width': '70%',
                # 表单中 label的宽度，对应element-ui的 label-width，默认80px
                'labelWidth': "120px",
            },
            {
                # 这里的type 对应el-input的原生input属性，默认为input
                'require': True,
                'type': 'input',
                # key 对应post参数中的key
                'key': 'express_company_code',
                # 显示的文本
                'label': '快递公司编码',
                'width': '70%',
                # 表单中 label的宽度，对应element-ui的 label-width，默认80px
                'labelWidth': "120px",
            },
            {
                # 这里的type 对应el-input的原生input属性，默认为input
                'require': True,
                'type': 'input',
                # key 对应post参数中的key
                'key': 'express_company_name',
                # 显示的文本
                'label': '快递公司名称',
                'width': '70%',
                # 表单中 label的宽度，对应element-ui的 label-width，默认80px
                'labelWidth': "120px",
            }
        ]
    }


class BlindBoxWinningRecordAdmin(WinningRecordAbstractAdmin):
    list_display = WinningRecordAbstractAdmin.list_display + ['blind_box_order', 'blind_box_title']
    readonly_fields = WinningRecordAbstractAdmin.readonly_fields + ['blind_box_order', 'blind_box_title']
    list_filter = WinningRecordAbstractAdmin.list_filter + ['blind_box_order']
    exclude = ['blind_box']
    fieldsets = (
        ('基本信息', {
            'fields': (
                'no', 'user', 'mobile', 'prize', 'source_type', 'blind_box_title', 'blind_box_order')
        }),
        ('状态信息', {
            'fields': ('status', 'winning_at', 'receive_at', 'ship_at', 'complete_at')
        }),
        ('物流信息', {
            'fields': ('express_user_name', 'express_phone', 'express_address', 'express_no', 'express_company_code',
                       'express_company_name')
        }),
    )

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.exclude(status=BlindBoxWinningRecord.ST_UNPAID)


class WheelWinningRecordAdmin(WinningRecordAbstractAdmin):
    list_display = WinningRecordAbstractAdmin.list_display + ['lottery_record', 'wheel_name']
    readonly_fields = WinningRecordAbstractAdmin.readonly_fields + ['lottery_record', 'wheel_name']
    list_filter = WinningRecordAbstractAdmin.list_filter + ['lottery_record']
    exclude = ['wheel_activity']
    fieldsets = (
        ('基本信息', {
            'fields': (
                'no', 'user', 'mobile', 'prize', 'source_type', 'instruction', 'wheel_name', 'lottery_record')
        }),
        ('状态信息', {
            'fields': ('status', 'winning_at', 'receive_at', 'ship_at', 'complete_at')
        }),
        ('物流信息', {
            'fields': ('express_user_name', 'express_phone', 'express_address', 'express_no', 'express_company_code',
                       'express_company_name')
        }),
    )


class WinningRecordShipmentReceiptAdmin(OnlyViewAdmin):
    list_display = ['create_at', 'operator', 'receipt_file', 'remark']
    list_filter = ['create_at']


# ========== 转盘活动相关 ==========
class WheelSectionInline(RemoveDeleteStackedInline):
    model = WheelSection
    extra = 0
    autocomplete_fields = ['prize']
    exclude = ['no']


def wheel_set_on(modeladmin, request, queryset):
    obj = queryset.first()
    num = obj.sections.filter(is_enabled=True).count()
    if num < 3:
        raise AdminException('转盘片区小于3,不可上架')
    obj.status = WheelActivity.STATUS_ON
    obj.save(update_fields=['status'])
    messages.success(request, '执行成功')


wheel_set_on.short_description = '上架'


def wheel_set_off(modeladmin, request, queryset):
    obj = queryset.first()
    obj.status = WheelActivity.STATUS_OFF
    obj.save(update_fields=['status'])
    messages.success(request, '执行成功')


wheel_set_off.short_description = '下架'


class WheelActivityAdmin(RemoveDeleteModelAdmin):
    list_display = ['name', 'status', 'create_at']
    list_filter = ['status', 'create_at']
    search_fields = ['name']
    inlines = [WheelSectionInline]
    readonly_fields = ['status']
    actions = [wheel_set_on, wheel_set_off]


# ========== 抽奖次数购买记录相关 ==========
def export_lottery_order(modeladmin, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format(
        timezone.now().strftime('%Y%m%d%H%M'))
    wb = Workbook()
    ws = wb.active
    ws.append(['订单号', '用户', '手机号', '实付金额', '购买次数', '状态', '创建时间', '付款时间', '微信支付单号', '已退款金额', '退款时间'])
    for record in queryset:
        create_at = record.create_at.strftime('%Y-%m-%d %H:%M')
        pay_at = record.pay_at.strftime('%Y-%m-%d %H:%M') if record.pay_at else None
        refund_at = record.refund_at.strftime('%Y-%m-%d %H:%M') if record.refund_at else None
        data = [record.order_no, str(record.user), record.mobile,
                record.amount, record.purchase_count,
                record.get_status_display(), create_at, pay_at, record.transaction_id, record.refund_amount, refund_at]
        ws.append(data)
    wb.save(response)
    return response


export_lottery_order.short_description = '导出选中记录'


def lottery_set_paid(modeladmin, request, queryset):
    queryset = queryset.filter(status=LotteryPurchaseRecord.ST_UNPAID)
    for order in queryset:
        order.set_paid()
    messages.success(request, '执行成功')


lottery_set_paid.short_description = u'后台付款'


class LotteryPurchaseRecordAdmin(OnlyViewAdmin):
    list_display = ['order_no', 'user', 'mobile', 'multiply', 'amount', 'refund_amount', 'status', 'create_at',
                    'pay_at',
                    'transaction_id']
    list_filter = ['status', 'create_at']
    search_fields = ['=mobile', '=order_no']
    readonly_fields = ['user', 'receipt', 'order_no']
    actions = ['lottery_refund', export_lottery_order, lottery_set_paid]

    def lottery_refund(self, request, queryset):
        qs = queryset.filter(status__in=LotteryPurchaseRecord.can_refund_status())
        if not qs:
            return JsonResponse(data={
                'status': 'error',
                'msg': '只有已付款可以退款！'
            })
        if qs.count() > 1:
            return JsonResponse(data={
                'status': 'error',
                'msg': '每次最多执行一条记录！'
            })
        order = queryset.first()
        key = get_redis_name('lottery_refund{}'.format(order.id))
        with run_with_lock(key, 3) as got:
            if not got:
                return JsonResponse(data={
                    'status': 'error',
                    'msg': '请勿点击多次！'
                })
        post = request.POST
        refund_reason = post.get('reason')
        amount = post.get('amount')
        amount = Decimal(amount)
        refund_amount = order.amount - order.refund_amount
        if amount <= 0:
            return JsonResponse(data={
                'status': 'error',
                'msg': '退款金额需要大于0！'
            })
        elif amount > refund_amount:
            return JsonResponse(data={
                'status': 'error',
                'msg': '退款金额不能大于实付金额！'
            })
        st, msg = order.do_refund(amount, refund_reason)
        if not st:
            return JsonResponse(data={
                'status': 'error',
                'msg': msg
            })
        return JsonResponse(data={
            'status': 'success',
            'msg': '操作成功！'
        })

    lottery_refund.short_description = '申请退款'
    lottery_refund.type = 'success'
    lottery_refund.icon = 'el-icon-s-promotion'
    # 指定为弹出层，这个参数最关键
    lottery_refund.layer = {
        # 弹出层中的输入框配置
        # 这里指定对话框的标题
        'title': '申请退款',
        # 提示信息
        'tips': '',
        # 确认按钮显示文本
        'confirm_button': '确认提交',
        # 取消按钮显示文本
        'cancel_button': '取消',
        # 弹出层对话框的宽度，默认50%
        'width': '50%',
        # 表单中 label的宽度，对应element-ui的 label-width，默认80px
        'labelWidth': "100px",
        'params': [
            {
                'require': True,
                'type': 'number',
                'width': '300px',
                'key': 'amount',
                'label': '退款金额',
                'value': 0
            },
            {
                # 这里的type 对应el-input的原生input属性，默认为input
                'require': True,
                'type': 'input',
                # key 对应post参数中的key
                'key': 'reason',
                # 显示的文本
                'label': '退款原因',
                'width': '70%',
                # 表单中 label的宽度，对应element-ui的 label-width，默认80px
                'labelWidth': "120px",
            }
        ]
    }


class UserLotteryTimesAdmin(OnlyViewAdmin):
    list_display = ['user', 'times', 'total_times', 'create_at', 'update_at']
    search_fields = ['=mobile']
    autocomplete_fields = ['user']


class UserLotteryRecordAdmin(OnlyViewAdmin):
    list_display = ['no', 'user', 'mobile', 'wheel_activity', 'is_prize', 'create_at']
    search_fields = ['=mobile']
    list_filter = ['is_prize', 'create_at']
    autocomplete_fields = ['user', 'wheel_activity']


# ========== 注册到admin ==========
admin.site.register(BlindBasic, BlindBasicAdmin)
admin.site.register(Prize, PrizeAdmin)
admin.site.register(BlindBox, BlindBoxAdmin)
admin.site.register(BlindBoxOrder, BlindBoxOrderAdmin)
admin.site.register(BlindBoxWinningRecord, BlindBoxWinningRecordAdmin)
# admin.site.register(WinningRecordShipmentReceipt, WinningRecordShipmentReceiptAdmin)
admin.site.register(WheelActivity, WheelActivityAdmin)
admin.site.register(LotteryPurchaseRecord, LotteryPurchaseRecordAdmin)
admin.site.register(UserLotteryTimes, UserLotteryTimesAdmin)
admin.site.register(UserLotteryRecord, UserLotteryRecordAdmin)
admin.site.register(WheelWinningRecord, WheelWinningRecordAdmin)

# 注册到technology_admin
technology_admin.register(BlindBasic, BlindBasicAdmin)
technology_admin.register(Prize, PrizeAdmin)
technology_admin.register(BlindBox, BlindBoxAdmin)
technology_admin.register(BlindBoxOrder, BlindBoxOrderAdmin)
technology_admin.register(BlindBoxWinningRecord, BlindBoxWinningRecordAdmin)
# technology_admin.register(WinningRecordShipmentReceipt, WinningRecordShipmentReceiptAdmin)
technology_admin.register(WheelActivity, WheelActivityAdmin)
technology_admin.register(LotteryPurchaseRecord, LotteryPurchaseRecordAdmin)
technology_admin.register(UserLotteryTimes, UserLotteryTimesAdmin)
technology_admin.register(UserLotteryRecord, UserLotteryRecordAdmin)
technology_admin.register(WheelWinningRecord, WheelWinningRecordAdmin)
