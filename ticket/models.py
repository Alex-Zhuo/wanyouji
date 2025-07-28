# coding=utf-8
from __future__ import unicode_literals
from django.db import models
from django.utils import timezone

from common.config import FILE_FIELD_PREFIX, IMAGE_FIELD_PREFIX
from express.models import Division
from mall.models import User, validate_positive_int_gen, Receipt, TheaterCardUserRecord, TheaterCardTicketLevel, \
    TheaterCard, TheaterCardCity, TheaterCardChangeRecord, TheaterCardUserDetail
from mall.utils import randomstrwithdatetime, qrcode_dir_tk
from django.db.models import Manager
import logging
from mp.models import WeiXinPayConfig, MaiZuoAccount, DouYinPayConfig
from restframework_ext.exceptions import CustomAPIException
from django.db.models import F
from django.db import close_old_connections
from datetime import timedelta
from django.db.transaction import atomic
import os
import json
from common.utils import get_config, random_new_digits, hash_ids, group_by_str, random_str, get_timestamp, sha256_str
from push import MpTemplateClient
from random import sample
from django.core.exceptions import ValidationError
from math import ceil
from decimal import Decimal
from common.utils import quantize
from django.forms.models import model_to_dict
from datetime import datetime
import pysnooper
from caches import get_pika_redis, get_redis_name
from django.core.validators import validate_image_file_extension, FileExtensionValidator
import uuid

from restframework_ext.models import UseNoAbstract

log = logging.getLogger(__name__)
tiktok_goods_url = 'pages/pagesKage/showDetail/showDetail'
tiktok_order_detail_url = 'pages/pagesKage/orderDetail/orderDetail'
SALE_ORDER_URL = 'pages/pagesKage/saleOrder/saleOrder'
tiktok_notify_url = '/api/receipts/tiktok_notify/'
tiktok_refund_notify_url = '/api/receipts/tiktok_refund_notify/'
wx_refund_notify_url = '/api/ticket_receipts/refund_notify/'

DY_STATUS = {'0': '无效值', '1': '待支付', '2': '待使用', '3': '已核销', '4': '订单关闭', '10': '待预约', '11': '核销中', '20': '退款中',
             '21': '已退款'}


def file_size(value):
    limit = 1024 * 1024
    if value.size > limit:
        raise ValidationError('图片大小不能超过1M')


def only_chinese(value):
    import re
    if re.findall('[\u4e00-\u9fa5]', value) != list(value):
        raise ValidationError('只能输入中文')


def validate_file_size(value):
    filesize = value.size
    if filesize > 1000000:
        raise ValidationError("图片必须小于1Mkb")
    else:
        return value


def randomstrwithdatetime_refund(tail_length=6):
    """
    使用当前时间(datetime)生成随机字符串,可以作为订单号
    :return:
    """
    now = timezone.now()
    return 'RF%s%s' % (now.strftime('%Y%m%d%H%M%S'), ''.join(sample(list(map(str, range(0, 10))), tail_length)))


def randomstrwithdatetime_booking(tail_length=6):
    """
    使用当前时间(datetime)生成随机字符串,可以作为订单号
    :return:
    """
    now = timezone.now()
    return 'BK%s%s' % (now.strftime('%Y%m%d%H%M%S'), ''.join(sample(list(map(str, range(0, 10))), tail_length)))


class DouYinStore(models.Model):
    name = models.CharField('店铺名称', max_length=100)
    supplier_ext_id = models.CharField('店铺ID', max_length=50)
    enable = models.BooleanField('是否使用', default=True)

    class Meta:
        verbose_name_plural = verbose_name = '抖音店铺'

    def __str__(self):
        return self.name


class ShowCategoryAbstract(models.Model):
    category_id = models.CharField(verbose_name='类目ID', null=True, max_length=10, help_text='抖音类目ID', unique=True)
    name = models.CharField(max_length=20, verbose_name='类目名称')
    enable = models.BooleanField('类目是否开放', default=True)
    goods_template = models.TextField('商品模板结构', null=True, blank=True)

    class Meta:
        abstract = True

    def __str__(self):
        return self.name


class ShowTopCategory(ShowCategoryAbstract):
    class Meta:
        verbose_name_plural = verbose_name = '抖音类目'

    @property
    def secondarys(self):
        return self.secondary

    @classmethod
    def update_or_create(cls):
        # 拉取全部
        from douyin import get_dou_yin
        inst = get_dou_yin()
        category_list = inst.get_category()
        category_ids = update_cate = cate_list = []
        for cate in category_list:
            inst = cls.objects.filter(category_id=cate['category_id']).first()
            if inst:
                inst.name = cate['name']
                inst.enable = cate['enable']
                update_cate.append(inst)
            else:
                cate_list.append(cls(category_id=cate['category_id'], name=cate['name'], enable=cate['enable']))
            category_ids.append(cate['category_id'])
        if cate_list:
            cls.objects.bulk_create(cate_list)
        if update_cate:
            cls.objects.bulk_update(update_cate, ['name', 'enable'])
        cls.update_or_create_second(category_ids)

    @classmethod
    def update_or_create_second(cls, category_ids: list):
        from douyin import get_dou_yin
        inst = get_dou_yin()
        update_second_cate = second_cate_list = []
        for cate in cls.objects.filter(category_id__in=category_ids):
            cate_second_list = inst.get_category(category_id=cate.category_id)
            for second_cate in cate_second_list:
                inst = ShowSecondaryCategory.objects.filter(category_id=second_cate['category_id']).first()
                if inst:
                    inst.name = cate['name']
                    inst.enable = cate['enable']
                    inst.superior = cate
                    update_second_cate.append(inst)
                else:
                    second_cate_list.append(
                        cls(superior=cate, category_id=cate['category_id'], name=cate['name'], enable=cate['enable']))
            if second_cate_list:
                ShowSecondaryCategory.objects.bulk_create(second_cate_list)
            if update_second_cate:
                ShowSecondaryCategory.objects.bulk_create(update_second_cate, ['name', 'enable', 'superior'])

    def get_goods_template(self):
        from douyin import get_dou_yin
        inst = get_dou_yin()
        goods_template = inst.goods_template(self.category_id)
        self.goods_template = json.dumps(goods_template)
        self.save(update_fields=['goods_template'])


class ShowSecondaryCategory(ShowCategoryAbstract):
    superior = models.ForeignKey(ShowTopCategory, verbose_name='一级分类', related_name='secondary',
                                 on_delete=models.CASCADE)

    class Meta:
        verbose_name_plural = verbose_name = '二级分类'

    def __str__(self):
        return '%s=>%s' % (self.superior.name, self.name)


class ShowThirdCategory(ShowCategoryAbstract):
    second = models.ForeignKey(ShowSecondaryCategory, verbose_name='二级分类', related_name='secondary',
                               on_delete=models.CASCADE)

    class Meta:
        verbose_name_plural = verbose_name = '三级分类'

    def __str__(self):
        return '%s=>%s' % (self.second.name, self.name)


class ShowContentCategory(models.Model):
    title = models.CharField(max_length=20, verbose_name='分类名称')
    display_order = models.PositiveSmallIntegerField('排序', default=0, help_text='从小到大排序,首页展示前5项')

    class Meta:
        verbose_name_plural = verbose_name = '节目大类'
        ordering = ['display_order']

    def __str__(self):
        return self.title

    def show_content_copy_to_pika(self):
        from caches import get_pika_redis, redis_show_content_copy_key
        with get_pika_redis() as pika:
            pika.hset(redis_show_content_copy_key, str(self.id), json.dumps(dict(id=self.id, title=self.title)))

    def get_index_data(self):
        qs = ShowProject.objects.filter(cate_id=self.id, status=ShowProject.STATUS_ON)
        data = dict(recent_num=qs.count())
        qs = qs.order_by('cate_second__display_order','display_order')[:2]
        data['shows'] = qs
        return data


class PerformerFlag(models.Model):
    title = models.CharField('名称', max_length=10)

    class Meta:
        verbose_name_plural = verbose_name = '演员标签'

    def __str__(self):
        return self.title


class ShowPerformer(models.Model):
    name = models.CharField('姓名', max_length=10)
    avatar = models.ImageField('头像', upload_to=f'{IMAGE_FIELD_PREFIX}/show/performer',
                               validators=[file_size, validate_image_file_extension])
    flag = models.ManyToManyField(PerformerFlag, verbose_name='演员标签', blank=True)
    desc = models.TextField('演员简介', max_length=500)
    focus_num = models.IntegerField('关注人数', default=0)
    is_show = models.BooleanField('是否展示', default=True)
    display_order = models.IntegerField('排序', default=0, help_text='按从小到大排列')

    class Meta:
        verbose_name_plural = verbose_name = '演员管理'
        ordering = ['display_order', 'pk']

    def __str__(self):
        return self.name

    @property
    def show_num(self):
        return ShowProject.objects.filter(performer=self, status=ShowProject.STATUS_ON).count()

    def set_focus_num(self, num=1):
        from caches import get_redis, performer_key
        redis = get_redis()
        if num:
            redis.lpush(performer_key, '{}_{}'.format(self.id, int(num)))

    @classmethod
    def update_focus_num(cls):
        from caches import get_redis, performer_key
        redis = get_redis()
        dd = redis.lindex(performer_key, 0)
        total_list = dict()
        i = 0
        while dd and i < 200:
            i += 1
            dd = redis.rpop(performer_key)
            id, num = dd.split('_')
            if total_list.get(str(id)):
                total_list[str(id)] += int(num)
            else:
                total_list[str(id)] = int(num)
        for key, val in total_list.items():
            self = cls.objects.filter(id=int(key)).first()
            if self:
                self.focus_num += val
                self.save(update_fields=['focus_num'])


class ShowPerformerBanner(models.Model):
    performer = models.ForeignKey(ShowPerformer, verbose_name='演员', on_delete=models.CASCADE)
    img = models.ImageField('海报', upload_to=f'{IMAGE_FIELD_PREFIX}/show/performer',
                            validators=[validate_file_size, validate_image_file_extension])
    is_show = models.BooleanField('是否展示', default=True)
    display_order = models.IntegerField('排序', default=0, help_text='按从小到大排列')

    class Meta:
        verbose_name_plural = verbose_name = '形象海报'
        ordering = ['-display_order']


class ShowType(models.Model):
    name = models.CharField('类型名称', max_length=20)
    SOURCE_MUSIC = 1
    SOURCE_DRAMA = 2
    SOURCE_QY = 3
    SOURCE_DANCE = 4
    SOURCE_ACROBATICS = 5
    SOURCE_EXHIBIT = 6
    SOURCE_MATCH = 7
    SOURCE_OTHER = 8
    SOURCE_CHOICES = [(SOURCE_MUSIC, '音乐'), (SOURCE_DRAMA, '戏剧'), (SOURCE_QY, '曲艺'), (SOURCE_DANCE, '舞蹈'),
                      (SOURCE_ACROBATICS, '杂技'), (SOURCE_EXHIBIT, '展览'),
                      (SOURCE_MATCH, '赛事'), (SOURCE_OTHER, '其他')]
    source_type = models.SmallIntegerField('演出行业标准分类', choices=SOURCE_CHOICES, default=SOURCE_MUSIC, null=True,
                                           blank=True, editable=False)
    category = models.ForeignKey(ShowTopCategory, verbose_name='抖音类目', related_name='cate', null=True, blank=True,
                                 on_delete=models.SET_NULL, editable=False)
    cy_cate = models.OneToOneField('caiyicloud.CyCategory', verbose_name='彩艺云类目', related_name='cy_cate', null=True,
                                   blank=True, on_delete=models.SET_NULL)
    is_use = models.BooleanField('是否启用', default=True)
    slug = models.CharField('标识', null=True, blank=True, max_length=15)

    class Meta:
        verbose_name_plural = verbose_name = '节目分类'
        ordering = ['-pk']

    def __str__(self):
        return self.name

    @classmethod
    def xunyan(cls):
        return cls.objects.get(slug='xunyan')

    @classmethod
    def dkxj(cls):
        return cls.objects.get(slug='dkxj')

    @classmethod
    def tkx(cls):
        return cls.objects.get(slug='tkx')

    @classmethod
    def dou_yin_cate(cls):
        data = dict()
        data['top'] = dict(name='休闲娱乐', category_id='4000000')
        data['second'] = dict(name='演出', category_id='4025000')
        data['third'] = dict(name='演唱会', category_id='4025001')
        data['third'] = dict(name='音乐节', category_id='4025002')
        data['third'] = dict(name='liveshow', category_id='4025003')
        data['third'] = dict(name='话剧音乐剧', category_id='4025004')
        data['third'] = dict(name='脱口秀', category_id='4025005')
        data['third'] = dict(name='儿童剧', category_id='4025006')
        data['third'] = dict(name='相声', category_id='4025007')
        data['third'] = dict(name='音乐会', category_id='4025008')
        data['third'] = dict(name='沉浸式演出', category_id='4025009')
        data['third'] = dict(name='舞蹈舞剧', category_id='4025010')
        data['third'] = dict(name='马戏', category_id='4025011')
        data['third'] = dict(name='戏曲', category_id='4025012')
        data['third'] = dict(name='其他演出', category_id='4025013')
        return data

    def show_type_copy_to_pika(self):
        from caches import get_pika_redis, redis_show_type_copy_key
        from ticket.serializers import ShowTypeSerializer
        data = ShowTypeSerializer(self).data
        with get_pika_redis() as pika:
            pika.hset(redis_show_type_copy_key, str(self.id), json.dumps(data))


class ShowContentCategorySecond(models.Model):
    cate = models.ForeignKey(ShowContentCategory, verbose_name='节目大类', on_delete=models.CASCADE)
    show_type = models.ForeignKey(ShowType, verbose_name='节目分类', on_delete=models.CASCADE)
    display_order = models.PositiveSmallIntegerField('排序', default=0, help_text='从小到大排序')

    class Meta:
        verbose_name_plural = verbose_name = '二级分类'
        unique_together = ['cate', 'show_type']
        ordering = ['display_order']

    def __str__(self):
        return f'{str(self.cate)}->{str(self.show_type)}'

    def show_content_second_copy_to_pika(self):
        from ticket.serializers import ShowContentCategorySecondSerializer
        from caches import get_pika_redis, redis_show_content_second_key
        with get_pika_redis() as pika:
            qs = ShowContentCategorySecond.objects.filter(cate=self.cate).order_by('display_order')
            show_type_list = ShowContentCategorySecondSerializer(qs, many=True).data
            pika.hset(redis_show_content_second_key, str(self.cate.id), json.dumps(dict(show_type_list=show_type_list)))


class Venues(UseNoAbstract):
    name = models.CharField('场馆名称', max_length=50, help_text='50个字内')
    layers = models.IntegerField('层数', default=1)
    city = models.ForeignKey(Division, verbose_name='城市', null=True, on_delete=models.SET_NULL,
                             limit_choices_to=models.Q(type=1))
    lat = models.FloatField('纬度', default=0)
    lng = models.FloatField('经度', default=0)
    address = models.CharField('详细地址', max_length=200)
    desc = models.TextField('详情描述', null=True)
    map = models.ImageField(u'剧场分区图', upload_to=f'{IMAGE_FIELD_PREFIX}/ticket/venues', null=True, blank=True,
                            validators=[validate_image_file_extension])
    is_use = models.BooleanField('是否启用', default=True)
    custom_mobile = models.CharField('客服电话', null=True, blank=True, max_length=20)
    custom_wechat = models.CharField('客服微信号', null=True, blank=True, max_length=50)
    custom_code = models.ImageField(u'客服微信二维码', upload_to=f'{IMAGE_FIELD_PREFIX}/ticket/venues', null=True, blank=True,
                                    validators=[validate_image_file_extension])
    display_order = models.IntegerField('排序', default=0, help_text='越小排越前')
    is_seat = models.BooleanField('是否已设座位', default=False)
    seat_data = models.TextField('座位数据', null=True, blank=True, editable=False)
    DIR_FORWARD = 1
    DIR_REVERSE = 2
    DIRECT_CHOICES = [(DIR_FORWARD, '正向'), (DIR_REVERSE, '反向')]
    direction = models.IntegerField('舞台方向', choices=DIRECT_CHOICES, default=DIR_FORWARD)

    class Meta:
        verbose_name_plural = verbose_name = '场馆'
        ordering = ['display_order']

    def __str__(self):
        return self.name

    def venues_detail_copy_to_pika(self):
        # log.debug('venues_detail_copy_to_pika')
        from caches import get_pika_redis, redis_venues_copy_key
        from ticket.serializers import VenuesSerializer
        data = VenuesSerializer(self).data
        data['price'] = float(data['price'])
        if hasattr(self, 'cy_venue'):
            data['cy_no'] = self.cy_venue.cy_no
        with get_pika_redis() as pika:
            pika.hset(redis_venues_copy_key, str(self.id), json.dumps(data))


class VenuesLayers(models.Model):
    venue = models.ForeignKey(Venues, verbose_name='场馆', on_delete=models.CASCADE)
    layer = models.IntegerField('楼层', default=0, help_text='1层填1')
    name = models.CharField('别名', max_length=10)

    class Meta:
        verbose_name_plural = verbose_name = '楼层管理'
        unique_together = ['venue', 'layer']

    def __str__(self):
        return str(self.id)


class VenuesLogoImage(models.Model):
    venue = models.ForeignKey(Venues, verbose_name='场馆', on_delete=models.CASCADE)
    image = models.ImageField('图片', upload_to=f'{IMAGE_FIELD_PREFIX}/mall/logo/venues',
                              validators=[file_size, validate_image_file_extension])

    class Meta:
        verbose_name_plural = verbose_name = '场馆封面图'

    def __str__(self):
        return str(self.id)


class VenuesDetailImage(models.Model):
    venue = models.ForeignKey(Venues, verbose_name='场馆', on_delete=models.CASCADE)
    image = models.ImageField('图片', upload_to=f'{IMAGE_FIELD_PREFIX}/mall/detail/venues',
                              validators=[validate_image_file_extension])

    class Meta:
        verbose_name_plural = verbose_name = '场馆详情图'

    def __str__(self):
        return str(self.id)


class Seat(models.Model):
    venue = models.ForeignKey(Venues, verbose_name='场馆', on_delete=models.CASCADE)
    seat_no = models.CharField('座位编号', null=True, max_length=100, editable=False, db_index=True)
    row = models.IntegerField('行数', default=0)
    column = models.IntegerField('列数', default=0)
    layers = models.IntegerField('层数', default=1)

    class Meta:
        verbose_name_plural = verbose_name = '场馆座位'
        unique_together = ['venue', 'row', 'column', 'layers']

    def __str__(self):
        return '{}楼-{}排{}座'.format(self.layers, self.row, self.column)


class ShowFlag(models.Model):
    title = models.CharField('名称', max_length=10, validators=[only_chinese])
    img = models.ImageField('标签图片', null=True, blank=True, upload_to=f'{IMAGE_FIELD_PREFIX}/ticket/show/flag',
                            validators=[validate_image_file_extension])

    class Meta:
        verbose_name_plural = verbose_name = '节目标签'

    def __str__(self):
        return self.title


class TikTokQualRecord(models.Model):
    qual_type_name = models.CharField('资质类型名称', max_length=50, null=True)
    qualification_id = models.CharField('资质ID', unique=True, max_length=20)
    Q_HOST = 5005
    Q_APPROVE = 5006
    Q_CHOICES = ((Q_HOST, u'演出主办⽅授权书'), (Q_APPROVE, '营业性演出准予许可决定'))
    qualification_type = models.IntegerField(u'资质类型', choices=Q_CHOICES, default=Q_HOST)
    STATUS_DEFAULT = 2
    STATUS_CHECK = 3
    STATUS_REJECT = 4
    STATUS_CHOICES = ((STATUS_DEFAULT, u'审核中'), (STATUS_CHECK, '审核通过'), (STATUS_REJECT, '审核拒绝'))
    status = models.IntegerField(u'资质状态', choices=STATUS_CHOICES, default=STATUS_DEFAULT)
    effective_date = models.CharField('有效期', max_length=30, help_text='永久/无期限的情况下传 9999-12-31', null=True)
    status_changed_time = models.DateTimeField('状态变更时间', null=True, blank=True)
    create_time = models.DateTimeField('创建时间', null=True, blank=True)
    life_account_id = models.CharField('资质对应账户id', max_length=20, null=True)
    qual_category = models.IntegerField('资质类目', null=True)

    class Meta:
        verbose_name_plural = verbose_name = '抖音资质记录'
        ordering = ['-pk']

    def __str__(self):
        return '{}({})'.format(self.qual_type_name, str(self.qualification_id))

    @classmethod
    def update_or_create(cls, page_index=1):
        from douyin import get_dou_yin
        dy = get_dou_yin()
        dd = dy.get_qual(page_index)
        update_list = []
        for data in dd['qual_dto_list']:
            inst, _ = cls.objects.get_or_create(qualification_id=str(data['qualification_id']),
                                                life_account_id=str(data['life_account_id']))
            inst.qual_type_name = data['attributes'][0]['attribute_values'][0]
            inst.qualification_type = data['qualification_type']
            inst.status = data['status']
            inst.effective_date = data['effective_date']
            inst.status_changed_time = data['status_changed_time']
            inst.create_time = data['create_time']
            inst.qual_category = data['qual_category']
            update_list.append(inst)
        if update_list:
            cls.objects.bulk_update(update_list, ['qual_type_name', 'qualification_type', 'status', 'effective_date',
                                                  'status_changed_time', 'create_time', 'qual_category'])
        if dd['total'] > dd['page'] * dd['page_size']:
            cls.update_or_create(dd['page'] + 1)

    @classmethod
    def pull_tiktok_qual(cls):
        close_old_connections()
        cls.update_or_create()


class ShowProject(UseNoAbstract):
    title = models.CharField('节目名称', max_length=100, help_text='100个字内')
    cate_second = models.ForeignKey(ShowContentCategorySecond, verbose_name='分类', on_delete=models.CASCADE, null=True)
    cate = models.ForeignKey(ShowContentCategory, verbose_name='内容分类', on_delete=models.SET_NULL, null=True,
                             editable=False)
    show_type = models.ForeignKey(ShowType, verbose_name='节目分类', on_delete=models.CASCADE, editable=False)
    venues = models.ForeignKey(Venues, verbose_name='场馆', on_delete=models.CASCADE, help_text='提交后不可修改')
    lat = models.FloatField('纬度', default=0, help_text='场馆纬度')
    lng = models.FloatField('经度', default=0, help_text='场馆经度')
    city_id = models.IntegerField('城市ID', editable=False, default=0)
    flag = models.ManyToManyField(ShowFlag, verbose_name='标签', blank=True)
    performer = models.ManyToManyField(ShowPerformer, verbose_name='演员', blank=True)
    logo_mobile = models.ImageField(u'宣传海报', upload_to=f'{IMAGE_FIELD_PREFIX}/ticket/shows',
                                    validators=[file_size, validate_image_file_extension], null=True, blank=True)
    sale_time = models.DateTimeField('开售时间')
    session_end_at = models.DateTimeField('场次最后结束时间', null=True, blank=True, editable=False, db_index=True)
    dy_show_date = models.CharField('本地演出日期', max_length=100, help_text='抖音用例如2024-01-01至2024-06-30', null=True,
                                    blank=True, editable=False)
    price = models.DecimalField('最低价格', max_digits=13, decimal_places=2, default=0, help_text='实际支付价格,用于展示和排序')
    # cert_type = models.ManyToManyField(CertificateType, verbose_name='支持的证件类型')
    name_limit_num = models.IntegerField('每个证件限购数量', default=1, help_text='实名制售票', editable=False)
    order_limit_num = models.IntegerField('每个订单限购数量', default=1, help_text='购票限制，0表示不限购', editable=False)
    # account_limit_num = models.IntegerField('每个账号限购数量', default=1, help_text='购票限制')
    origin_amount = models.DecimalField('抖音原价', max_digits=13, decimal_places=2, default=0, editable=False)
    ID_DEFAULT = 0
    ID_ORGANIZER = 1
    ID_TICKETAAGENT = 2
    # 1：主办方 ：主办方资质必填
    # 2：票务代理：主办方资质和票务代理资质必填
    # 主办方资质对应资质查询接口中的“营业性演出准予许可决定”
    # 票务代理资质对应资质查询接口中的“演出主办方授权书”
    ID_CHOICES = [(ID_DEFAULT, '无'), (ID_ORGANIZER, '主办方'), (ID_TICKETAAGENT, '票务代理')]
    qualification_identity = models.SmallIntegerField('商家资质身份', choices=ID_CHOICES, default=ID_DEFAULT, editable=False)
    host_approval_qual = models.ManyToManyField(TikTokQualRecord, verbose_name='营业性演出准予许可决定',
                                                limit_choices_to=models.Q(qualification_type=5006),
                                                help_text='主办方和票务代理,要传“营业性演出准予许可决定”5006', blank=True, related_name='+',
                                                editable=False)
    ticket_agent_qual = models.ManyToManyField(TikTokQualRecord, verbose_name='演出主办方授权书',
                                               limit_choices_to=models.Q(qualification_type=5005),
                                               help_text='票务代理,要传“演出主办方授权书” 5005', blank=True, related_name='+',
                                               editable=False)
    content = models.TextField('节目介绍', null=True)
    notice = models.TextField('购票须知')
    other_notice = models.TextField('其他说明信息', help_text='抖音商品使用', null=True, blank=True, editable=False)
    STATUS_ON = 1
    STATUS_OFF = 0
    STATUS_CHOICES = ((STATUS_ON, u'上架'), (STATUS_OFF, u'下架'))
    status = models.IntegerField(u'状态', choices=STATUS_CHOICES, default=STATUS_OFF)
    is_recommend = models.BooleanField('是否近期节目', default=True)
    wx_pay_config = models.ForeignKey(WeiXinPayConfig, verbose_name='微信支付', blank=True, null=True,
                                      on_delete=models.SET_NULL)
    dy_pay_config = models.ForeignKey(DouYinPayConfig, verbose_name='抖音支付商户', null=True, blank=True,
                                      on_delete=models.SET_NULL, editable=False)
    display_order = models.PositiveIntegerField('排序', default=0, help_text='越大越排前')
    create_at = models.DateTimeField('创建时间', auto_now_add=True)
    updated_at = models.DateTimeField(auto_now=True, verbose_name='更新时间')
    version = models.IntegerField('版本号', default=0)
    is_test = models.BooleanField('是否测试商品', default=False, editable=False)
    tiktok_code = models.ImageField('抖音分享二维码', null=True, blank=True, upload_to=f'{IMAGE_FIELD_PREFIX}/ticket/project',
                                    validators=[validate_image_file_extension], editable=False)
    wxa_code = models.ImageField('小程序分享二维码', null=True, blank=True, upload_to=f'{IMAGE_FIELD_PREFIX}/ticket/project',
                                 validators=[validate_image_file_extension])

    def __str__(self):
        return self.title

    class Meta:
        verbose_name_plural = verbose_name = '节目'
        ordering = ['-pk']

    @classmethod
    def import_record(cls, file_path):
        from django.conf import settings
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        venue = Venues.objects.get(name=data['source'])
        show_type = ShowType.objects.first()
        from common.qrutils import open_image_by_url
        for show in data['shows']:
            logo_mobile = open_image_by_url(show['cover_image'])
            sale_time = datetime.fromtimestamp(show['timestamps'][0])
            logo_mobile_path = f'{IMAGE_FIELD_PREFIX}/ticket/shows'
            file_path = os.path.join(settings.MEDIA_ROOT, logo_mobile_path)
            if not os.path.isdir(file_path):
                os.makedirs(file_path)
            file_name = f'{sha256_str(show["cover_image"])}.png'
            img = f'{file_path}/{file_name}'
            logo_mobile_path = f'{logo_mobile_path}/{file_name}'
            logo_mobile.save(img)
            end_at = sale_time + timedelta(hours=2)
            project, _ = cls.objects.get_or_create(title=show['title'], venues=venue, show_type=show_type,
                                                   sale_time=timezone.now(), content=show['introduction'],
                                                   notice=show['notice'],
                                                   lat=venue.lat, lng=venue.lng, city_id=venue.city.id,
                                                   price=Decimal(show['prices'][-1]), status=1,
                                                   logo_mobile=logo_mobile_path, session_end_at=end_at)
            session, _ = SessionInfo.objects.get_or_create(show=project, start_at=sale_time,
                                                           end_at=end_at, has_seat=2, status=1)
            colors = [['蓝色', '#5B9BD5'], ['浅绿', '#e2efda'], ['绿色', '#54BF31'], ['黄色', '#D5DB28'], ['紫色', '#9000FF'],
                      ['粉色', '#E863C2'], ['红色', '#F80B0B'], ['黄色', '#FFFF00'], ['粉色', '#F96AAD'], ['橙色', '#EE6E06']]
            i = 0
            for price in show['prices']:
                color, _ = TicketColor.objects.get_or_create(name=colors[i][0], code=colors[i][1])
                level = TicketFile.objects.get_or_create(session=session, title=project.title, color=color,
                                                         origin_price=Decimal(price), price=Decimal(price), stock=100)
                i += 1
                level.redis_stock()
            session.redis_show_date_copy()

    def change_session_end_at(self, end_at: datetime):
        if not self.session_end_at or self.session_end_at < end_at:
            self.session_end_at = end_at
            self.save(update_fields=['session_end_at'])

    def set_shows_no_pk(self):
        from caches import get_pika_redis, redis_shows_no_key
        with get_pika_redis() as redis:
            redis.hset(redis_shows_no_key, self.no, self.id)

    @classmethod
    def get_cache(cls, data: dict, show_id: int, user: User, is_tiktok: bool, is_ks: bool, is_xhs: bool = False):
        # inst = ShowCollectRecord.objects.filter(show_id=int(show_id), user=user).first()
        # data['is_collect'] = inst.is_collect if inst else False
        now = timezone.now()
        sale_time = datetime.strptime(data['sale_time'], '%Y-%m-%dT%H:%M:%S')
        session_end_at = datetime.strptime(data['session_end_at'], '%Y-%m-%dT%H:%M:%S') if data.get(
            'session_end_at') else None
        data['can_buy'] = data[
                              'status'] == ShowProject.STATUS_ON and sale_time <= now and session_end_at and session_end_at >= now
        from caches import get_pika_redis, redis_show_date_copy, redis_session_info_copy, \
            redis_session_info_tiktok_copy, redis_ticket_level_ks_cache, redis_ticket_level_xhs_cache, \
            redis_ticket_level_cache, redis_ticket_level_tiktok_cache, redis_show_type_copy_key, \
            redis_venues_copy_key, redis_show_content_copy_key, redis_session_info_ks_copy, redis_session_info_xhs_copy
        from ticket.stock_updater import tfc
        with get_pika_redis() as pika:
            data['show_type'] = json.loads(pika.hget(redis_show_type_copy_key, data['show_type']))
            data['venues'] = json.loads(pika.hget(redis_venues_copy_key, data['venues']))
            if data['cate']:
                data['cate'] = json.loads(pika.hget(redis_show_content_copy_key, data['cate']))
                data['cate'].pop('show_type_list', None)
            date_data = pika.hget(redis_show_date_copy, show_id)
            data['date'] = json.loads(date_data) if date_data else None
            if is_tiktok:
                session_data = pika.hget(redis_session_info_tiktok_copy, show_id)
            elif is_ks:
                session_data = pika.hget(redis_session_info_ks_copy, show_id)
            elif is_xhs:
                session_data = pika.hget(redis_session_info_xhs_copy, show_id)
            else:
                session_data = pika.hget(redis_session_info_copy, show_id)
            data['sessions'] = json.loads(session_data) if session_data else []
            if data['sessions']:
                # log.debug(data['sessions'])
                for session in data['sessions']:
                    session['ticket_level'] = []
                    if is_tiktok:
                        name = redis_ticket_level_tiktok_cache.format(session['no'])
                    elif is_ks:
                        name = redis_ticket_level_ks_cache.format(session['no'])
                    elif is_xhs:
                        name = redis_ticket_level_xhs_cache.format(session['no'])
                    else:
                        name = redis_ticket_level_cache.format(session['no'])
                    ticket_level_keys_list = pika.hkeys(name)
                    # log.debug(ticket_level_keys_list)
                    for key in ticket_level_keys_list:
                        level = pika.hget(name, key)
                        if level:
                            level = json.loads(level)
                            level['stock'] = tfc.get_stock(level['id'])
                            session['ticket_level'].append(level)
        return data

    @classmethod
    def init_all_cache(cls):
        from user_agents import parsers
        from functools import partial
        parsers.MOBILE_BROWSER_FAMILIES = parsers.MOBILE_BROWSER_FAMILIES + ('Chrome Mobile',)
        from concu import set_redis_provider
        from caches import get_redis_with_db
        set_redis_provider(partial(get_redis_with_db, db=3))
        from ticket.stock_updater import tfc
        tfc.pre_cache()
        st_qs = ShowType.objects.all()
        for inst in st_qs:
            inst.show_type_copy_to_pika()
        cate_qs = ShowContentCategory.objects.all()
        for inst in cate_qs:
            inst.show_content_copy_to_pika()
        venues_qs = Venues.objects.all()
        for inst in venues_qs:
            inst.venues_detail_copy_to_pika()
        shows_qs = cls.objects.filter(status=cls.STATUS_ON)
        for inst in shows_qs:
            session_qs = SessionInfo.objects.filter(show_id=inst.id, status=SessionInfo.STATUS_ON, is_delete=False)
            if session_qs:
                session = session_qs.first()
                session.redis_show_date_copy()
            for session in session_qs:
                level_qs = TicketFile.objects.filter(session_id=session.id)
                for level in level_qs:
                    level.redis_ticket_level_cache()
                    if level.status:
                        level.redis_stock()
            inst.shows_detail_copy_to_pika()

    def shows_detail_copy_to_pika(self):
        # log.debug('show_detail_init_or_delete')
        from caches import get_pika_redis, redis_shows_copy_key, redis_shows_no_key
        redis = get_pika_redis()
        if self.status == self.STATUS_OFF:
            # 下架删除，避免后面一直叠加
            redis.hdel(redis_shows_copy_key, str(self.id))
            redis.hdel(redis_shows_no_key, self.no)
        else:
            self.set_shows_no_pk()
            from common.config import get_config
            config = get_config()
            domain = config.get('template_url')
            show_dict = model_to_dict(self)
            pop_list = ['no', 'title', 'lat', 'lng', 'content', 'notice', 'display_order', 'is_test', 'status', 'cate',
                        'venues', 'show_type', 'is_recommend', 'dy_show_date']
            show_cache = dict()
            for key in pop_list:
                if show_dict.get(key):
                    show_cache[key] = show_dict[key]
            show_cache['id'] = self.no
            show_cache['price'] = float(self.price)
            show_cache['sale_time'] = datetime.strftime(self.sale_time, '%Y-%m-%dT%H:%M:%S')
            show_cache['session_end_at'] = datetime.strftime(self.session_end_at,
                                                             '%Y-%m-%dT%H:%M:%S') if self.session_end_at else None
            show_cache['origin_amount'] = float(self.origin_amount)
            show_cache['logo_mobile'] = '{}/{}'.format(domain, self.logo_mobile.url)
            # log.debug(show_cache)
            image_data = []
            qs = ShowsDetailImage.objects.filter(show_id=self.id)
            if qs:
                for img in qs:
                    image_data.append(dict(id=img.id, image='{}/{}'.format(domain, img.image.url)))
            show_cache['images'] = image_data if image_data else None
            from common.utils import get_timestamp
            show_cache['sale_time_timestamp'] = get_timestamp(self.sale_time) if self.sale_time else None
            if hasattr(self, 'cy_show'):
                show_cache['cy_no'] = self.cy_show.event_id
            redis.hset(redis_shows_copy_key, str(self.id), json.dumps(show_cache))

    @classmethod
    def get_calendar_key(cls, year: int, month: int, city_id: int = 0):
        from caches import pika_show_calendar_key
        key = pika_show_calendar_key.format(city_id)
        name = '{}_{}'.format(year, month)
        tiktok_name = '{}_tiktok'.format(name)
        ks_name = '{}_ks'.format(name)
        xhs_name = '{}_xhs'.format(name)
        return key, name, tiktok_name, ks_name, xhs_name

    @classmethod
    def get_show_calendar(cls, year: int, month: int, city_id: int = 0, is_tiktok=False, is_init=False, is_ks=False,
                          is_xhs=False):
        # is_tiktok 和is_init不会同时是true
        key, name, tiktok_name, ks_name, xhs_name = cls.get_calendar_key(year, month, city_id)
        from caches import get_pika_redis
        tiktok_data = None
        with get_pika_redis() as pika:
            if is_tiktok:
                data = pika.hget(key, tiktok_name)
            elif is_ks:
                data = pika.hget(key, ks_name)
            elif is_xhs:
                data = pika.hget(key, xhs_name)
            else:
                data = pika.hget(key, name)
            if not data or is_init:
                from common.dateutils import get_month_day
                start_at, end_at = get_month_day(year, month, 1)
                qs = SessionInfo.objects.filter(status=SessionInfo.STATUS_ON, start_at__gte=start_at,
                                                start_at__lt=end_at)
                if city_id:
                    qs = qs.filter(show__city_id=city_id)
                tiktok_qs = qs.filter(dy_status=SessionInfo.STATUS_ON)
                data = dict()
                for inst in qs:
                    d_key = inst.start_at.strftime('%Y-%m-%d')
                    data[d_key] = data[d_key] + 1 if data.get(d_key) else 1
                tiktok_data = dict()
                for inst in tiktok_qs:
                    d_key = inst.start_at.strftime('%Y-%m-%d')
                    tiktok_data[d_key] = tiktok_data[d_key] + 1 if tiktok_data.get(d_key) else 1
                if data:
                    pika.hset(key, name, json.dumps(data))
                if tiktok_data:
                    pika.hset(key, tiktok_name, json.dumps(tiktok_data))
                # 快手
                from kuaishou_wxa.models import KsGoodsConfig
                ks_data = KsGoodsConfig.ks_show_calendar(qs)
                if ks_data:
                    pika.hset(key, ks_name, json.dumps(ks_data))
                # 小红书
                from xiaohongshu.models import XhsGoodsConfig
                xhs_data = XhsGoodsConfig.xhs_show_calendar(qs)
                if xhs_data:
                    pika.hset(key, xhs_name, json.dumps(xhs_data))
                is_init = True
            else:
                data = json.loads(data)
        return data, tiktok_data, is_init

    def get_show_time(self):
        inst = SessionInfo.objects.filter(show=self).first()
        return '{}--{}'.format(inst.start_at.strftime('%Y-%m-%d %H:%M'), inst.end_at.strftime('%Y-%m-%d %H:%M'))

    @property
    def can_buy(self):
        return self.status == self.STATUS_ON and self.sale_time <= timezone.now() and self.session_end_at >= timezone.now()

    @classmethod
    def auto_expire_off(cls):
        close_old_connections()
        # 定时任务触发不了signal
        qs = cls.objects.filter(status=cls.STATUS_ON, session_end_at__lte=timezone.now(), session_end_at__isnull=False)
        for show in qs:
            show.status = cls.STATUS_OFF
            show.save(update_fields=['status'])
            show.shows_detail_copy_to_pika()
        qs_s = SessionInfo.objects.filter(status=SessionInfo.STATUS_ON, end_at__lte=timezone.now())
        for inst in qs_s:
            inst.status = SessionInfo.STATUS_OFF
            inst.save(update_fields=['status'])
            inst.change_show_calendar()
            inst.redis_show_date_copy()

    def get_wxa_code(self):
        return None
        from mp.models import SystemWxMP
        from django.db.models import signals
        sy = SystemWxMP.get()
        if not sy:
            return None
        url = 'pages/pagesKage/showDetail/showDetail'
        from mall.utils import qrcode_dir_tk
        dir, rel_url, img_dir = qrcode_dir_tk()
        filename = 'wxa_code_new{}_1_v{}.png'.format(self.pk, self.version)
        filepath = os.path.join(dir, filename)
        if not os.path.isfile(filepath) or not self.wxa_code:
            from mp.wechat_client import get_wxa_client
            wxa = get_wxa_client()
            scene = 'sg_{}_'.format(self.pk)
            buf = wxa.biz_get_wxa_code_unlimited(scene, url)
            if buf:
                from common.qrutils import save_code
                save_code(buf, filepath)
                self.wxa_code = '{}/{}'.format(img_dir, filename)
                self.save(update_fields=['wxa_code'])
        return self.wxa_code.url if self.wxa_code else None

    def get_tiktok_code(self):
        from mp.models import SystemDouYinMP
        dy = SystemDouYinMP.get()
        if not dy:
            return None
        url = 'pages/pagesKage/showDetail/showDetail'
        from mall.utils import qrcode_dir_tk
        dir, rel_url, img_dir = qrcode_dir_tk()
        from douyin import get_tiktok
        tk = get_tiktok()
        url = '{}?&id={}'.format(url, self.id)
        filename = 'dy_code_new{}_v{}.png'.format(self.pk, self.version)
        filepath = os.path.join(dir, filename)
        if not os.path.isfile(filepath) or not self.tiktok_code:
            buf = tk.get_qrcode(url)
            if buf:
                from common.qrutils import save_code
                save_code(buf, filepath)
                self.tiktok_code = '{}/{}'.format(img_dir, filename)
                self.save(update_fields=['tiktok_code'])
        return self.tiktok_code.url if self.tiktok_code else None


class ShowNotification(models.Model):
    show = models.ForeignKey(ShowProject, verbose_name='项目', on_delete=models.CASCADE, related_name='notification')
    title = models.CharField('标题', max_length=20, help_text='20字以内')
    content = models.TextField('内容', help_text='200字以内')

    class Meta:
        verbose_name_plural = verbose_name = '抖音商品使用规则'

    def __str__(self):
        return self.title


class ShowsDetailImage(models.Model):
    show = models.ForeignKey(ShowProject, verbose_name='节目', on_delete=models.CASCADE)
    image = models.ImageField('图片', upload_to=f'{IMAGE_FIELD_PREFIX}/ticket/shows/detail',
                              validators=[file_size, validate_image_file_extension])

    class Meta:
        verbose_name_plural = verbose_name = '节目详情图'

    def __str__(self):
        return str(self.id)

    def shows_detail_images_copy_to_pika(self):
        self.show.shows_detail_copy_to_pika()


# class TicketNotice(models.Model):
#     show = models.OneToOneField(ShowProject, verbose_name='项目', on_delete=models.CASCADE)
#     title = models.CharField('演出/活动时长', max_length=200, default='以现场为准')
#     appointment_desc = models.CharField('预约说明', max_length=200, default='无需预约')
#     children_desc = models.TextField('儿童说明', default='16岁以下谢绝入场')
#     prohibit_items = models.TextField('禁止携带物品', default='因版权限制，演出过程中禁止录音录像，入场前观众手机需放入定制手机套中并上锁，'
#                                                         '除不能拍照录像外不影响正常使用，手机套损坏照价赔偿，感谢您的支持与理解，谢谢配合！')
#     time_limit = models.TextField('付款时效提醒', default='下单成功后需在指定时间内完成支付，未支付成功的订单，'
#                                                     '将在下单指定时间后系统自动取消，请及时刷新购票页面进行购买。')
#
#     def __str__(self):
#         return self.title
#
#     class Meta:
#         verbose_name_plural = verbose_name = '购票须知'
#
#
# class CustomizeNotice(models.Model):
#     show = models.ForeignKey(ShowProject, verbose_name='项目', on_delete=models.CASCADE)
#     title = models.CharField('标题', max_length=20, help_text='20字以内')
#     desc = models.TextField('内容', help_text='1000字以内')
#
#     def __str__(self):
#         return self.title
#
#     class Meta:
#         verbose_name_plural = verbose_name = '自定义购票须知'


class ShowCompany(models.Model):
    name = models.CharField('名称', max_length=50, help_text='50个字内')
    create_at = models.DateTimeField('创建时间', auto_now_add=True)

    class Meta:
        verbose_name_plural = verbose_name = '合同主办方'


class ContractInfo(models.Model):
    show = models.OneToOneField(ShowProject, verbose_name='节目', on_delete=models.CASCADE)
    company = models.ForeignKey(ShowCompany, verbose_name='合同主办', on_delete=models.CASCADE,
                                help_text='主办方提交保存后不可变更，请谨慎选择')
    police_approval = models.ImageField(u'公安批文', upload_to=f'{IMAGE_FIELD_PREFIX}/ticket/contract',
                                        help_text=u'支持格式JPG/JPEG/PNG', null=True,
                                        blank=True, validators=[validate_image_file_extension])
    act_approval = models.ImageField(u'活动批文', upload_to=f'{IMAGE_FIELD_PREFIX}/ticket/contract',
                                     help_text=u'支持格式JPG/JPEG/PNG', null=True,
                                     blank=True, validators=[validate_image_file_extension])

    def __str__(self):
        return self.show.title

    class Meta:
        verbose_name_plural = verbose_name = '合同信息'


# class CertificateType(models.Model):
#     name = models.CharField('名称', max_length=50, help_text='50个字内')
#
#     class Meta:
#         verbose_name_plural = verbose_name = '证件类型'
#
#
# class SalesRules(models.Model):
#     show = models.OneToOneField(ShowProject, verbose_name='演出', on_delete=models.CASCADE)
#     # inside_sale_time = models.DateTimeField('内部销售时间')
#     # SALE_IN = 1
#     # SALE_OUT = 2
#     # SALE_CHOICES = [(SALE_IN, '同内部销售时间'), (SALE_OUT, '自定义开售时间')]
#     # outside_sale = models.SmallIntegerField('外部销售时间', choices=SALE_CHOICES, default=SALE_IN)
#     sale_time = models.DateTimeField('开售时间')
#     # BUY_NO = 1
#     # BUY_NAME = 2
#     # BUY_CHOICES = [(BUY_NO, '非实名制'), (BUY_NAME, '实名制')]
#     # buy_type = models.SmallIntegerField('实名制售票', choices=BUY_CHOICES, default=BUY_NAME)
#     cert_type = models.ManyToManyField(CertificateType, verbose_name='支持的证件类型')
#     name_limit_num = models.IntegerField('每个证件限购数量', default=1, help_text='实名制售票')
#     order_limit_num = models.IntegerField('每个订单限购数量', default=1, help_text='购票限制')
#     account_limit_num = models.IntegerField('每个账号限购数量', default=1, help_text='购票限制')
#     discount = models.BooleanField('会员折扣', default=False)
#
#     def __str__(self):
#         return self.show.title
#
#     class Meta:
#         verbose_name_plural = verbose_name = '销售规则'


class SessionInfo(UseNoAbstract):
    show = models.ForeignKey(ShowProject, verbose_name='项目', on_delete=models.CASCADE, related_name='session_info')
    venue_id = models.IntegerField('场馆ID', editable=False, default=0)
    title = models.CharField('场次名称', max_length=60, help_text='60个字内,不填则默认项目名称，其他平台使用', null=True,
                             blank=True)
    start_at = models.DateTimeField('场次开始时间', db_index=True)
    end_at = models.DateTimeField('场次结束时间', db_index=True)
    dy_sale_time = models.DateTimeField('开售时间', null=True, blank=True, help_text='其他平台用')
    tiktok_store = models.ForeignKey(DouYinStore, verbose_name='抖音店铺', null=True, blank=True, help_text='推送商品到抖音必填',
                                     on_delete=models.SET_NULL, editable=False)
    valid_start_time = models.DateTimeField('门票有效期开始时间', null=True, blank=True, help_text='抖音用：不填默认场次开始时间前2个小时',
                                            editable=False)
    desc = models.CharField('场次备注', max_length=100, null=True, blank=True)
    # 限购相关
    order_limit_num = models.IntegerField('每个订单限购数量', default=0, help_text='购票限制,0表示不限购', editable=False)
    is_name_buy = models.BooleanField('是否一单一证', default=False, help_text='0.01购票权限用户不走一单一证流程')
    name_buy_num = models.PositiveIntegerField('单证限购数量', default=0, help_text='用于限制一证最多购买多少张票，配合一单一证使用')
    one_id_one_ticket = models.BooleanField(default=False, verbose_name='是否一票一证购买')
    SEAT_HAS = 1
    SEAT_NO = 2
    SEAT_CHOICES = [(SEAT_HAS, '有座'), (SEAT_NO, '无座')]
    has_seat = models.IntegerField('是否有座', choices=SEAT_CHOICES, default=SEAT_HAS)
    SR_DEFAULT = 1
    SR_CY = 2
    SR_CHOICES = ((SR_DEFAULT, U'自建'), (SR_CY, '彩艺云'))
    source_type = models.SmallIntegerField(u'渠道类型', choices=SR_CHOICES, default=SR_DEFAULT)
    main_session = models.ForeignKey('self', verbose_name='关联主场次', on_delete=models.SET_NULL, null=True, blank=True,
                                     related_name='m_session', limit_choices_to=models.Q(has_seat=SEAT_HAS),
                                     help_text='有座场次才能选为主场次')
    STATUS_ON = 1
    STATUS_OFF = 2
    STATUS_CHOICES = ((STATUS_ON, u'上架'), (STATUS_OFF, u'下架'))
    status = models.IntegerField(u'状态', choices=STATUS_CHOICES, default=STATUS_OFF)
    is_delete = models.BooleanField('是否作废', default=False)
    is_theater_discount = models.BooleanField('是否参与剧场会员卡优惠', default=False)
    is_sale_off = models.BooleanField('标记售罄', default=False)
    is_paper = models.BooleanField('是否纸质票', default=False, help_text='勾选的场次对应的订单需要邮寄纸质票给客户')
    express_template = models.ForeignKey('express.Template', verbose_name='邮费模板', null=True, blank=True,
                                         help_text='选择物流模板',
                                         on_delete=models.SET_NULL)
    express_days = models.IntegerField('开演前多少天下单可邮寄', default=0, help_text='单位/天，设置后开演前n天下单的可以邮寄，'
                                                                           '不足n天的需要现场取票入场')
    close_comment = models.BooleanField('是否关闭评论', default=False)
    is_dy_code = models.BooleanField('是否动态码', default=True, help_text='勾选后需要填写动态有效时间才能有效')
    dc_expires_in = models.PositiveIntegerField('动态码有效时间', default=20, help_text='单位：分钟')
    create_at = models.DateTimeField('创建时间', auto_now_add=True)
    updated_at = models.DateTimeField(auto_now=True, verbose_name='更新时间')
    actual_amount = models.DecimalField('总实收', max_digits=9, decimal_places=2, default=0)

    is_price = models.BooleanField('是否已设价格', default=False)
    cache_seat = models.TextField('座位价格数据', editable=False, null=True, blank=True)
    PUSH_DEFAULT = 1
    PUSH_APPROVE = 2
    PUSH_SUCCESS = 3
    PUSH_NEED = 4
    PUSH_FAIL = 5
    PUSH_AUTH_FAIL = 6
    PUSH_CHOICES = (
        (PUSH_DEFAULT, u'未推送'), (PUSH_NEED, u'已推送'), (PUSH_APPROVE, u'审核中'), (PUSH_SUCCESS, u'审核完成'),
        (PUSH_FAIL, u'推送创建商品失败'), (PUSH_AUTH_FAIL, u'审核失败'))
    dy_status = models.IntegerField(u'抖音状态', choices=STATUS_CHOICES, default=STATUS_OFF, editable=False)
    push_status = models.IntegerField(u'推送抖音状态', choices=PUSH_CHOICES, default=PUSH_DEFAULT, editable=False)
    product_id = models.CharField('抖音商品ID', null=True, blank=True, max_length=50, editable=False)
    plan_id = models.CharField('通用计划ID', null=True, blank=True, max_length=50, editable=False)
    fail_msg = models.TextField('抖音错误信息', max_length=1000, null=True, blank=True, editable=False)
    mz_account = models.ForeignKey(MaiZuoAccount, verbose_name='麦座账户', null=True, blank=True, on_delete=models.SET_NULL,
                                   editable=False)
    PULL_DEFAULT = 1
    PULL_APPROVE = 2
    PULL_SUCCESS = 3
    PULL_FAIL = 4
    PULL_CHOICES = (
        (PULL_DEFAULT, U'未拉取'), (PULL_APPROVE, U'更新中'), (PULL_SUCCESS, U'已完成'), (PULL_FAIL, U'更新失败'))
    pull_mz_status = models.IntegerField(u'麦座同步状态', choices=PULL_CHOICES, default=PULL_DEFAULT, editable=False)
    maizuo_data = models.TextField('麦座基础数据', null=True, blank=True, editable=False)
    display_order = models.IntegerField('抖音商品排序', default=0, help_text='数字越大,商品越靠前', editable=False)

    def __str__(self):
        return '{}{}({})({})'.format(self.show.title, self.start_at.strftime('%Y-%m-%d %H:%M'),
                                     self.get_has_seat_display(), self.id)

    class Meta:
        verbose_name_plural = verbose_name = '场次'
        ordering = ['-start_at']

    @classmethod
    def push_to_ai_mysql(cls):
        def image_to_hex(image_path):
            """读取图片文件并转换为16进制字符串"""
            with open(image_path, 'rb') as f:
                binary_data = f.read()
            hex_data = binary_data.hex()
            return hex_data

        from django.db import connections
        import random
        qs = cls.objects.all().order_by('id')
        data = []
        sql = "INSERT INTO shows (slug,name,show_at,place,price,logo,rate) VALUES (%s,%s,%s,%s,%s,%s,%s)"
        with connections['sz_ai'].cursor() as cursor:
            for session in qs:
                show_at = session.start_at.strftime('%Y-%m-%d %H:%M:%S')
                show = session.show
                venue = show.venues
                logo = image_to_hex(show.logo_mobile.path.replace('\\', '/'))
                rate = f"{random.uniform(2.0, 5.0):.1f}"
                data.append([show.no, show.title, show_at, venue.name, str(show.price), logo, rate])
                cursor.execute(sql, data)

    @property
    def is_real_name_buy(self):
        return self.is_name_buy or self.one_id_one_ticket

    def clean(self):
        if self.is_real_name_buy and self.is_paper:
            raise ValidationError('场次的实名购票项与是否纸质票不能同时勾选')
        if self.has_seat == self.SEAT_HAS and self.main_session:
            raise ValidationError('有座场次不予许选择主场次')

    @property
    def check_is_dy_code(self):
        # 是否动态码
        return self.is_dy_code and self.dc_expires_in > 0

    def redis_show_date_copy(self):
        # 场次发生更新时都需要变化
        log.debug('redis_session_info_copy')
        from caches import get_pika_redis, redis_show_date_copy, redis_session_info_copy, \
            redis_session_info_tiktok_copy, redis_session_info_ks_copy, redis_session_info_xhs_copy, \
            redis_session_no_key

        inst = SessionInfo.objects.filter(show_id=self.show_id, status=self.STATUS_ON, is_delete=False).order_by(
            'start_at').first()
        create_at = self.show.create_at.strftime("%Y-%m-%dT%H:%M:%S")
        data_date = dict(start_at=inst.start_at.strftime("%Y-%m-%dT%H:%M:%S"),
                         end_at=inst.end_at.strftime("%Y-%m-%dT%H:%M:%S")) if inst else dict(start_at=create_at,
                                                                                             end_at=create_at)
        # 有排序，只能每次更改后一个场次后都排序重新写入
        qs = SessionInfo.objects.filter(show_id=self.show.id, status=SessionInfo.STATUS_ON, is_delete=False).order_by(
            'start_at')
        qs_tiktok = qs.filter(dy_status=SessionInfo.STATUS_ON, push_status=SessionInfo.PUSH_SUCCESS).order_by(
            'start_at')
        from ticket.serializers import ShowSessionCacheSerializer
        session_list = ShowSessionCacheSerializer(qs, many=True).data
        session_tiktok_list = session_ks_list = session_xhs_list = None
        # session_tiktok_list = ShowSessionCacheSerializer(qs_tiktok, many=True).data
        # from kuaishou_wxa.models import KsGoodsConfig
        # qs_ks = KsGoodsConfig.get_session_qs(qs).order_by('start_at')
        # session_ks_list = ShowSessionCacheSerializer(qs_ks, many=True).data
        # from xiaohongshu.models import XhsGoodsConfig
        # xhs_ks = XhsGoodsConfig.get_session_qs(qs).order_by('start_at')
        # session_xhs_list = ShowSessionCacheSerializer(xhs_ks, many=True).data
        with get_pika_redis() as pika:
            if self.status == SessionInfo.STATUS_ON:
                pika.hset(redis_session_no_key, self.no, self.id)
            else:
                pika.hdel(redis_session_no_key, self.no)
            pika.hset(redis_show_date_copy, str(self.show_id), json.dumps(data_date))
            pika.hset(redis_session_info_copy, str(self.show_id), json.dumps(session_list))
            if session_tiktok_list:
                pika.hset(redis_session_info_tiktok_copy, str(self.show_id), json.dumps(session_tiktok_list))
            else:
                pika.hdel(redis_session_info_tiktok_copy, str(self.show_id))
            if session_ks_list:
                pika.hset(redis_session_info_ks_copy, str(self.show_id), json.dumps(session_ks_list))
            else:
                pika.hdel(redis_session_info_ks_copy, str(self.show_id))
            if session_xhs_list:
                pika.hset(redis_session_info_xhs_copy, str(self.show_id), json.dumps(session_xhs_list))
            else:
                pika.hdel(redis_session_info_xhs_copy, str(self.show_id))
        tf_qs = TicketFile.objects.filter(session_id=self.id)
        # 更改票档缓存
        if tf_qs:
            if self.status == self.STATUS_OFF:
                for level in tf_qs:
                    level.redis_ticket_level_cache(is_create=False)
            else:
                for level in tf_qs:
                    level.redis_ticket_level_cache()

    @property
    def express_end_at(self):
        dd = self.start_at - timedelta(days=self.express_days)
        from common.utils import get_timestamp
        return get_timestamp(dd)

    def check_express_fee_date(self):
        dd = self.start_at - timedelta(days=self.express_days)
        if timezone.now() > dd:
            return True
        return False

    def clear_actual_amount(self):
        self.actual_amount = 0
        self.save(update_fields=['actual_amount'])

    def update_actual_amount(self, amount):
        """
        更改实收
        """
        from caches import get_redis, session_actual_amount_key
        redis = get_redis()
        redis.lpush(session_actual_amount_key, json.dumps([self.id, float(amount)]))

    def change_show_calendar(self):
        year = self.start_at.year
        month = self.start_at.month
        city_id = self.show.city_id
        ShowProject.get_show_calendar(year, month, city_id, is_init=True)
        # 全国也要初始化一次
        if city_id:
            ShowProject.get_show_calendar(year, month, 0, is_init=True)

    @classmethod
    def task_add_actual_amount(cls):
        from caches import get_redis, session_actual_amount_key
        redis = get_redis()
        amount_list = redis.lrange(session_actual_amount_key, 0, -1)
        if amount_list:
            data_dict = dict()
            for value in amount_list:
                val = redis.rpop(session_actual_amount_key)
                if val:
                    val = json.loads(val)
                    session_id = val[0]
                    amount = Decimal(val[1])
                    key = str(session_id)
                    if data_dict.get(key):
                        data_dict[key] += amount
                    else:
                        data_dict[key] = amount
            if data_dict:
                for key, value in data_dict.items():
                    session_id = int(key)
                    session = cls.objects.filter(id=session_id).first()
                    if session:
                        session.actual_amount += Decimal(value)
                        session.save(update_fields=['actual_amount'])

    def set_delete(self, is_delete):
        self.is_delete = is_delete
        self.save(update_fields=['is_delete'])

    def set_mz_status(self, status):
        self.pull_mz_status = status
        self.save(update_fields=['pull_mz_status'])

    @property
    def mz_lock_msg(self):
        return '已售卖锁座,'

    def mai_zuo_lock(self, lock_ticket_ids: list):
        if self.maizuo_data:
            maizuo_data = json.loads(self.maizuo_data)
            from maizuo import get_mai_zuo
            mz = get_mai_zuo(self.mz_account.name, self.mz_account.password)
            try:
                ticket_ids = ','.join(lock_ticket_ids)
                remark = '{}{}'.format(self.mz_lock_msg, self.id)
                mz.ticket_lock(maizuo_data['event_id'], ticket_ids, maizuo_data['project_id'],
                               maizuo_data['perform_template_id'], remark)
            except Exception as e:
                log.error('锁座失败,{}'.format(e))

    def mai_zuo_unlock(self, unlock_ticket_ids: list, session_seat_ids: list):
        if self.maizuo_data:
            from maizuo import get_mai_zuo
            mz = get_mai_zuo(self.mz_account.name, self.mz_account.password)
            try:
                maizuo_data = json.loads(self.maizuo_data)
                st = mz.ticket_unlock(maizuo_data['event_id'], unlock_ticket_ids, maizuo_data['project_id'],
                                      maizuo_data['perform_template_id'])
                if st:
                    SessionSeat.objects.filter(id__in=session_seat_ids).update(push_mz_lock=False)
            except Exception as e:
                log.error('解锁座失败,{}'.format(e))

    def get_mai_zuo_data_new(self, is_init=False, seat_dict=None):
        """
        is_init:是否初始化
        """
        log.warning('新开始同步,{},{}'.format(timezone.now().strftime('%Y%m%d%H%M%S'), self.id))
        from maizuo import get_mai_zuo
        mz_account = self.mz_account
        if not mz_account:
            return False, '场次未选择麦座账户', None
        mz = get_mai_zuo(mz_account.name, mz_account.password)
        maizuo_data = json.loads(self.maizuo_data) if self.maizuo_data else None
        content, ret = mz.get_seat_data(show_name=self.show.title, start_at=self.start_at, maizuo_data=maizuo_data)
        if ret and content:
            from caches import get_pika_redis, matrix_seat_data_key
            matrix_seat_data_key = matrix_seat_data_key.format(self.id)
            # 做异操作的话要两边数据一致，不然^出来会错误。因为没对上，每个循环找的话就不存在这种错误,而且字典不按顺序排的，还要做排序算法
            if not self.maizuo_data or is_init:
                self.maizuo_data = json.dumps(ret)
                self.save(update_fields=['maizuo_data'])
            # 先排序，麦座座位开始买后不能修改，不然同步就出错了
            content = sorted(content, key=lambda x: x['performSeatId'])
            lock_ticket_ids = []
            unlock_ticket_ids = []
            session_seat_ids = []
            # 用于存放初始化的麦座结构数据
            matrix_seat_data = None
            need_change_seat_list = []
            # 开始坐标
            start_index = 0
            # 结束坐标
            end_index = 2
            if not is_init:
                # 非初始化的数据匹配处理
                with get_pika_redis as redis:
                    # 已缓存的矩阵二进制数据
                    matrix_seat_data_cache = redis.get(matrix_seat_data_key)
                    if not matrix_seat_data_cache:
                        return False, '该场次未初始化同步麦座', None
                    seat_list = []
                    for seat_s in content:
                        # 两个一组
                        ticket_s = seat_s['ticket']
                        has_lock = True if ticket_s.get('lockTagId') else False
                        # 判断是否这边售卖然后推的锁座
                        own_bug = 0
                        if ticket_s.get('remark') and self.mz_lock_msg in ticket_s['remark']:
                            own_bug = 1
                        m_val = '{}{}{}'.format(int(not ticket_s['canOperate']), own_bug, int(has_lock))
                        matrix_seat_data = matrix_seat_data + m_val if matrix_seat_data else m_val
                        seat_list.append(seat_s)
                    origin_len = len(matrix_seat_data)
                    if origin_len != len(matrix_seat_data_cache):
                        return False, '该场次座位数量发生了变化，导致同步失败', None
                    if matrix_seat_data != matrix_seat_data_cache:
                        # 转整型，异或
                        matrix_seat_data_int = int(matrix_seat_data, 2)
                        matrix_seat_data_cache_int = int(matrix_seat_data_cache, 2)
                        # 执行异或操作判断是否有变化
                        result = matrix_seat_data_int ^ matrix_seat_data_cache_int
                        # 转2进制去除前缀'0b'
                        result = bin(result)[2:]
                        # 补0
                        need_zero_len = origin_len - len(result)
                        if need_zero_len > 0:
                            result = '0' * need_zero_len + result
                        st_width = end_index - start_index + 1
                        # 000表示没有变化
                        not_change_val = '0' * st_width
                        # 数据分组
                        group_list = group_by_str(result, st_width)
                        i = 0
                        # 遍历寻找变更了的数据
                        for group in group_list:
                            val = ''.join(group)
                            if val != not_change_val:
                                need_change_seat_list.append(seat_list[i])
                            i += 1
                        # 把需要遍历更改的填入
                        content = need_change_seat_list
            for seat in content:
                # 一楼(二楼)，1排(2排)，'35', '1914438029'
                # seat_list.append(dict(layer=seat['standName'], row=seat['rowName'], column=seat['seatNum'],
                #                       ticketIds=seat['performSeatId']))
                ticket = seat['ticket']
                has_lock = True if ticket.get('lockTagId') else False
                own_bug = 0
                if ticket.get('remark') and self.mz_lock_msg in ticket['remark']:
                    own_bug = 1
                matrix_val = '{}{}{}'.format(int(not ticket['canOperate']), own_bug, int(has_lock))
                lock_ticket_id, unlock_ticket_id, sessionseat_id = SessionSeat.mai_zuo_set_seat_new(
                    session_id=self.id, venue_id=self.venue_id, stand_name=seat['standName'],
                    floor_name=seat['floorName'], row=seat['rowName'], column=seat['seatNum'],
                    ticket_ids=seat['performSeatId'], can_buy=ticket['canOperate'], has_lock=has_lock, is_init=is_init,
                    seat_dict=seat_dict, matrix_val=matrix_val, start_index=start_index, end_index=end_index)
                if not is_init:
                    if lock_ticket_id:
                        lock_ticket_ids.append(lock_ticket_id)
                    if unlock_ticket_id:
                        unlock_ticket_ids.append(unlock_ticket_id)
                        session_seat_ids.append(sessionseat_id)
                else:
                    # 初始化座位的矩阵数据
                    # 麦座是否已卖，小程序默认未卖0，是否锁定，以麦座为标准
                    val = '{}{}{}'.format(int(not ticket['canOperate']), 0, int(has_lock))
                    matrix_seat_data = matrix_seat_data + val if matrix_seat_data else val
                    width = end_index - start_index
                    start_index = end_index + 1
                    end_index = start_index + width
            if lock_ticket_ids:
                # 锁座
                self.mai_zuo_lock(lock_ticket_ids)
            if unlock_ticket_ids:
                # 解锁座
                self.mai_zuo_unlock(unlock_ticket_ids, session_seat_ids)
            if is_init and matrix_seat_data:
                # 初始化数据
                with get_pika_redis as redis:
                    redis.set(matrix_seat_data_key, matrix_seat_data)
            log.warning('结束同步,{}'.format(timezone.now().strftime('%Y%m%d%H%M%S')))
            return True, None, None
        else:
            msg = '拉取数据失败, 场次标题和时间没对上或登录失败,{}'.format(str(self))
            log.error(msg)
        return False, msg, ret

    @classmethod
    def auto_sync_mai_zuo_key(cls):
        from caches import pika_auto_sync_mai_zuo_key
        return pika_auto_sync_mai_zuo_key

    @classmethod
    def auto_sync_mai_zuo(cls):
        # 定时同步和卖座的状态
        # 1分钟执行一次
        import time
        from random import randint
        # key = cls.auto_sync_mai_zuo_key()
        # 正在执行更新的话，任务直接结束
        # 已完成状态才能推送
        qs = cls.objects.filter(status=cls.STATUS_ON, has_seat=cls.SEAT_HAS, end_at__gt=timezone.now(),
                                pull_mz_status=cls.PULL_SUCCESS)
        for inst in qs:
            ret, msg, is_login = inst.get_mai_zuo_data_new()
            if not ret:
                if is_login == False:
                    # 登录超时会自动登录，直接结束
                    break
                else:
                    msg = '同步场次数据失败,{}'.format(msg)
                    log.error(msg)
                    MaiZuoLoginLog.create_record('同步场次数据失败,{}'.format(msg), MaiZuoLoginLog.TY_SESSION)
            # 随机延迟1-3秒
            # time.sleep(randint(1, 3))

    def get_product_id(self):
        product_id = self.product_id
        # if not product_id:
        #     tf = TicketFile.objects.filter(session_id=self.id, push_status=TicketFile.PUSH_SUCCESS, is_tiktok=True,
        #                                    product_id__isnull=False).order_by('origin_price').first()
        #     if tf:
        #         product_id = tf.product_id
        return product_id

    def update_start_at_and_end_at(self, start_at, end_at):
        self.start_at = start_at
        self.end_at = end_at
        self.venue_id = self.show.venues.id
        self.save(update_fields=['start_at', 'end_at', 'venue_id'])
        # 复制场次初始化
        self.change_show_calendar()
        # if valid_start_time:
        #     session.valid_start_time = valid_start_time
        #     fields.append('valid_start_time')
        if not self.show.session_end_at or self.show.session_end_at < self.end_at:
            self.show.session_end_at = self.end_at
            self.show.save(update_fields=['session_end_at'])

    @atomic
    def layer_session(self, level_dict=None):
        from django.forms.models import model_to_dict
        from caches import get_pika_redis, pika_session_seat_list_key, pika_session_seat_key, pika_level_seat_key
        pika = get_pika_redis()
        session = self
        ticket_levels = TicketFile.objects.filter(session_id=session.id)
        session_dict = model_to_dict(session)
        # editable=False 不会转换
        session_dict['cache_seat'] = session.cache_seat
        session_id = session_dict.pop('id')
        session_dict.pop('no', None)
        session_dict['show_id'] = session_dict.pop('show')
        session_dict['tiktok_store_id'] = session_dict.pop('tiktok_store')
        session_dict['express_template_id'] = session_dict.pop('express_template')
        session_dict['product_id'] = None
        session_dict['fail_msg'] = None
        session_dict['plan_id'] = None
        session_dict['push_status'] = self.PUSH_DEFAULT
        session_dict['maizuo_data'] = None
        session_dict['actual_amount'] = 0
        session_dict['is_delete'] = False
        # 主场次不复制
        session_dict.pop('main_session')
        # session_dict['create_at'] = timezone.now()
        ss = SessionInfo.objects.create(**session_dict)
        session_inst = ss
        session_seat_list_key = pika_session_seat_list_key.format(ss.id)
        seat_list = []
        pika_list = []
        if ticket_levels:
            i = 0
            session_seat_key = pika_session_seat_key.format(ss.id)
            for inst in ticket_levels:
                dd = model_to_dict(inst)
                level_id = dd.pop('id')
                dd['stock'] = 0
                if level_dict and level_dict.get(str(level_id)):
                    dd['stock'] = level_dict[str(level_id)]
                dd['session'] = ss
                dd['color_id'] = dd.pop('color')
                dd['sales'] = 0
                dd['product_id'] = None
                dd['fail_msg'] = None
                dd['can_cps'] = False
                dd['plan_id'] = None
                dd['push_status'] = TicketFile.PUSH_DEFAULT
                # dd['create_at'] = timezone.now()
                level = TicketFile.objects.create(**dd)
                level.redis_ticket_level_cache()
                level.redis_stock()
                seats_qs = SessionSeat.objects.filter(session_id=session_id, ticket_level=inst)
                ticket_level_id = level.id
                for st in seats_qs:
                    seats = st.seats
                    seat_list.append(
                        SessionSeat(ticket_level_id=int(ticket_level_id), seats=seats, row=seats.row,
                                    column=seats.column,
                                    layers=seats.layers, session_id=ss.id,
                                    color_id=level.color.id, price=level.price, color_code=level.color.code,
                                    is_reserve=st.is_reserve, showRow=st.showRow, box_no_special=st.box_no_special,
                                    showCol=st.showCol, desc=st.desc))
                    pika_seat = dict(ticket_level=ticket_level_id, seats=seats.id, row=seats.row,
                                     column=seats.column,
                                     layers=seats.layers, session_id=ss.id,
                                     color_id=level.color.id, price=float(level.price), color_code=level.color.code,
                                     is_reserve=st.is_reserve, showRow=st.showRow, box_no_special=st.box_no_special,
                                     showCol=st.showCol, desc=st.desc, can_buy=not st.is_reserve)
                    level_seat_key = pika_level_seat_key.format(ticket_level_id, seats.id)
                    pika_list.append(pika_seat)
                    pika_seat['index'] = i
                    pika.hdel(session_seat_key, level_seat_key)
                    r = pika.hset(session_seat_key, level_seat_key, json.dumps(pika_seat))
                    if r != 1:
                        return False, '执行失败, pika错误', None
                    i += 1
            if pika_list:
                # clogger.debug(pika_list)
                pika.set(session_seat_list_key, json.dumps(pika_list))
            if seat_list:
                SessionSeat.objects.bulk_create(seat_list)
        return True, None, session_inst

    @classmethod
    def can_push_status(cls):
        return [cls.PUSH_DEFAULT, cls.PUSH_APPROVE, cls.PUSH_FAIL, cls.PUSH_AUTH_FAIL]

    def set_status(self, status):
        if self.push_status == self.PUSH_SUCCESS:
            self.push_on_or_off_to_dy(status)
        self.status = status
        self.dy_status = status
        self.save(update_fields=['status', 'dy_status'])

    def set_dy_status(self, dy_status):
        self.push_on_or_off_to_dy(dy_status)
        self.dy_status = dy_status
        self.save(update_fields=['dy_status'])

    def push_on_or_off_to_dy(self, status):
        from douyin import get_dou_yin
        dy = get_dou_yin()
        if self.product_id:
            # 新
            ret = dy.goods_operate(self.product_id, status)
            if ret['error_code'] != 0:
                log.error('上下架失败,{}'.format(self.product_id))
        else:
            # 旧
            tf_qs = TicketFile.objects.filter(session=self, product_id__isnull=False,
                                              push_status=TicketFile.PUSH_SUCCESS)
            if tf_qs:
                for inst in tf_qs:
                    ret = dy.goods_operate(inst.product_id, status)
                    if ret['error_code'] != 0:
                        log.error('上下架失败,{}'.format(inst.product_id))

    @property
    def can_buy(self):
        return self.status == self.STATUS_ON and self.end_at > timezone.now()

    def dy_can_buy(self):
        return self.status == self.STATUS_ON and self.dy_status == self.STATUS_ON and self.end_at > timezone.now()

    # def skus_push_dou_yin(self):
    #     # 需要更改
    #     tf_qs = TicketFile.objects.filter(session=self)
    #     skus = []
    #     for tf in tf_qs:
    #         skus.append(tf.get_dy_sku_data())
    #     params = dict(product_id=self.product_id, skus=skus)
    #     from douyin import get_dou_yin
    #     dy = get_dou_yin()
    #     ret = dy.skus_batch_save(params)
    #     log.debug(ret)
    #     if ret['error_code'] == 0:
    #         return True
    def check_goods_from_dou_yin(self):
        from douyin import get_dou_yin
        dy = get_dou_yin()
        session = self
        product_ids = [self.product_id]
        data = dict(get_draft_type=0, product_ids=product_ids)
        ret = dy.product_draft(data)
        log.debug(ret)
        if ret['error_code'] == 0 and ret['product_drafts']:
            for dd in ret['product_drafts']:
                product_id = dd['sku']['sku_id']
                draft_status = dd['draft_status']
                # tf = TicketFile.objects.filter(product_id=product_id).first()
                if session.product_id == product_id:
                    if draft_status == 12:
                        session.push_status = session.PUSH_AUTH_FAIL
                        push_status = TicketFile.PUSH_FAIL
                    elif draft_status in [1, 11]:
                        session.push_status = session.PUSH_SUCCESS
                        push_status = TicketFile.PUSH_SUCCESS
                    else:
                        session.push_status = session.PUSH_APPROVE
                        push_status = TicketFile.PUSH_APPROVE
                    session.fail_msg = dd['audit_msg'] if dd['audit_msg'] else None
                    session.save(update_fields=['push_status', 'fail_msg'])
                    qs = TicketFile.objects.filter(session_id=session.id, is_tiktok=True)
                    # if push_status == TicketFile.PUSH_SUCCESS:
                    #     if not qs.filter(can_cps=True):
                    #         tf = qs.order_by('origin_price').first()
                    #         tf.can_cps = True
                    #         tf.save(update_fields=['can_cps'])
                    qs.update(push_status=push_status)
                    session.redis_show_date_copy()

    @classmethod
    def goods_draft_dou_yin(cls):
        """
        draft_status
        草稿状态 10-审核中 12-审核失败 1-审核通过 11-审核通过
        """
        close_old_connections()
        qs = cls.objects.filter(push_status__in=[cls.PUSH_NEED, cls.PUSH_APPROVE])
        from douyin import get_dou_yin
        dy = get_dou_yin()
        for session in qs:
            if session.product_id:
                #  新
                session.check_goods_from_dou_yin()
            # else:
            #     # 旧
            #     product_id_list = list(
            #         TicketFile.objects.filter(session=session, is_tiktok=True,
            #                                   push_status__in=TicketFile.check_status()).values_list(
            #             "product_id", flat=True))
            #     if not product_id_list:
            #         session.push_status = cls.PUSH_SUCCESS
            #         session.save(update_fields=['push_status'])
            #     else:
            #         product_ids_list = []
            #         total = ceil(len(product_id_list) / 10)
            #         if total > 1:
            #             # 每次最多查询10个
            #             for i in list(range(1, total + 1)):
            #                 product_ids_list.append(','.join(product_ids_list[(i - 1) * 10:i * 10]))
            #         else:
            #             product_ids_list.append(product_id_list[0])
            #         for product_ids in product_ids_list:
            #             data = dict(get_draft_type=0, product_ids=product_ids)
            #             ret = dy.product_draft(data)
            #             log.debug(ret)
            #             if ret['error_code'] == 0 and ret['product_drafts']:
            #                 for dd in ret['product_drafts']:
            #                     product_id = dd['sku']['sku_id']
            #                     draft_status = dd['draft_status']
            #                     tf = TicketFile.objects.filter(product_id=product_id).first()
            #                     if tf:
            #                         if draft_status == 12:
            #                             tf.push_status = TicketFile.PUSH_FAIL
            #                         elif draft_status in [1, 11]:
            #                             tf.push_status = TicketFile.PUSH_SUCCESS
            #                         else:
            #                             tf.push_status = TicketFile.PUSH_APPROVE
            #                         tf.fail_msg = dd['audit_msg'] if dd['audit_msg'] else None
            #                         tf.save(update_fields=['push_status', 'fail_msg'])

    def get_session_out_id(self):
        return 'p{}'.format(self.id)

    def get_dy_sku_data_session(self):
        # from dj_ext.exceptions import AdminException
        # if self.show.price < 50:
        #     AdminException('价格需要大于50元才可以推送')
        if self.show.origin_amount <= 0 or self.show.origin_amount < self.show.price:
            raise CustomAPIException('抖音原价需要大于0元和最低价才可以推送')
        limit_rule = {"is_limit": False, "total_buy_num": 99}
        stock = {"limit_type": 2, "stock_qty": 99999}
        data = {
            "sku_name": '{}'.format(self.show.price),
            "origin_amount": int(self.show.origin_amount * 100),
            "actual_amount": int(self.show.price * 100),
            "stock": stock,
            # "out_sku_id": 'sk'.format(tf.id),
            "status": 1,
            "attr_key_value_map": {
                "code_source_type": "2",
                "limit_rule": json.dumps(limit_rule, ensure_ascii=False),
                # "market_price": "900",
                "settle_type": "1",
                "use_type": "1",
            },
        }
        return data

    def get_dy_product_name(self):
        return self.title or '{}·{}'.format(self.start_at.strftime("%m月%d日"), self.show.title)

    def goods_push_dou_yin_new(self):
        # https://bytedance.feishu.cn/docx/doxcnnyH289B98IgcPiLWxpSChc 创建商品
        show = self.show
        if not show.show_type.category or not show.show_type.category.category_id:
            return False, '请先关联抖音类目'
        else:
            # if not force:
            #     tf_qs = TicketFile.objects.filter(session=self, is_tiktok=True,
            #                                       push_status__in=TicketFile.can_push_status())
            # else:

            # 取价格最小的记录
            tf_qs = TicketFile.objects.filter(session=self, is_tiktok=True).order_by('origin_price')
            if not tf_qs:
                return False, '没有满足条件的票档，请确认是否推送到抖音'
            if not self.tiktok_store:
                return False, '请先选择抖音店铺'
            from douyin import get_dou_yin
            from mp.models import SystemDouYinMP
            dy_mp = SystemDouYinMP.get()
            dy = get_dou_yin()
            config = get_config()
            appointment = {"need_appointment": False}
            # 不可使用日期
            can_no_use_date = {"enable": False}
            # 是否留客户资料
            customer_reserved_info = {"allow": True, "allow_tel": True, "allow_name": True, "require_for_tel": True}
            # note_type 1是文本 2是图片
            description_rich_text = None
            if show.other_notice:
                description_rich_text = [{"note_type": 1, "content": show.other_notice}]
            pois = []
            image_list = []
            environment_image_list = []
            detail_image_list = []
            qs = ShowsDetailImage.objects.filter(show=show)
            for img in qs:
                detail_image_list.append({"url": '{}{}'.format(config['template_url'], img.image.url)})
            qs = VenuesLogoImage.objects.filter(venue=show.venues)
            for img in qs:
                environment_image_list.append({"url": '{}{}'.format(config['template_url'], img.image.url)})
            image_list.append({"url": '{}{}'.format(config['template_url'], show.logo_mobile.url)})
            FrontCategoryTag = ['团购']
            Notification = []
            TagList = ','.join([str(flag) for flag in show.flag.all()])
            for nt in show.notification.all():
                Notification.append({"title": nt.title, "content": nt.content})
            """
            ONLY_ONE_INFO  = 2   // 仅填写一位游客信息
            EVERY_ONE_INFO = 1   // 每张门票都要填写用户信息
            """
            real_name_info = {"enable": False, "scene": 2}
            # 可使用时间加7天,取消了。因为要按资质
            use_date = {"use_date_type": 1, "use_start_date": self.start_at.strftime('%Y-%m-%d'),
                        "use_end_date": self.end_at.strftime('%Y-%m-%d')}
            # 1全天2时间段
            use_time = {"use_time_type": 1}
            # 主办批文资质ID
            ticket_agent_qual = ""
            host_approval_qual = ""
            if show.host_approval_qual.all():
                approval_qual_list = list(show.host_approval_qual.all().values_list('qualification_id', flat=True))
                host_approval_qual = {"QualificationId": approval_qual_list}
            if show.ticket_agent_qual.all():
                agent_qual_list = list(show.ticket_agent_qual.all().values_list('qualification_id', flat=True))
                ticket_agent_qual = {"QualificationId": agent_qual_list}
            pois.append({"poi_id": self.tiktok_store.supplier_ext_id})
            app_id = dy_mp.app_id
            params = {"id": str(show.id), 'session_id': str(self.id)}
            # out_id = "ss{}_{}".format(show.id, self.id)
            sold_start_time = int(get_timestamp(self.dy_sale_time) / 1000) if self.dy_sale_time else int(
                get_timestamp(show.sale_time) / 1000)
            product_name = self.get_dy_product_name()
            product_ext = dict(auto_online=self.dy_status != self.STATUS_OFF)
            config = get_config()
            is_tiktok_debug = False
            if config.get('tiktok') and config['tiktok']['debug']:
                is_tiktok_debug = True
            if is_tiktok_debug:
                product_ext['test_extra'] = dict(uids=config['tiktok']['uids'], test_flag=True)
            data = {
                "product": {
                    "attr_key_value_map": {
                        "qualification_identity": str(show.qualification_identity),
                        "host_approval_qual": json.dumps(host_approval_qual),
                        # "ticket_agent_qual": json.dumps(ticket_agent_qual) if ticket_agent_qual else "{}",
                        "appointment": json.dumps(appointment),
                        "auto_renew": "false",
                        # "bring_out_meal": "false",
                        "can_no_use_date": json.dumps(can_no_use_date),
                        "customer_reserved_info": json.dumps(customer_reserved_info),
                        "Description": json.dumps([show.content], ensure_ascii=False),
                        # 详情图
                        "detail_image_list": json.dumps(detail_image_list),
                        # "dishes_image_list": "[{\"url\":\"https:xxxxx\"}]",
                        "EntryType": "2",
                        # 环境图
                        "environment_image_list": json.dumps(environment_image_list),
                        # "free_pack": "false",
                        "FrontCategoryTag": json.dumps(FrontCategoryTag, ensure_ascii=False),
                        # 封面图
                        "image_list": json.dumps(image_list),
                        "IndustryType": "其他",
                        "IsConfirmImme": "true",
                        "Notification": json.dumps(Notification, ensure_ascii=False),
                        #  是否可以使用包间
                        # "private_room": "false",
                        "real_name_info": json.dumps(real_name_info, ensure_ascii=False),
                        # 推荐语
                        # "RecommendWord": "",
                        # "rec_person_num": "99",
                        # "rec_person_num_max": "999",
                        # 1-允许退款 2-不可退款 3-有条件退
                        "RefundPolicy": "2",
                        "refund_need_merchant_confirm": "true",
                        # 投放渠道 1-不限制 2-仅直播间可见
                        "show_channel": "1",
                        # 排序权重
                        "SortWeight": str(self.display_order) or "0",
                        # "superimposed_discounts": "true",
                        "TagList": TagList,
                        "use_date": json.dumps(use_date),
                        "use_time": json.dumps(use_time),
                        # "account_name": "lh测试商家",
                        # "poi_list": "[{\"poi_id\": \"123123123123\"}]",
                        # "product_name": "migrate_openapi_0711_01",
                        # "sold_start_time": "1646724999",
                        # "sold_end_time": "1745607528",
                        # 1营业性演出准予许可证2演出主办方授权书
                    },
                    "product_ext": product_ext,
                    # 5 小程序，不可更新
                    "out_url": json.dumps({
                        "app_id": app_id,
                        "params": json.dumps(params),
                        "path": tiktok_goods_url
                    }),
                    "biz_line": 5,
                    # 不可更新
                    "category_id": show.show_type.category.category_id,
                    # "out_id": "ss{}_{}".format(show.id, self.id),
                    "pois": pois,
                    "product_name": product_name,
                    # 不可更新
                    "product_type": 1,
                    "sold_end_time": int(get_timestamp(self.end_at) / 1000),
                    "sold_start_time": sold_start_time,
                    # "telephone": ["1234-4321"]
                }
            }
            if ticket_agent_qual:
                data['product']['attr_key_value_map']['ticket_agent_qual'] = json.dumps(ticket_agent_qual)
            if description_rich_text:
                data['product']['attr_key_value_map']['description_rich_text'] = json.dumps(description_rich_text,
                                                                                            ensure_ascii=False)
            if is_tiktok_debug:
                # 测试商品必须填写
                data['product']['attr_key_value_map']['trade_url'] = data['product']['out_url']
            skus = self.get_dy_sku_data_session()
            if is_tiktok_debug:
                # 测试库存不能大于50
                skus['stock'] = {"limit_type": 1, "stock_qty": 40}
            data['sku'] = skus
            data['product']['out_id'] = self.get_session_out_id()
            try:
                ret = dy.goods_dy_create(data)
                log.debug(ret)
                if ret['error_code'] == 0:
                    self.push_status = self.PUSH_APPROVE
                    self.product_id = ret['product_id']
                    self.save(update_fields=['push_status', 'product_id'])
                    tf_qs.update(push_status=TicketFile.PUSH_APPROVE, product_id=ret['product_id'])
                    # tf.push_status = tf.PUSH_APPROVE
                    # tf.product_id = ret['product_id']
                    # tf.save(update_fields=['product_id', 'push_status'])
                # 推skus
                # self.skus_push_dou_yin()
            except Exception as e:
                self.push_status = self.PUSH_FAIL
                self.fail_msg = e
                self.save(update_fields=['push_status', 'fail_msg'])
                from dj_ext.exceptions import AdminException
                log.error(e)
                tf_qs.update(push_status=TicketFile.PUSH_PUSH_FAIL)
                # tf.push_status = tf.PUSH_PUSH_FAIL
                # tf.fail_msg = e
                # tf.save(update_fields=['fail_msg', 'push_status'])
                raise AdminException(e)
            return True, ''

    @property
    def is_ks_session(self):
        return hasattr(self, 'ks_session')

    @property
    def is_xhs_session(self):
        return hasattr(self, 'xhs_session')


class SessionChangeSaleTimeRecord(models.Model):
    session = models.ForeignKey(SessionInfo, verbose_name='场次', on_delete=models.CASCADE)
    old_sale_time = models.DateTimeField('旧开售时间', null=True, blank=True)
    new_sale_time = models.DateTimeField('新开售时间', null=True, blank=True)

    class Meta:
        verbose_name_plural = verbose_name = '抖音开售时间修改记录'
        ordering = ['-pk']

    def __str__(self):
        return str(self.id)

    @classmethod
    def create(cls, session, sale_time):
        st, msg = SessionPushTiktokTask.create_record(session, '抖音开售时间修改')
        if not st:
            return st, msg
        cls.objects.create(session=session, new_sale_time=sale_time,
                           old_sale_time=session.dy_sale_time or session.show.sale_time)
        session.dy_sale_time = sale_time
        session.save(update_fields=['dy_sale_time'])
        return True, ''


class SessionChangeRecord(models.Model):
    user = models.ForeignKey(User, verbose_name='操作用户', on_delete=models.SET_NULL, null=True)
    session = models.ForeignKey(SessionInfo, verbose_name='场次', on_delete=models.CASCADE)
    old_start_at = models.DateTimeField('旧开始时间', null=True, blank=True)
    new_start_at = models.DateTimeField('新开始时间', null=True, blank=True)
    old_end_at = models.DateTimeField('旧结束时间', null=True, blank=True)
    new_end_at = models.DateTimeField('新结束时间', null=True, blank=True)
    create_at = models.DateTimeField('操作时间', auto_now_add=True, null=True)

    class Meta:
        verbose_name_plural = verbose_name = '场次时间延长记录'
        ordering = ['-pk']

    def __str__(self):
        return str(self.id)

    @classmethod
    def create(cls, session, user, new_end_at=None, new_start_at=None):
        st, msg = SessionPushTiktokTask.create_record(session, '场次时间延长')
        if session.is_ks_session:
            # 重新推快手
            session.ks_session.re_push()
        if not st:
            return st, msg
        inst = cls.objects.create(session=session, user=user)
        fields = []
        fields_s = []
        if new_start_at:
            inst.new_start_at = new_start_at
            inst.old_start_at = session.start_at
            session.start_at = new_start_at
            fields.append('start_at')
            fields_s.append('new_start_at')
            fields_s.append('old_start_at')
        if new_end_at:
            # 用来搜索的
            if not session.show.session_end_at or session.show.session_end_at < new_end_at:
                session.show.session_end_at = new_end_at
                session.show.save(update_fields=['session_end_at'])
            inst.new_end_at = new_end_at
            inst.old_end_at = session.end_at
            session.end_at = new_end_at
            fields.append('end_at')
            fields_s.append('new_end_at')
            fields_s.append('old_end_at')
        if fields:
            inst.save(update_fields=fields_s)
            session.save(update_fields=fields)
            if new_start_at:
                session.change_show_calendar()
        return st, msg


class TicketColor(models.Model):
    name = models.CharField('名称', max_length=10)
    code = models.CharField('色号', max_length=10, help_text='16进制色号')
    is_use = models.BooleanField('是否使用', default=True)

    class Meta:
        verbose_name_plural = verbose_name = '票档颜色'

    def __str__(self):
        return self.name


class TicketFile(models.Model):
    session = models.ForeignKey(SessionInfo, verbose_name='场次', on_delete=models.CASCADE, related_name='session_level')
    title = models.CharField('节目名称', max_length=100, null=True, blank=True)
    color = models.ForeignKey(TicketColor, verbose_name='票档颜色', on_delete=models.CASCADE, null=True, blank=True)
    color_code = models.CharField('色号', max_length=10, help_text='16进制色号', null=True, blank=True, editable=False)
    origin_price = models.DecimalField('票档', max_digits=13, decimal_places=2, default=0)
    price = models.DecimalField('售价', max_digits=13, decimal_places=2, default=0)
    stock = models.PositiveIntegerField('库存数量', default=0)
    sales = models.PositiveIntegerField('销量', default=0)
    desc = models.CharField('票档描述', max_length=20, null=True, blank=True)
    status = models.BooleanField(u'是否上架', default=True)
    create_at = models.DateTimeField('创建时间', auto_now_add=True)

    total_stock = models.PositiveIntegerField('总数量', default=0, editable=False)
    lock_stock = models.BooleanField('无限库存', default=False, editable=False)
    limit_num = models.PositiveIntegerField('每单限购份数', default=0, editable=False)
    version = models.IntegerField('版本', editable=False, default=0)
    product_id = models.CharField('抖音商品ID', null=True, blank=True, max_length=50, editable=False)
    plan_id = models.CharField('通用计划ID', null=True, blank=True, max_length=50, editable=False)
    is_tiktok = models.BooleanField('是否推送到抖音', default=False, editable=False)
    is_ks = models.BooleanField('是否推送到快手', default=False, editable=False)
    is_xhs = models.BooleanField('是否推送到小红书', default=False, editable=False)
    push_xhs = models.BooleanField('是否已经推送小红书', default=False, editable=False)
    PUSH_DEFAULT = 1
    PUSH_PUSH_FAIL = 2
    PUSH_APPROVE = 3
    PUSH_SUCCESS = 4
    PUSH_FAIL = 5
    PUSH_CHOICES = (
        (PUSH_DEFAULT, u'未推送'), (PUSH_PUSH_FAIL, u'推送失败'), (PUSH_APPROVE, u'审核中'),
        (PUSH_SUCCESS, u'审核成功'), (PUSH_FAIL, u'审核失败'))
    push_status = models.IntegerField(u'推送抖音状态', choices=PUSH_CHOICES, default=PUSH_DEFAULT, editable=False)
    fail_msg = models.TextField('抖音错误信息', max_length=1000, null=True, blank=True, editable=False)
    can_cps = models.BooleanField('是否cps选择', default=False, editable=False)

    class Meta:
        verbose_name_plural = verbose_name = '票档'
        unique_together = ['session', 'color']

    def __str__(self):
        return '{},{}'.format(self.session.id, str(self.session))

    @classmethod
    def update_stock_from_redis(cls):
        close_old_connections()
        from ticket.stock_updater import tfc
        tfc.persist()

    def redis_ticket_level_cache(self, is_create=True):
        # 创建和修改的时候触发
        # log.debug('redis_ticket_level_cache')
        from caches import get_pika_redis, redis_ticket_level_cache, redis_ticket_level_tiktok_cache, \
            redis_ticket_level_ks_cache, redis_ticket_level_xhs_cache
        name = redis_ticket_level_cache.format(self.session.no)
        tiktok_name = redis_ticket_level_tiktok_cache.format(self.session.no)
        ks_name = redis_ticket_level_ks_cache.format(self.session.no)
        xhs_name = redis_ticket_level_xhs_cache.format(self.session.no)
        key = str(self.id)
        with get_pika_redis() as pika:
            if is_create and self.status:
                from ticket.serializers import TicketFileCacheSerializer
                data = TicketFileCacheSerializer(self).data
                data['origin_price'] = float(data['origin_price'])
                data['price'] = float(data['price'])
                dd = json.dumps(data)
                pika.hset(name, key, dd)
                if self.is_tiktok:
                    pika.hset(tiktok_name, key, dd)
                else:
                    pika.hdel(tiktok_name, key)
                if self.is_ks:
                    pika.hset(ks_name, key, dd)
                else:
                    pika.hdel(ks_name, key)
                if self.is_xhs:
                    pika.hset(xhs_name, key, dd)
                else:
                    pika.hdel(xhs_name, key)
            else:
                pika.hdel(name, key)
                pika.hdel(tiktok_name, key)
                pika.hdel(ks_name, key)

    def redis_stock(self, stock=None):
        # 初始化库存
        if self.session.has_seat == SessionInfo.SEAT_NO:
            from ticket.stock_updater import tfc, StockModel
            if stock == None:
                stock = self.stock
            tfc.append_cache(StockModel(_id=self.id, stock=stock))
            # gc_list = GoodSpecConfig.objects.filter(good=self)
            # for spec in gc_list:
            #     gscc.append_cache(StockModel(_id=spec.id, stock=spec.stock))

    def delete_redis_stock(self):
        if self.session.has_seat == SessionInfo.SEAT_NO:
            from ticket.stock_updater import tfc
            tfc.remove(self.id)

    @classmethod
    def delete_redis_stock_all(cls, session):
        if session.has_seat == SessionInfo.SEAT_NO:
            qs = cls.objects.filter(session_id=session.id)
            for inst in qs:
                inst.delete_redis_stock()

    def get_price(self, card, tc_card, pay_type, multiply, use_old_card=False):
        # 会员卡金额大于0才会有优惠
        init_price = self.price * multiply
        discount_theater = False
        if card and tc_card and pay_type == Receipt.PAY_CARD_JC and self.session.show.show_type == ShowType.dkxj():
            t_card_ids = None
            if not use_old_card:
                t_card_ids = list(
                    TheaterCardUserDetail.objects.filter(user_card_id=tc_card.id, amount__gt=0).values_list('card_id',
                                                                                                            flat=True))
            city = self.session.show.venues.city
            has_city = False
            if city:
                if t_card_ids:
                    inst = TheaterCardCity.objects.filter(card_id__in=t_card_ids, cities=city).order_by(
                        'discount').first()
                else:
                    inst = TheaterCardCity.objects.filter(card_id=card.id, cities=city).first()
                if inst and tc_card and tc_card.amount > 0:
                    init_price = quantize(init_price * inst.discount / 100, 2)
                    has_city = True
                    discount_theater = True
            if not has_city and tc_card and tc_card.amount > 0:
                if t_card_ids:
                    inst = TheaterCardTicketLevel.objects.filter(card_id__in=t_card_ids, title=self.desc).order_by(
                        'discount').first()
                else:
                    inst = TheaterCardTicketLevel.objects.filter(card_id=card.id, title=self.desc).first()
                if inst:
                    init_price = quantize(init_price * inst.discount / 100, 2)
                    discount_theater = True
        return init_price, discount_theater

    def get_out_id(self):
        return "ss{}_{}_{}".format(self.session.show_id, self.session_id, self.id)

    def get_product_id(self):
        product_id = self.session.product_id
        if not product_id:
            tf = TicketFile.objects.filter(session_id=self.session_id, push_status=self.PUSH_SUCCESS, is_tiktok=True,
                                           product_id__isnull=False).order_by('origin_price').first()
            if tf:
                product_id = tf.product_id
        return product_id

    def set_log(self, request, msg):
        from django.contrib import admin
        log_inst = admin.ModelAdmin(TicketFile, admin.site)
        log_inst.log_change(request, self, msg)

    @classmethod
    def update_goods_sales(cls):
        from caches import get_redis, level_sales_key
        redis = get_redis()
        sales_list = redis.lrange(level_sales_key, 0, -1)
        if sales_list:
            total_list = dict()
            for val in sales_list:
                dd = redis.rpop(level_sales_key)
                id, num = dd.split('_')
                if total_list.get(str(id)):
                    total_list[str(id)] += int(num)
                else:
                    total_list[str(id)] = int(num)
            for key, val in total_list.items():
                self = cls.objects.filter(id=int(key)).first()
                if self:
                    sales = self.sales + val
                    self.sales = sales if sales > 0 else 0
                    self.save(update_fields=['sales'])

    def update_sales(self, num):
        """
        更改销量
        """
        from caches import get_redis, level_sales_key
        redis = get_redis()
        if num:
            redis.lpush(level_sales_key, '{}_{}'.format(self.id, num))

    @classmethod
    def check_status(cls):
        return [cls.PUSH_APPROVE, cls.PUSH_PUSH_FAIL]

    @classmethod
    def can_push_status(cls):
        return [cls.PUSH_DEFAULT, cls.PUSH_PUSH_FAIL, cls.PUSH_FAIL]

    # def change_stock(self, mul):
    #     """
    #     变更库存.是否考虑安全更新(即库存不能为负数)
    #     :param mul: 正数为加，负数为扣
    #     """
    #     from caches import get_redis
    #     from restframework_ext.exceptions import CustomAPIException
    #     redis = get_redis()
    #     key = 'update_goods_stock_lock'.format(self.id)
    #     if redis.llen(key) <= 50:
    #         redis.lpush(key, mul)
    #         self.refresh_from_db(fields=['stock', 'version'])
    #         rows = 0
    #         try:
    #             rows = TicketFile.objects.filter(pk=self.pk, version=self.version, stock__gte=-mul).update(
    #                 version=models.F('version') + 1, stock=models.F('stock') + mul)
    #         except Exception as e:
    #             redis.delete(key)
    #         finally:
    #             redis.rpop(key)
    #         if rows <= 0:
    #             raise CustomAPIException('下单失败,库存不足')
    #     else:
    #         raise CustomAPIException('抢购失败,请稍后再试')
    @pysnooper.snoop(log.debug)
    def change_stock(self, mul):
        levels_upd = []
        levels_upd.append((self.pk, mul, 0))
        from ticket.stock_updater import tfc
        succ1, tfc_result = tfc.batch_incr(levels_upd)
        if succ1:
            tfc.batch_record_update_ts(tfc.resolve_ids(tfc_result))
        else:
            log.warning(f"ticket_levels incr failed")
            raise CustomAPIException('抢购失败,库存不足')

    def get_dy_sku_data(self):
        from dj_ext.exceptions import AdminException
        if self.price < 50:
            AdminException('价格需要大于50元才可以推送')
        if self.limit_num > 0:
            limit_rule = {"is_limit": True, "total_buy_num": self.limit_num}
        else:
            limit_rule = {"is_limit": False, "total_buy_num": 99}
        if self.session.has_seat == SessionInfo.SEAT_HAS:
            stock = {"limit_type": 2, "stock_qty": 99999}
        else:
            stock = {"limit_type": 1, "stock_qty": self.stock}
        data = {
            "sku_name": self.desc if self.desc else self.session.show.title,
            "origin_amount": self.origin_price * 100,
            "actual_amount": self.price * 100,
            "stock": stock,
            # "out_sku_id": 'sk'.format(tf.id),
            "status": 1,
            "attr_key_value_map": {
                "code_source_type": "2",
                "limit_rule": json.dumps(limit_rule, ensure_ascii=False),
                # "market_price": "900",
                "settle_type": "1",
                "use_type": "1",
            },
        }
        return data

    def get_xhs_sku_data(self, sku_image: str):
        data = {
            "out_sku_id": self.get_out_id(),
            "name": self.desc if self.desc else self.session.show.title,
            "sku_image": sku_image,
            "origin_price": int(self.origin_price * 100),
            "sale_price": int(self.price * 100),
        }
        return data

    @classmethod
    def get_order_no_seat_amount(cls, user, ticket_list: list, pay_type: int, session, is_tiktok=False, express_fee=0):
        total_multiply = 0
        amount = 0
        actual_amount = 0
        level_list = []
        discount_type = TicketOrder.DISCOUNT_DEFAULT
        tc_card = TheaterCardUserRecord.objects.filter(user=user).first()
        card = None
        use_old_card = False
        if pay_type == Receipt.PAY_CARD_JC and session.show.show_type == ShowType.dkxj():
            old_card_qs = TheaterCardUserDetail.get_old_cards(user.id)
            old_card_detail = old_card_qs.first()
            if old_card_detail:
                card_amount = 0
                for data in ticket_list:
                    multiply = int(data['multiply'])
                    inst = data.get('level')
                    if inst:
                        price, _ = inst.get_price(card, tc_card, pay_type, multiply, use_old_card)
                        card_amount += price
                    else:
                        raise CustomAPIException('下单失败，请重新选择')
                if old_card_detail.amount > card_amount + express_fee:
                    use_old_card = True
                    card = old_card_detail.card
                log.debug(use_old_card)
        if not card:
            card = TheaterCard.get_inst()
        for data in ticket_list:
            multiply = int(data['multiply'])
            # inst = cls.objects.filter(id=data['level_id'], session=session).first()
            inst = data.get('level')
            if inst:
                if is_tiktok and inst.push_status != TicketFile.PUSH_SUCCESS:
                    raise CustomAPIException('该商品未上架{}'.format(inst.price))
                if inst.limit_num > 0 and multiply > inst.limit_num:
                    raise CustomAPIException(
                        '价格{}的票,超过限购数量{}，请重新选择'.format(inst.price, inst.limit_num))
                # if not inst.lock_stock:
                #     if multiply > inst.stock:
                #         raise CustomAPIException('下单失败，库存不足')
                #     else:
                #         inst.change_stock(-multiply)
                total_multiply += multiply
                amount += inst.price * multiply
                price, discount_theater = inst.get_price(card, tc_card, pay_type, multiply, use_old_card)
                if discount_theater:
                    discount_type = TicketOrder.DISCOUNT_THEATER
                actual_amount += price
                level_list.append(dict(level=inst, multiply=multiply))
            else:
                raise CustomAPIException('下单失败，请重新选择')
        if pay_type in [Receipt.PAY_WeiXin_LP, Receipt.PAY_KS]:
            account = user.account
            discount = account.get_discount()
            actual_amount = amount * discount
            if discount < 1:
                discount_type = TicketOrder.DISCOUNT_YEAR
        from common.utils import quantize
        return total_multiply, amount, quantize(actual_amount, 2), level_list, discount_type


class ShowUser(models.Model):
    user = models.ForeignKey(User, verbose_name='用户', on_delete=models.CASCADE)
    name = models.CharField('姓名', max_length=30)
    mobile = models.CharField('手机号', max_length=20, null=True)
    id_card = models.CharField('身份证号', max_length=20, null=True, blank=True, db_index=True)
    create_at = models.DateTimeField('创建时间', auto_now_add=True)

    class Meta:
        verbose_name_plural = verbose_name = '常用联系人'
        unique_together = ['user', 'name', 'mobile']
        ordering = ['-pk']

    def __str__(self):
        return self.name

    @classmethod
    def auth_cert_no(cls, name, id_card):
        from alibaba import get_inst
        inst = get_inst()
        ret = inst.check_cert_no(name, id_card)
        return ret


class TicketReceiptManager(Manager):
    """
    filter by role implicitly.
    """

    def get_queryset(self):
        return super().get_queryset().filter(biz=Receipt.BIZ_TICKET)


class TicketReceipt(Receipt):
    objects = TicketReceiptManager()

    class Meta:
        proxy = True
        verbose_name = verbose_name_plural = '收款记录'
        ordering = ['-pk']


class SessionSeat(models.Model):
    ticket_level = models.ForeignKey(TicketFile, verbose_name='票档', on_delete=models.CASCADE, null=True)
    seats = models.ForeignKey(Seat, verbose_name='座位', on_delete=models.CASCADE, null=True)
    row = models.IntegerField('行数', default=0)
    column = models.IntegerField('列数', default=0)
    layers = models.IntegerField('层数', default=1)
    price = models.DecimalField('售价', max_digits=13, decimal_places=2, default=0)
    color_code = models.CharField('色号', max_length=10, help_text='16进制色号', null=True, blank=True)
    session_id = models.IntegerField('场次ID', default=0, editable=False, db_index=True)
    color_id = models.IntegerField('票档颜色ID', default=0, editable=False)
    showRow = models.CharField('座位排号', editable=False, default=0, null=True, blank=True, max_length=10)
    showCol = models.CharField('座位列号', editable=False, default=0, null=True, blank=True, max_length=10)
    box_no_special = models.CharField('区域', max_length=50, null=True, blank=True)
    is_reserve = models.BooleanField('是否预留', default=False)
    desc = models.CharField('预留描述', max_length=50, null=True, blank=True)
    is_buy = models.BooleanField('是否被锁定', default=False)
    order_no = models.CharField(u'订单号', max_length=128, null=True, blank=True, db_index=True)
    version = models.IntegerField('版本', editable=False, default=0)
    ticket_ids = models.CharField('麦座座位号', max_length=30, null=True, blank=True)
    buy_desc = models.CharField('购买描述', max_length=50, null=True, blank=True)
    start_index = models.PositiveIntegerField('数据缓存开始索引', default=0, editable=False, help_text='用于同步麦座')
    end_index = models.PositiveIntegerField('数据缓存结束始索引', default=0, editable=False, help_text='用于同步麦座')
    push_mz_lock = models.BooleanField('是否推送麦座锁座', default=False)

    class Meta:
        verbose_name_plural = verbose_name = '场次座位'
        unique_together = ['seats', 'session_id']

    def __str__(self):
        if self.box_no_special:
            return '{}层{}{}排{}列'.format(self.layers, self.box_no_special, self.showRow, self.showCol)
        return '{}层{}排{}列'.format(self.layers, self.showRow, self.showCol)

    def seat_desc(self, venue):
        vl = VenuesLayers.objects.filter(venue=venue, layer=self.layers).first()
        if vl:
            if not self.box_no_special:
                seat = '{}{}排{}列'.format(vl.name, self.showRow, self.showCol)
            else:
                seat = '{}{}{}排{}列'.format(vl.name, self.box_no_special, self.showRow,
                                           self.showCol)
        else:
            seat = '{}'.format(str(self))
        return seat

    @classmethod
    def get_pika_mz_seat_key(cls, session_id: int, layers: int, showRow, showCol, box_no_special=None):
        from caches import pika_session_mz_layer
        key = '{}_seat_{}_{}_{}'.format(pika_session_mz_layer, layers, showRow, showCol)
        if box_no_special:
            key = '{}_seat_{}_{}_{}_{}'.format(pika_session_mz_layer, layers, showRow, showCol, box_no_special)
        return '{}_{}'.format(pika_session_mz_layer, session_id), key

    @classmethod
    def get_pika_buy_key(cls, session_id: int, ticket_ids: str):
        from caches import pika_session_mz_buy
        key = '{}_seat_{}_{}'.format(pika_session_mz_buy, session_id, ticket_ids)
        return '{}_{}'.format(pika_session_mz_buy, session_id), key

    def get_row_and_box_no_special(self):
        # 新的同步使用
        try:
            # 如果这边只是排数只是数字,则需要补个排字。因为麦座返回的是1排
            row = int(self.showRow)
            row = '{}排'.format(row)
        except Exception as e:
            row = self.showRow
        box_no_special = self.box_no_special
        chinese_numbers = {
            0: '零', 1: '一', 2: '二', 3: '三', 4: '四',
            5: '五', 6: '六', 7: '七', 8: '八', 9: '九'
        }
        if not box_no_special:
            # 麦座默认是**楼
            box_no_special = '{}楼'.format(chinese_numbers[self.layers])
        return row, box_no_special

    def change_pika_mz_seat(self, is_buy):
        if self.ticket_ids:
            from caches import get_pika_redis, matrix_seat_data_key
            result = '{}_{}_{}'.format(self.ticket_ids, int(is_buy), int(self.is_reserve))
            # 麦座肯定是未售的，3种情况会进来，1这边售卖了，麦座未售 2.退款，限制麦座未售才进来，3取消和退款一样的逻辑
            matrix_val = '{}{}{}'.format(0, int(is_buy), int(self.is_reserve))
            with get_pika_redis() as pika:
                name, key = SessionSeat.get_pika_mz_seat_key(self.session_id, self.layers, self.showRow, self.showCol,
                                                             self.box_no_special)
                pika.hset(name, key, result)
                row, box_no_special = self.get_row_and_box_no_special()
                new_name, new_key = SessionSeat.get_pika_mz_seat_key(self.session_id, self.layers, row,
                                                                     self.showCol,
                                                                     box_no_special)
                pika.hset(new_name, new_key, result)
                self.matrix_seat_change_pika(matrix_val)

    def set_pika_buy(self, is_buy):
        if self.ticket_ids:
            from caches import get_pika_redis
            with get_pika_redis() as pika:
                name, key = SessionSeat.get_pika_buy_key(self.session_id, self.ticket_ids)
                pika.hset(name, key, 1 if is_buy else 0)

    def matrix_seat_change_pika(self, matrix_val):
        # matrix_val新的值
        from caches import get_pika_redis, matrix_seat_data_key
        with get_pika_redis() as pika:
            # 外面方法确保了座位数量没有发生变化
            matrix_seat_data_key = matrix_seat_data_key.format(self.session_id)
            pika.setrange(matrix_seat_data_key, self.start_index, matrix_val)

    @classmethod
    def mai_zuo_set_seat_new(cls, session_id: int, venue_id: int, stand_name: str, floor_name: str, row: str,
                             column: str, ticket_ids: str, can_buy: bool, has_lock: bool, is_init: bool,
                             matrix_val: str, start_index: int, end_index: int, seat_dict=None):
        """
        floor_name: 一层
        layer: standName 区名
        row: 1排
        can_buy: False 说明已经被卖
        has_lock：True 说明麦座加了锁
        is_init: True 是否初始化
        start_index, end_index 初始化的时候设置索引,非初始化，无用
        """
        import cn2an
        from caches import get_pika_redis
        lock_ticket_id = None
        unlock_ticket_id = None
        inst_id = None
        box_no_special = stand_name
        vl = VenuesLayers.objects.filter(venue_id=venue_id, name=floor_name).first()
        if vl:
            layers = vl.layer
        else:
            layers = cn2an.cn2an(floor_name[:-1])
        column = int(column)
        update_list = []
        result = '{}_{}_{}'.format(ticket_ids, int(not can_buy), int(has_lock))
        inst = None
        name, key = SessionSeat.get_pika_mz_seat_key(session_id, layers, row, column, box_no_special)
        if is_init:
            # inst = cls.objects.filter(session_id=session_id, layers=layers, showRow=row, showCol=column).first()
            if not seat_dict:
                seat_dict = dict()
            inst = seat_dict.get(key)
            if inst:
                from caches import get_pika_redis
                with get_pika_redis() as pika:
                    # name, key = SessionSeat.get_pika_mz_seat_key(session_id, layers, row, column, box_no_special)
                    pika.hset(name, key, result)
                inst.ticket_ids = ticket_ids
                inst.is_reserve = has_lock
                inst.end_index = end_index
                inst.start_index = start_index
                if has_lock:
                    inst.desc = '麦座锁定'
                else:
                    inst.desc = None
                update_list.append(inst)
                if update_list:
                    cls.objects.bulk_update(update_list,
                                            ['ticket_ids', 'is_reserve', 'desc', 'end_index', 'start_index'])
        else:
            with get_pika_redis() as pika:
                # name, key = SessionSeat.get_pika_mz_seat_key(session_id, layers, row, column, box_no_special)
                val = pika.hget(name, key)
                if val and result != val:
                    inst = cls.objects.filter(ticket_ids=ticket_ids, ticket_level__session_id=session_id).first()
                    if not inst:
                        log.error('{},{}'.format(session_id, box_no_special))
                    if inst:
                        pika.hset(name, key, result)
                        # 外面方法确保了座位数量没有发生变化
                        inst.matrix_seat_change_pika(matrix_val)
                        if not inst.push_mz_lock:
                            # 更新result,不是这边推送的锁座才需要更新锁座状态
                            # 不是这边推送的锁座，根据那边的锁座状态，锁座
                            inst.mz_set_reserve(has_lock)
                        else:
                            # 麦座已锁,且可售，这边可售，且push_mz_lock为True时表示这边推的锁座，才会推解锁
                            if has_lock and inst.can_buy() and can_buy:
                                unlock_ticket_id = inst.ticket_ids
        if inst:
            lock_ticket_id = inst.mz_set_buy(not can_buy, has_lock)
            # set_not_buy true的话是这边推的锁座
            is_buy = not can_buy or inst.is_buy
            can_buy_y = (not is_buy) and (unlock_ticket_id or not has_lock)
            # is_reserve是否锁定根据麦座来解
            inst.mz_change_and_set_pika(is_buy=is_buy, can_buy=can_buy_y, is_reserve=has_lock,
                                        unlock_ticket_id=unlock_ticket_id)
            inst_id = inst.id
        if not can_buy:
            # 两边同时已售的时候，标注
            with get_pika_redis() as pika:
                name, key = SessionSeat.get_pika_buy_key(session_id, ticket_ids)
                buy = pika.hget(name, key) or 0
                if int(buy) == 1:
                    inst = cls.objects.filter(session_id=session_id, ticket_ids=ticket_ids).first()
                    if inst:
                        TicketUserCode.check_refund_mz(inst.id)
                    pika.hset(name, key, 2)
        return lock_ticket_id, unlock_ticket_id, inst_id

    def can_buy(self):
        return (not self.is_reserve) and (not self.is_buy)

    def change_pika_redis(self, is_buy, can_buy, order_no=None):
        from caches import get_pika_redis, pika_session_seat_list_key, pika_session_seat_key, pika_level_seat_key
        pika = get_pika_redis()
        pika_session_seat_list_key = pika_session_seat_list_key.format(self.session_id)
        session_seat_key = pika_session_seat_key.format(self.session_id)
        level_seat_key = pika_level_seat_key.format(self.ticket_level_id, self.seats_id)
        seat = pika.hget(session_seat_key, level_seat_key)
        seat = json.loads(seat)
        seat['is_buy'] = is_buy
        seat['can_buy'] = can_buy
        seat['order_no'] = order_no
        # pika.hdel(session_seat_key, level_seat_key)
        pika.hset(session_seat_key, level_seat_key, json.dumps(seat))
        pika_session_seat_list = pika.get(pika_session_seat_list_key)
        pika_session_seat_list = json.loads(pika_session_seat_list)
        pika_session_seat_list[seat['index']]['is_buy'] = is_buy
        pika_session_seat_list[seat['index']]['can_buy'] = can_buy
        pika_session_seat_list[seat['index']]['order_no'] = order_no
        pika.set(pika_session_seat_list_key, json.dumps(pika_session_seat_list))

    def mz_change_and_set_pika(self, is_buy, can_buy, is_reserve, unlock_ticket_id=None):
        from caches import get_pika_redis, pika_session_seat_list_key, pika_session_seat_key, pika_level_seat_key
        with get_pika_redis() as pika:
            pika_session_seat_list_key = pika_session_seat_list_key.format(self.session_id)
            session_seat_key = pika_session_seat_key.format(self.session_id)
            level_seat_key = pika_level_seat_key.format(self.ticket_level_id, self.seats_id)
            seat = pika.hget(session_seat_key, level_seat_key)
            seat = json.loads(seat)
            seat['is_buy'] = is_buy
            seat['can_buy'] = can_buy
            seat['is_reserve'] = is_reserve
            # pika.hdel(session_seat_key, level_seat_key)
            pika.hset(session_seat_key, level_seat_key, json.dumps(seat))
            pika_session_seat_list = pika.get(pika_session_seat_list_key)
            pika_session_seat_list = json.loads(pika_session_seat_list)
            pika_session_seat_list[seat['index']]['is_buy'] = is_buy
            pika_session_seat_list[seat['index']]['can_buy'] = can_buy
            pika_session_seat_list[seat['index']]['is_reserve'] = is_reserve
            pika.set(pika_session_seat_list_key, json.dumps(pika_session_seat_list))
            if unlock_ticket_id:
                self.set_push_mz_lock(False)

    # @classmethod
    # def set_seat_price(cls):
    #     close_old_connections()
    #     from caches import get_redis, set_price_seat
    #     key = set_price_seat
    #     redis = get_redis()
    #     seat_list = redis.lrange(key, 0, -1)
    #     if seat_list:
    #         for val in seat_list:
    #             data = redis.rpop(key)
    #             dd = json.loads(data)
    #             cls.create_record(dd['data'], dd['cache_seat'])

    def set_log(self, request, msg):
        from django.contrib import admin
        log_inst = admin.ModelAdmin(SessionSeat, admin.site)
        log_inst.log_change(request, self, msg)

    @classmethod
    def create_record(cls, data: list, cache_seat, request=None):
        create_list = []
        update_list = []
        levels = dict()
        session = None
        from caches import get_pika_redis, pika_session_seat_list_key, pika_session_seat_key, pika_level_seat_key
        pika = get_pika_redis()
        list_dd = []
        i = 0
        session_seat_list_key = None
        is_delete = False
        for u in data:
            ticket_level_id = u['ticket_level_id']
            if not levels.get(str(ticket_level_id)):
                level = TicketFile.objects.filter(id=ticket_level_id).first()
                levels[str(ticket_level_id)] = level
                if not session:
                    session = level.session
                    if session.status != session.STATUS_OFF:
                        raise CustomAPIException('请先下架场次再修改,{}'.format(session.id))
                    if not is_delete:
                        session_seat_key = pika_session_seat_key.format(session.id)
                        pika.delete(session_seat_key)
                        is_delete = True
            else:
                level = levels[str(ticket_level_id)]
            session_seat_key = pika_session_seat_key.format(session.id)
            session_seat_list_key = pika_session_seat_list_key.format(session.id)
            seat_no = u['seat_no']
            seats = Seat.objects.filter(seat_no=seat_no).first()
            if seats:
                ss = None
                level_seat_key_old = None
                is_reserve = True if u.get('is_reserve') else False
                if session.is_price:
                    ss = cls.objects.filter(seats=seats, session_id=session.id).first()
                if ss:
                    is_update = False
                    desc = ''
                    if ss.ticket_level_id != int(ticket_level_id):
                        level_seat_key_old = pika_level_seat_key.format(ss.ticket_level_id, seats.id)
                        tf = TicketFile.objects.get(id=int(ticket_level_id))
                        desc = '{}修改价格:{},原价格：{}'.format(str(ss), tf.price, ss.ticket_level.price)
                        is_update = True
                        ss.ticket_level_id = int(ticket_level_id)
                        ss.color_id = level.color.id
                        ss.price = level.price
                        ss.color_code = level.color.code
                    if ss.is_reserve != is_reserve:
                        is_update = True
                        ss.is_reserve = is_reserve
                        if is_reserve:
                            desc = desc + ',{}锁座'.format(str(ss))
                        else:
                            desc = desc + ',{}销售座位'.format(str(ss))
                    if ss.box_no_special != u.get('box_no_special'):
                        is_update = True
                        ss.box_no_special = u.get('box_no_special')
                    if is_update:
                        if ss.is_buy:
                            raise CustomAPIException('座位已被购买，不能更改,{}'.format(str(ss)))
                        update_list.append(ss)
                    if desc:
                        ss.set_log(request, desc)
                else:
                    create_list.append(
                        cls(ticket_level_id=int(ticket_level_id), seats=seats, row=seats.row, column=seats.column,
                            layers=seats.layers, session_id=session.id, box_no_special=u.get('box_no_special'),
                            color_id=level.color.id, price=level.price, color_code=level.color.code,
                            is_reserve=is_reserve, showRow=u.get('showRow') or 0,
                            showCol=u.get('showCol') or 0,
                            desc=u.get('desc')))
                dd = dict(ticket_level=int(ticket_level_id), seats=seats.id, row=seats.row, column=seats.column,
                          layers=seats.layers, session_id=session.id,
                          color_id=level.color.id, price=float(level.price), color_code=level.color.code,
                          is_reserve=is_reserve, showRow=u.get('showRow') or 0,
                          showCol=u.get('showCol') or 0, box_no_special=u.get('box_no_special'),
                          desc=u.get('desc'), can_buy=ss.can_buy() if ss else not is_reserve)
                level_seat_key = pika_level_seat_key.format(ticket_level_id, seats.id)
                list_dd.append(dd)
                dd['index'] = i
                if level_seat_key_old:
                    pika.hdel(session_seat_key, level_seat_key_old)
                r = pika.hset(session_seat_key, level_seat_key, json.dumps(dd))
                if r not in [1, 0]:
                    raise CustomAPIException('执行失败, pika错误')
                i += 1
        pika.set(session_seat_list_key, json.dumps(list_dd))
        if create_list:
            cls.objects.bulk_create(create_list)
            session.is_price = True
            session.cache_seat = cache_seat
            session.save(update_fields=['is_price', 'cache_seat'])
        if update_list:
            cls.objects.bulk_update(update_list,
                                    ['ticket_level_id', 'is_reserve', 'box_no_special', 'color_id', 'price',
                                     'color_code'])
            session.cache_seat = cache_seat
            session.save(update_fields=['cache_seat'])

    def set_buy(self):
        # 锁定位置
        self.refresh_from_db(fields=['is_buy', 'version'])
        if SessionSeat.objects.filter(version=self.version, pk=self.pk, is_buy=False).update(version=F('version') + 1,
                                                                                             is_buy=True) <= 0:
            raise CustomAPIException('座位已被占用，请重新选座')
        else:
            self.change_pika_mz_seat(True)

    def mz_set_reserve(self, has_lock):
        fields = ['is_reserve', 'desc']
        self.is_reserve = has_lock
        if has_lock:
            self.desc = '麦座锁定'
        else:
            self.desc = None
        self.save(update_fields=fields)

    def set_push_mz_lock(self, status: bool):
        self.push_mz_lock = status
        self.save(update_fields=['push_mz_lock'])

    buy_desc_content = '麦座已售'

    def mz_set_buy(self, is_buy, has_lock):
        # 同步麦座
        ticket_ids = None
        if is_buy:
            # 麦座已售
            qs = SessionSeat.objects.filter(version=self.version, pk=self.pk, is_buy=False)
            if self.is_buy or qs.update(version=F('version') + 1, is_buy=True, buy_desc=self.buy_desc_content) <= 0:
                # 这边也已售
                qs.update(buy_desc='{},小程序也已售，需退款'.format(self.buy_desc_content))
                TicketUserCode.check_refund_mz(self.id)
        else:
            # 麦座未售且未锁
            if self.is_buy and not has_lock:
                if self.buy_desc != self.buy_desc_content:
                    # 不是麦座同步的已售,推送锁定
                    self.set_push_mz_lock(True)
                    ticket_ids = self.ticket_ids
                else:
                    # 麦座推的已售，这里解锁
                    self.is_buy = False
                    self.buy_desc = None
                    self.save(update_fields=['is_buy', 'buy_desc'])
        return ticket_ids

    @classmethod
    def get_order_amount(cls, user, session, ticket_list: list, pay_type, is_tiktok=False, express_fee=0):
        from caches import get_redis, seat_lock, session_seat_key
        redis = get_redis()
        multiply = 0
        amount = 0
        actual_amount = 0
        ticket = dict()
        discount_type = TicketOrder.DISCOUNT_DEFAULT
        session_seat_list = []
        tc_card = TheaterCardUserRecord.objects.filter(user=user).first()
        card = None
        use_old_card = False
        if pay_type == Receipt.PAY_CARD_JC and session.show.show_type == ShowType.dkxj():
            old_card_qs = TheaterCardUserDetail.get_old_cards(user.id)
            old_card_detail = old_card_qs.first()
            if old_card_detail:
                card_amount = 0
                for data in ticket_list:
                    inst = data.get('seat')
                    if inst:
                        ticket_level = inst.ticket_level
                        price, _ = ticket_level.get_price(old_card_detail.card, tc_card, pay_type, 1, True)
                        card_amount += price
                    else:
                        raise CustomAPIException('下单失败，请重新选择')
                if old_card_detail.amount > card_amount + express_fee:
                    use_old_card = True
                    card = old_card_detail.card
            log.debug(use_old_card)
        if not card:
            card = TheaterCard.get_inst()
        for data in ticket_list:
            kk = session_seat_key.format(data['level_id'], data['seat_id'])
            if redis.setnx(kk, 1):
                # 5分钟自动过期
                redis.expire(kk, 60)
            else:
                raise CustomAPIException('座位已被占用，请重新选座')
            # inst = cls.objects.filter(ticket_level_id=data['level_id'], seats_id=data['seat_id']).first()
            inst = data.get('seat')
            if inst:
                ticket_level = inst.ticket_level
                if is_tiktok and ticket_level.push_status != TicketFile.PUSH_SUCCESS:
                    raise CustomAPIException('该商品未上架{}'.format(ticket_level.price))
                if inst.session_id == session.id and inst.can_buy() and not inst.order_no:
                    if ticket.get(str(data['level_id'])):
                        ticket[str(data['level_id'])] += 1
                    else:
                        ticket[str(data['level_id'])] = 1
                    if ticket_level.limit_num > 0 and ticket[str(data['level_id'])] > ticket_level.limit_num:
                        raise CustomAPIException(
                            '价格{}的票,超过限购数量{}，请重新选座'.format(ticket_level.price, ticket_level.limit_num))
                    key = seat_lock.format(inst.id)
                    if redis.setnx(key, 1):
                        try:
                            # 5分钟自动过期
                            redis.expire(key, 300)
                            inst.set_buy()
                            amount += ticket_level.price
                            price, discount_theater = ticket_level.get_price(card, tc_card, pay_type, 1, use_old_card)
                            if discount_theater:
                                discount_type = TicketOrder.DISCOUNT_THEATER
                            actual_amount += price
                            multiply += 1
                            session_seat_list.append(inst)
                        finally:
                            redis.delete(key)
                    else:
                        raise CustomAPIException('座位{}已被占用，请重新选座'.format(str(inst)))
                else:
                    raise CustomAPIException('座位{}已被占用，请重新选座'.format(str(inst)))
            else:
                raise CustomAPIException('座位选择错误，请重新选座，')
        if pay_type in [Receipt.PAY_WeiXin_LP, Receipt.PAY_KS]:
            account = user.account
            discount = account.get_discount()
            actual_amount = amount * account.get_discount()
            if discount < 1:
                discount_type = TicketOrder.DISCOUNT_YEAR
        from common.utils import quantize
        return multiply, amount, quantize(actual_amount, 2), session_seat_list, discount_type


class TicketOrder(models.Model):
    user = models.ForeignKey(User, verbose_name='用户', on_delete=models.SET_NULL, null=True, blank=True)
    agent = models.ForeignKey(User, verbose_name='推荐人(代理)', on_delete=models.SET_NULL, null=True, blank=True,
                              related_name='agent')
    u_user_id = models.IntegerField('用户id', default=0)
    u_agent_id = models.IntegerField('代理id', default=0)
    title = models.CharField('节目名称', max_length=60)
    session = models.ForeignKey(SessionInfo, verbose_name=u'场次', null=True, on_delete=models.CASCADE)
    venue = models.ForeignKey(Venues, verbose_name='场馆', on_delete=models.CASCADE, null=True)
    mobile = models.CharField('手机号', max_length=20, null=True)
    express_address = models.CharField('送货地址', max_length=200, null=True, blank=True)
    order_no = models.CharField(u'订单号', max_length=128, unique=True, default=randomstrwithdatetime, db_index=True)
    tiktok_order_id = models.CharField(u'抖音订单号', max_length=128, unique=True, db_index=True, null=True, blank=True)
    ks_order_no = models.CharField(u'快手订单号', max_length=128, unique=True, db_index=True, null=True, blank=True)
    multiply = models.IntegerField(u'数量', validators=[validate_positive_int_gen])
    amount = models.DecimalField(u'订单总价', max_digits=13, decimal_places=2)
    actual_amount = models.DecimalField('实付金额', max_digits=9, decimal_places=2, default=0)
    discount_amount = models.DecimalField('折扣金额', max_digits=9, decimal_places=2, default=0, editable=False)
    express_fee = models.DecimalField('邮费', max_digits=9, decimal_places=2, default=0)
    card_jc_amount = models.DecimalField('剧场会员卡支付数额', max_digits=9, decimal_places=2, default=0)
    DISCOUNT_DEFAULT = 0
    DISCOUNT_YEAR = 1
    DISCOUNT_THEATER = 2
    DISCOUNT_CHOICES = (
        (DISCOUNT_DEFAULT, U'无'), (DISCOUNT_YEAR, '年度会员卡'), (DISCOUNT_THEATER, '剧场会员卡'))
    discount_type = models.IntegerField(u'优惠会员卡类型', choices=DISCOUNT_CHOICES, default=DISCOUNT_DEFAULT)
    refund_amount = models.DecimalField('已退款金额', max_digits=9, decimal_places=2, default=0)
    TY_HAS_SEAT = 1
    TY_NO_SEAT = 2
    TY_MARGIN = 3
    TY_CHOICES = ((TY_HAS_SEAT, u'有座订单'), (TY_NO_SEAT, '无座订单'), (TY_MARGIN, '补差订单'))
    order_type = models.SmallIntegerField(u'订单类型', choices=TY_CHOICES, default=TY_HAS_SEAT)
    SR_DEFAULT = 1
    SR_CY = 2
    SR_CHOICES = ((SR_DEFAULT, U'自建'), (SR_CY, '彩艺云'))
    source_type = models.SmallIntegerField(u'渠道类型', choices=SR_CHOICES, default=SR_DEFAULT)
    STATUS_UNPAID = 1
    STATUS_PAID = 2
    STATUS_CANCELED = 3
    STATUS_FINISH = 4
    STATUS_REFUNDING = 5
    STATUS_REFUNDED = 6
    STATUS_OVER_TIME = 7
    STATUS_REFUNDED_FAIL = 8
    STATUS_CHOICES = ((STATUS_UNPAID, u'未付款'), (STATUS_PAID, '待核销'), (STATUS_FINISH, '已完成'), (STATUS_CANCELED, '已取消'),
                      (STATUS_REFUNDING, '退款中'), (STATUS_REFUNDED, '已退款'), (STATUS_REFUNDED_FAIL, '退款失败'),
                      (STATUS_OVER_TIME, '付款超时(请联系客服)'))
    status = models.IntegerField(u'状态', choices=STATUS_CHOICES, default=STATUS_UNPAID)
    EXPRESS_DEFAULT = 1
    EXPRESS_UNSENT = 2
    EXPRESS_DELIVER = 3
    EXPRESS_FINISH = 4
    EXPRESS_CHOICES = (
        (EXPRESS_DEFAULT, U'无需发货'), (EXPRESS_UNSENT, U'待发货'), (EXPRESS_DELIVER, '待收货'), (EXPRESS_FINISH, '已完成'))
    express_status = models.IntegerField(u'发货状态', choices=EXPRESS_CHOICES, default=EXPRESS_DEFAULT)
    over_express_time = models.BooleanField('是否已过邮寄时间', default=False)
    SOURCE_DEFAULT = 0
    SOURCE_VIDEO = 1
    SOURCE_LIVE = 2
    SOURCE_NO = 3
    SOURCE_CHOICES = (
        (SOURCE_DEFAULT, u'无'), (SOURCE_VIDEO, '短视频'), (SOURCE_LIVE, '直播间'), (SOURCE_NO, '非cps订单'))
    source_type = models.IntegerField(u'带货场景', choices=SOURCE_CHOICES, default=SOURCE_DEFAULT)
    tiktok_nickname = models.CharField('达人抖音昵称', max_length=50, null=True, blank=True)
    tiktok_douyinid = models.CharField('达人抖音/快手号', max_length=50, null=True, blank=True)
    tiktok_commission_amount = models.DecimalField('达人佣金', max_digits=9, decimal_places=2, default=0)
    plan_id = models.CharField('计划ID', null=True, blank=True, max_length=50)
    is_cancel_pay = models.BooleanField('是否订单自动取消后付款', default=False)
    REFUND_TYPE_DEFAULT = 0
    REFUND_TYPE_USER = 1
    REFUND_TYPE_API = 2
    REFUND_TYPE_DEADLINE = 3
    REFUND_TYPE_KF = 4
    REFUND_TYPE_APPOINT = 5
    REFUND_TYPE_CHOICES = (
        (REFUND_TYPE_DEFAULT, u'无'), (REFUND_TYPE_USER, '用户发起退款'), (REFUND_TYPE_API, '开放者发起退款'),
        (REFUND_TYPE_DEADLINE, '过期自动退款'), (REFUND_TYPE_KF, '抖音客服退款'), (REFUND_TYPE_APPOINT, '预约失败自动退款'))
    tiktok_refund_type = models.IntegerField(u'抖音退款类型', choices=REFUND_TYPE_CHOICES, default=REFUND_TYPE_DEFAULT)
    status_before_refund = models.PositiveSmallIntegerField('退款前状态', default=STATUS_UNPAID,
                                                            choices=[(STATUS_PAID, u'待核销'), (STATUS_REFUNDED, '已退款')],
                                                            editable=False)
    receipt = models.OneToOneField(Receipt, verbose_name=u'收款信息', null=True, blank=True, on_delete=models.SET_NULL,
                                   related_name='ticket_order')
    pay_type = models.SmallIntegerField('付款类型', choices=Receipt.PAY_CHOICES, default=Receipt.PAY_NOT_SET)
    wx_pay_config = models.ForeignKey(WeiXinPayConfig, verbose_name='微信支付', blank=True, null=True,
                                      on_delete=models.SET_NULL)
    dy_pay_config = models.ForeignKey(DouYinPayConfig, verbose_name='抖音支付商户', null=True, blank=True,
                                      on_delete=models.SET_NULL)
    pay_at = models.DateTimeField(u'支付时间', null=True, blank=True)
    deliver_at = models.DateTimeField(u'发货时间', null=True, blank=True)
    create_at = models.DateTimeField(u'下单时间', auto_now_add=True)
    start_at = models.DateTimeField(u'场次开场时间', null=True, blank=True)
    end_at = models.DateTimeField('场次结束时间', null=True, blank=True)
    snapshot = models.TextField('场次演出快照', null=True, blank=True, help_text='下单时保存的快照', editable=False, max_length=2048)
    item_order_info_list = models.TextField('抖音下单明细', null=True, blank=True)
    push_message = models.BooleanField('是否已推送开场2小时前短信', default=False)
    push_message_day = models.BooleanField('是否已推送开场24小时前短信', default=False)
    is_check_num = models.IntegerField('已核验数量', default=0)
    is_low_buy = models.BooleanField('是否0.01购票', default=False)
    is_comment = models.BooleanField('是否已经评论', default=False)
    is_paper = models.BooleanField('是否纸质票场次', default=False)
    is_lock_seat = models.BooleanField('是否已手动出票', default=False)
    express_no = models.CharField(u'快递单号', max_length=30, null=True, blank=True,
                                  help_text=u'物流单号')
    express_name = models.CharField(u'快递公司', max_length=30, null=True, blank=True, help_text=u'物流公司名称')
    express_comp_no = models.CharField(u'快递公司编码', max_length=30, null=True, blank=True, help_text=u'物流公司名称')
    CHECK_DEFAULT = 0
    CHECK_SUCCESS = 1
    CHECK_FAIL = 2
    CHECK_CHOICES = (
        (CHECK_DEFAULT, ''), (CHECK_SUCCESS, '成功'), (CHECK_FAIL, '失败'))
    auto_check = models.IntegerField(u'抖音/小红书自动核销状态', choices=CHECK_CHOICES, default=CHECK_DEFAULT)
    booking_at = models.DateTimeField('补单时间', null=True, blank=True, editable=False)
    need_refund_mz = models.BooleanField('同步卖座是否需要退款', default=False)
    source_order = models.ForeignKey('self', verbose_name='来源订单', null=True, blank=True, on_delete=models.SET_NULL,
                                     related_name='sr_order')
    transaction_id = models.CharField('微信(抖音)支付单号', max_length=32, null=True, blank=True)
    ks_report = models.IntegerField(u'快手推送核销状态', choices=CHECK_CHOICES, default=CHECK_DEFAULT)

    def __str__(self):
        return self.order_no

    class Meta:
        verbose_name_plural = verbose_name = '订单'
        ordering = ['-pk']

    @classmethod
    def export_fields(cls):
        return [u'下单用户', '姓名', '联系电话', '详细地址', '推荐人', '付款类型', '微信/抖音商户', '演出场次座位', '票档描述', '订单号',
                '商户订单号', '微信(抖音)支付单号', '数量', '订单总价', '剧场会员卡支付数额', '实际支付金额', '邮费', '状态', '演出名称',
                '下单时间', '支付时间', '开演时间', '带货场景', '达人抖音昵称', '达人抖音号', '计划ID', '计划类型', '演出场馆']

    @classmethod
    def export_express_fields(cls):
        return ['下单用户', '姓名', '联系电话', '详细地址', '演出场次座位', '票档描述', '订单号', '商户订单号', '数量', '订单总价', '剧场会员卡支付数额',
                '实际支付金额', '状态', '演出名称', '下单时间', '支付时间', '开演时间', '演出场馆', '快递公司', '快递单号', '快递公司编码']

    @classmethod
    def can_refund_status(cls):
        return [cls.STATUS_PAID, cls.STATUS_REFUNDED_FAIL]

    def auto_refund(self):
        refund_amount = self.actual_amount
        st, msg_or_obj = TicketOrderRefund.create_record(self, refund_amount, '下单确认失败自动退款',
                                                         TicketOrderRefund.ST_WX)
        if st:
            self.refund_amount += refund_amount
            self.status_before_refund = self.status
            self.status = self.STATUS_REFUNDING
            self.save(update_fields=['refund_amount', 'status', 'status_before_refund'])
            try:
                re_st, msg = msg_or_obj.set_confirm()
                if not re_st:
                    msg_or_obj.error_msg = msg
            except Exception as e:
                msg_or_obj.error_msg = str(e)
            msg_or_obj.save(update_fields=['error_msg'])
        else:
            log.error(msg_or_obj)

    def cps_stl(self, num):
        from statistical.models import TotalStatistical
        dy_order_num = 0
        dy_live_order_num = 0
        dy_video_order_num = 0
        if self.source_type == self.SOURCE_VIDEO:
            dy_video_order_num = num
        elif self.source_type == self.SOURCE_LIVE:
            dy_live_order_num = num
        elif self.source_type == self.SOURCE_NO:
            dy_order_num = num
        TotalStatistical.change_ticket_order_stl(dy_live_order_num=dy_live_order_num,
                                                 dy_video_order_num=dy_video_order_num, dy_order_num=dy_order_num)

    def add_stl(self, is_refund=False):
        from statistical.models import DayStatistical, CityStatistical, TotalStatistical, MonthSales
        dy_amount = 0
        wx_amount = 0
        wx_order_num = 0
        if self.pay_type == Receipt.PAY_TikTok_LP:
            dy_amount = -self.actual_amount if is_refund else self.actual_amount
        else:
            wx_amount = -self.actual_amount if is_refund else self.actual_amount
            wx_order_num = -1 if is_refund else 1
        TotalStatistical.change_ticket_order_stl(dy_amount=dy_amount, wx_amount=wx_amount,
                                                 wx_order_num=wx_order_num)
        num = 1
        amount = self.actual_amount
        if is_refund:
            num = -1
            amount = -self.actual_amount
        DayStatistical.change_order_sum(self.pay_at, num, amount)
        CityStatistical.change_order_sum(self.venue.city, num, amount)
        MonthSales.change_order_sum(self.pay_at, num, amount)

    @classmethod
    def can_express_status(cls):
        return [cls.EXPRESS_UNSENT, cls.EXPRESS_DELIVER]

    def set_express_finish(self):
        self.express_status = self.EXPRESS_FINISH
        self.save(update_fields=['express_status'])

    def set_lock_seats(self, st):
        self.is_lock_seat = st
        self.save(update_fields=['is_lock_seat'])

    def query_dy_status(self):
        from douyin import get_dou_yin
        dy = get_dou_yin()
        ret = dy.query_item_order_info(self.tiktok_order_id)
        status = ret['item_list'][0]['item_order_status']
        return DY_STATUS.get(str(status)) or '未知状态'

    def before_pay_check(self):
        order = self
        now = timezone.now()
        from mp.models import BasicConfig
        bc = BasicConfig.get()
        auto_cancel_minutes = bc.auto_cancel_minutes if bc else 10
        expire_at = now + timedelta(minutes=-auto_cancel_minutes)
        if not order.session.can_buy:
            raise CustomAPIException('该场次已停止购买')
        if not order.session.show.can_buy:
            raise CustomAPIException('演出已停止购买')
        receipt = order.receipt
        receipt.query_status(self.order_no)
        if receipt.paid:
            raise CustomAPIException('该订单已经付款，请尝试刷新订单页面')
        if order.status != order.STATUS_UNPAID:
            raise CustomAPIException('订单状态错误')
        if order.create_at < expire_at:
            order.cancel()
            raise CustomAPIException('该订单已经过期，请重新下单')

    def cancel_lock_seats(self):
        qs = SessionSeat.objects.filter(order_no=self.order_no)
        code_qs = TicketUserCode.objects.filter(order_id=self.id)
        from caches import with_redis, session_seat_key
        with with_redis() as redis:
            for inst in qs:
                inst.change_pika_redis(is_buy=False, can_buy=(not inst.is_reserve), order_no=None)
                inst.change_pika_mz_seat(False)
                inst.set_pika_buy(False)
                key = session_seat_key.format(inst.ticket_level_id, inst.seats_id)
                # 下单的锁
                redis.delete(key)
        qs.update(order_no=None, is_buy=False, buy_desc=None)
        code_qs.update(session_seat=None)
        self.set_lock_seats(False)

    @classmethod
    def get_or_set_real_name_buy_num(cls, session_id: int, id_card: str, num: int, is_get=True):
        from caches import get_pika_redis, real_name_buy_session_key, real_name_buy_id_card_key
        with get_pika_redis() as redis:
            real_name_buy_session_key = real_name_buy_session_key.format(session_id)
            real_name_buy_id_card_key = real_name_buy_id_card_key.format(id_card)
            if is_get:
                buy_num = redis.hget(real_name_buy_session_key, real_name_buy_id_card_key)
            if num != 0:
                buy_num = redis.hincrby(real_name_buy_session_key, real_name_buy_id_card_key, num)
        return int(buy_num) if buy_num else 0

    @property
    def show_express_address(self):
        """

        :return:
        """
        return self.express_address.replace('#', '') if self.express_address else ''

    def change_agent_c_amount(self, c_amount):
        if self.session and self.agent and self.pay_at:
            from statistical.models import SessionAgentDaySum
            if self.source_type == self.SOURCE_DEFAULT:
                s_type = SessionAgentDaySum.ST_WX
            else:
                s_type = self.source_type
            SessionAgentDaySum.change_record(self.session, self.agent, self.pay_at, s_type, c_amount=c_amount)

    def change_agent_amount(self, is_refund=False):
        if self.session and self.agent and self.pay_at:
            from statistical.models import SessionAgentDaySum
            if self.source_type == self.SOURCE_DEFAULT:
                s_type = SessionAgentDaySum.ST_WX
            else:
                s_type = self.source_type
            amount = -self.refund_amount if is_refund else self.award_amount
            SessionAgentDaySum.change_record(self.session, self.agent, self.pay_at, s_type, amount=amount)

    def change_cps_agent_amount(self, source_type, c_amount, amount, platform):
        from statistical.models import SessionCpsDaySum
        SessionCpsDaySum.change_cps_record(self.session, self.tiktok_douyinid, self.tiktok_nickname, self.pay_at,
                                           source_type,
                                           platform=platform,
                                           c_amount=c_amount, amount=amount)

    @classmethod
    def check_cps_source_new(cls):
        close_old_connections()
        from douyin import get_dou_yin
        from caches import get_redis, check_cps_source_key
        dy = get_dou_yin()
        now = timezone.now()
        start_at = now - timedelta(days=2)
        redis = get_redis()
        order_qs = cls.objects.filter(source_type=TicketOrder.SOURCE_DEFAULT, tiktok_order_id__isnull=False,
                                      status__in=[cls.STATUS_PAID, cls.STATUS_FINISH],
                                      pay_type=Receipt.PAY_TikTok_LP, pay_at__lt=now, pay_at__gt=start_at)
        for order in order_qs:
            key = check_cps_source_key.format(order.id)
            try:
                if redis.setnx(key, order.id):
                    # 避免同一订单同时执行
                    redis.expire(key, 5 * 60)
                    st, ret = dy.query_cps(order.order_no)
                    if st and ret.get('data') and ret['data'].get('cps_info') and \
                            ret['data']['cps_info'][
                                'cps_item_list']:
                        source_type = ret['data']['cps_info']['cps_item_list'][0]['source_type']
                        order.tiktok_douyinid = ret['data']['cps_info']['cps_item_list'][0][
                            'commission_user_douyinid']
                        order.tiktok_nickname = ret['data']['cps_info']['cps_item_list'][0][
                            'commission_user_nickname']
                        order.plan_id = ret['data']['cps_info']['cps_item_list'][0]['task_id']
                        order.source_type = source_type
                        c_amount = ret['data']['cps_info']['total_commission_amount'] / 100
                        order.tiktok_commission_amount = c_amount
                        # CPS订单不发放任何奖励，并且不记录最新推荐人在订单中
                        # order.agent = None
                        order.save(update_fields=['source_type', 'tiktok_douyinid', 'tiktok_nickname', 'plan_id',
                                                  'source_type', 'tiktok_commission_amount'])
                        order.change_cps_agent_amount(source_type, c_amount, order.actual_amount, TiktokUser.ST_DY)
                    else:
                        # 非CPS订单或带货场景为空的订单按照目前的逻辑正常计算
                        source_type = TicketOrder.SOURCE_NO
                        order.source_type = TicketOrder.SOURCE_NO
                        order.save(update_fields=['source_type'])
                        # 订单发奖励
                        order.send_award()
                        # 补差订单发奖励
                        order_qs = TicketOrder.objects.filter(source_order_id=order.id)
                        for od in order_qs:
                            od.send_award()
                    TicketUserCode.objects.filter(order_id=order.id).update(source_type=source_type)
            except Exception as e:
                continue
            finally:
                redis.delete(key)

    @classmethod
    def check_add_booking(cls):
        order_list = cls.objects.filter(auto_check=cls.CHECK_FAIL, status=cls.STATUS_FINISH)
        for order in order_list:
            order.add_booking()

    def add_booking(self):
        tb = TicketBooking.objects.filter(order_id=self.id, status=TicketBooking.STATUS_SUCCESS).first()
        data = None
        if tb:
            snapshot = json.loads(self.snapshot)
            book_info = []
            booking = None
            for price_list in snapshot['price_list']:
                multiply = int(price_list['multiply'])
                ticket_level = TicketFile.objects.get(id=int(price_list['level_id']))
                goods_id = ticket_level.get_product_id()
                if goods_id:
                    booking_item = TicketBookingItem.objects.filter(booking_id=tb, goods_id=goods_id).first()
                    item_num = TicketBookingItem.objects.filter(booking_id=tb, goods_id=goods_id).count()
                    if item_num < multiply:
                        if not booking:
                            booking = TicketBooking.objects.create(user=tb.user, order=tb.order, order_no=tb.order_no,
                                                                   tiktok_order_id=tb.tiktok_order_id)
                        need_num = multiply - item_num
                        item_data = dict(booking=booking, goods_id=booking_item.goods_id,
                                         ext_shopid=booking_item.ext_shopid,
                                         poi_id=booking_item.poi_id,
                                         shop_name=booking_item.shop_name,
                                         book_start_time=booking_item.book_start_time,
                                         book_end_time=booking_item.book_end_time, sku_info=booking_item.sku_info,
                                         user_info=booking_item.user_info)
                        dd = {
                            "bookStartTime": get_timestamp(booking_item.book_start_time),
                            "bookEndTime": get_timestamp(booking_item.book_end_time),
                            "bookRangeType": 0,  # 非日历房不传，或者传0
                            "outShopId": str(booking_item.ext_shopid),
                            "goodsId": booking_item.goods_id,
                            "poiId": booking_item.poi_id,
                            "shopName": booking_item.shop_name,
                            "userInfo": json.loads(booking_item.user_info),
                            "skuInfo": json.loads(booking_item.sku_info)
                        }
                        for i in list(range(need_num)):
                            book_info.append(dd)
                            TicketBookingItem.objects.create(**item_data)
            if booking:
                data = {
                    "orderId": booking.tiktok_order_id,
                    "outBookNo": booking.out_book_no,
                    "bookInfo": book_info,
                }
        if (tb and data) or not tb:
            st = TicketBooking.create_book(self, data)
            if st:
                self.auto_check = self.CHECK_DEFAULT
                self.booking_at = timezone.now()
                self.save(update_fields=['auto_check', 'booking_at'])
                TicketUserCode.objects.filter(order_id=self.id).update(tiktok_check=False)
        else:
            log.error(self.order_no)

    def check_refund_order(self):
        msg = ''
        if self.tiktok_order_id:
            source = {"1": '用户发起退款', "2": '开放者发起退款', "3": '自动退款', "4": '抖音客服退款', "5": '预约失败自动发起退款', "6": '开发者拒绝接单退款',
                      "7": '后约单触发先买单退款'}
            status = {"PROCESSING": '退款中', "SUCCESS": '已退款', "FAIL": '退款失败'}
            from douyin import get_dou_yin
            dy = get_dou_yin()
            try:
                ret = dy.query_refund_order(self.tiktok_order_id)
                for data in ret['refund_list']:
                    refund_source = source.get(str(data['refund_source']))
                    status_display = status.get(data['refund_status'])
                    msg += '|退款来源:{}，退款金额:{},状态:{}'.format(refund_source, data['refund_total_amount'] / 100,
                                                           status_display)
            except Exception as e:
                msg = '订单不存在退款'
        return msg

    @classmethod
    def scroll_key(cls):
        from caches import scroll_key
        return scroll_key

    @property
    def encrypt_mobile(self):
        return '{}***{}'.format(self.mobile[:3], self.mobile[-2:])

    def change_scroll_list(self, is_pay=False):
        scroll_key = TicketOrder.scroll_key()
        from caches import with_redis
        with with_redis() as redis:
            if is_pay:
                desc = '{}刚买了{}'.format(self.encrypt_mobile, self.title)
            else:
                desc = '{}正在买{}'.format(self.encrypt_mobile, self.title)
            redis.lpush(scroll_key, desc)
            if redis.llen(scroll_key) > 10:
                redis.rpop(scroll_key)

    @classmethod
    def scroll_list(cls):
        scroll_key = TicketOrder.scroll_key()
        from caches import with_redis
        with with_redis() as redis:
            scroll_list = redis.lrange(scroll_key, 0, -1) or []
        return scroll_list

    def change_actual_amount(self, after_amount):
        self.actual_amount = after_amount
        self.save(update_fields=['actual_amount'])
        self.receipt.amount = after_amount
        self.receipt.save(update_fields=['amount'])

    def set_need_refund_mz(self, status):
        self.need_refund_mz = status
        self.save(update_fields=['need_refund_mz'])

    def get_plan_name(self):
        ret = None
        if self.plan_id:
            if self.source_type == self.SOURCE_LIVE:
                if LiveRoomCps.objects.filter(plan_id=self.plan_id):
                    ret = '直播间定向佣金计划'
                else:
                    ret = '通用佣金计划'
            elif self.source_type == self.SOURCE_VIDEO:
                if ShortVideoCps.objects.filter(plan_id=self.plan_id):
                    ret = '短视频定向佣金计划'
                else:
                    ret = '通用佣金计划'
        return ret

    def set_tiktok_order_id(self, tiktok_order_id, item_order_info_list):
        self.tiktok_order_id = tiktok_order_id
        self.item_order_info_list = item_order_info_list
        self.save(update_fields=['tiktok_order_id', 'item_order_info_list'])

    def get_wx_pay_end_at(self):
        from datetime import timedelta
        from mp.models import BasicConfig
        bc = BasicConfig.get()
        cancel_minutes = bc.auto_cancel_minutes - 1 if bc.auto_cancel_minutes > 1 else bc.auto_cancel_minutes
        pay_end_at = self.create_at + timedelta(minutes=cancel_minutes)
        if hasattr(self, 'cy_order'):
            cy_pay_end_at = self.cy_order.auto_cancel_order_time - timedelta(seconds=30)
            if pay_end_at > cy_pay_end_at:
                pay_end_at = cy_pay_end_at
        return pay_end_at

    def get_end_at(self):
        # 转时间戳
        pay_end_at = self.get_wx_pay_end_at()
        from common.utils import get_timestamp
        end_at = get_timestamp(pay_end_at)
        return end_at

    def team_award(self, parent, lv, check_parent_list=None):
        is_award = False
        if not check_parent_list:
            check_parent_list = []
        if not parent or self.id == parent.id or parent in check_parent_list:
            # 发奖励的是自己或者已经发过奖励了直接不再发
            return is_award
        if parent and lv < 25:
            if parent.account.level:
                if self.session and self.session.show and self.session.show.show_type == ShowType.xunyan():
                    ratio = parent.account.level.team_ratio_xy
                else:
                    ratio = parent.account.level.team_ratio
                if ratio > 0:
                    # 代理分销奖
                    amount = self.award_amount * ratio / 100
                    if amount > 0.01:
                        from shopping_points.models import UserCommissionChangeRecord
                        UserCommissionChangeRecord.add_record(parent.account, amount,
                                                              UserCommissionChangeRecord.SOURCE_TYPE_GROUP,
                                                              '代理团购奖励,{}级'.format(lv),
                                                              status=UserCommissionChangeRecord.STATUS_CAN_WITHDRAW,
                                                              order=self)
                        is_award = True
                        try:
                            if parent.openid:
                                self.send_order_parent_notice(parent.openid, '团队奖{}'.format(amount))
                        except Exception as e:
                            log.error('发送消息失败')
            if parent.parent:
                check_parent_list.append(parent)
                self.team_award(parent.parent, lv + 1, check_parent_list)
        return is_award

    @classmethod
    def send_show_start_notice(cls):
        close_old_connections()
        hours = 2
        start_at = timezone.now() + timedelta(hours=hours)
        start_at_day = timezone.now() + timedelta(hours=24)
        from qcloud.sms import get_sms
        sms = get_sms()
        # log.error(start_at)
        two_qs = cls.objects.filter(push_message=False, session__start_at__lte=start_at, status=cls.STATUS_PAID)
        for order in two_qs:
            code = TicketUserCode.objects.filter(order_id=order.id, give_status=TicketUserCode.GIVE_DEFAULT)
            if code:
                try:
                    data = dict(name=order.title, mobile=order.mobile, number=2,
                                time=order.session.start_at.strftime("%Y-%m-%d %H:%M"))
                    sms.smsvrcode(data)
                except Exception as e:
                    log.error('发送短息消息失败')
            order.push_message = True
            order.push_message_day = True
            order.save(update_fields=['push_message', 'push_message_day'])
        day_qs = cls.objects.filter(push_message_day=False, session__start_at__lte=start_at_day, status=cls.STATUS_PAID)
        for order in day_qs:
            code = TicketUserCode.objects.filter(order_id=order.id, give_status=TicketUserCode.GIVE_DEFAULT)
            if code:
                try:
                    data = dict(name=order.title, mobile=order.mobile, number=24,
                                time=order.session.start_at.strftime("%Y-%m-%d %H:%M"))
                    sms.smsvrcode(data)
                except Exception as e:
                    log.error('发送短息消息失败')
            order.push_message_day = True
            order.save(update_fields=['push_message_day'])

    def wx_template_msg(self):
        url = 'pages/pagesKage/orderDetail/orderDetail?id={}'.format(self.id)
        address = '{} | {}'.format(self.venue.city.city, self.venue.name)
        start_at = self.session.start_at.strftime('%Y-%m%-d %H:%M')
        order_desc = '演出即将开始，请尽快前往观看~'
        from mp.wechat_client import get_wxa_client
        wxa = get_wxa_client()
        wxa.show_start_notice(self.user.lp_openid, address, self.title, start_at, order_desc, url)
        # mini_program = dict(id=self.id)
        # MpTemplateClient.show_start_notice(self.user.openid, address, self.title, start_at, order_desc, url,
        #                                    mini_program)

    def tiktok_template_msg(self):
        from douyin import get_tiktok
        client = get_tiktok()
        url = 'pages/pagesKage/orderDetail/orderDetail?id={}'.format(self.id)
        data = dict()
        data['项目名称'] = self.title
        data['演出地点'] = '{} | {}'.format(self.venue.city.city, self.venue.name)
        data['演出时间'] = self.session.start_at.strftime('%Y-%m%-d %H:%M')
        data['订单详情'] = '演出即将开始，请尽快前往观看~'
        client.send_notify(self.user.openid_tiktok, data, url)

    def send_order_parent_notice(self, openid, desc):
        from mp.wechat_client import get_wxa_client
        # wxa = get_wxa_client()
        # session_name = self.session.start_at.strftime('%Y年%月%日 %H:%M')
        # wxa.award_notice(lp_openid, self.title, session_name, self.actual_amount, desc, SALE_ORDER_URL)
        name = '{}({})'.format(self.user.get_full_name(), desc)
        MpTemplateClient.order_parent_notice(openid, self.order_no, self.title, self.actual_amount, name,
                                             SALE_ORDER_URL)

    @property
    def award_amount(self):
        return self.actual_amount - self.express_fee

    def share_award(self):
        is_award = False
        if self.agent and self.agent.account.level:
            if self.session and self.session.show and self.session.show.show_type == ShowType.xunyan():
                ratio = self.agent.account.level.share_ratio_xy
            else:
                ratio = self.agent.account.level.share_ratio
            if ratio > 0:
                # 代理分销奖
                amount = self.award_amount * ratio / 100
                if amount > 0.01:
                    from shopping_points.models import UserCommissionChangeRecord
                    UserCommissionChangeRecord.add_record(self.agent.account, amount,
                                                          UserCommissionChangeRecord.SOURCE_TYPE_SHARE_AWARD, '代理分销奖励',
                                                          status=UserCommissionChangeRecord.STATUS_CAN_WITHDRAW,
                                                          order=self)
                    is_award = True
                    try:
                        if self.agent.openid:
                            self.send_order_parent_notice(self.agent.openid, '分销奖{}'.format(amount))
                    except Exception as e:
                        log.error('发送消息失败')
        return is_award

    def set_paid(self):
        if self.status == self.STATUS_UNPAID:
            self.status = self.STATUS_PAID
            self.transaction_id = self.receipt.transaction_id if self.receipt else None
            self.pay_at = timezone.now()
            if self.session.is_paper and self.order_type in [self.TY_HAS_SEAT,
                                                             self.TY_NO_SEAT] and not self.over_express_time:
                # 纸质且未过邮寄时间
                self.express_status = TicketOrder.EXPRESS_UNSENT
            else:
                self.express_status = TicketOrder.EXPRESS_DEFAULT
            self.save(update_fields=['status', 'pay_at', 'transaction_id', 'express_status'])
            self.settle_order()
            # 下面注释的丢到任务做
            # self.do_order()
            if self.order_type == self.TY_HAS_SEAT:
                # 有座的
                self.set_code()
            elif self.order_type == self.TY_NO_SEAT:
                self.no_seat_set_code()
            elif self.order_type == self.TY_MARGIN:
                # 补差价订单退出付款页面就直接取消，付款后直接变为已完成；
                self.set_finish()
            if hasattr(self, 'cy_order'):
                from caiyicloud.tasks import async_confirm_order
                async_confirm_order.delay(self.id)
        elif self.status == self.STATUS_CANCELED:
            if self.receipt.status == Receipt.STATUS_FINISHED:
                self.is_cancel_pay = True
                self.pay_at = timezone.now()
                self.status = self.STATUS_OVER_TIME
                self.save(update_fields=['is_cancel_pay', 'status', 'pay_at'])

    def no_seat_set_code(self):
        # 无座的
        snapshot = json.loads(self.snapshot)
        price_list = snapshot['price_list']
        for ll in price_list:
            tf = TicketFile.objects.filter(id=int(ll['level_id']), session_id=self.session.id).first()
            if tf:
                tf.update_sales(int(ll['multiply']))
                for i in list(range(ll['multiply'])):
                    TicketUserCode.create_record(order=self, ticket_level=tf)

    @classmethod
    def award_status_list(cls):
        return [cls.STATUS_PAID, cls.STATUS_FINISH]

    @classmethod
    def send_award_task(cls):
        from caches import with_redis, settle_order_award_key
        with with_redis() as redis:
            order_list = redis.lrange(settle_order_award_key, 0, -1)
            if order_list:
                for order_id in order_list:
                    val = redis.rpop(settle_order_award_key)
                    if val:
                        val = int(val)
                        self = cls.objects.filter(id=val).first()
                        if self:
                            self.do_order()

    def do_order(self):
        if self.order_type == self.TY_MARGIN:
            if self.source_order and self.source_order.pay_type == Receipt.PAY_WeiXin_LP or \
                    self.source_order.source_type in [self.SOURCE_VIDEO, self.SOURCE_LIVE]:
                # 补差订单，来源订单是微信支付的则直接发奖励
                self.send_award()
        else:
            # 非补差订单，订单是微信支付的则直接发奖励
            if self.pay_type in [Receipt.PAY_WeiXin_LP]:
                self.send_award()
            if self.pay_type == Receipt.PAY_CARD_JC and self.card_jc_amount > 0:
                # 增加优惠金额
                user_card = TheaterCardUserRecord.get_inst(self.user)
                amount = self.amount - self.actual_amount
                if amount > 0:
                    user_card.add_discount_total(amount)
        self.change_scroll_list(is_pay=True)
        # 增加场次实收
        self.session.update_actual_amount(self.actual_amount)

    def settle_order(self):
        from caches import with_redis, settle_order_award_key
        with with_redis() as redis:
            redis.lpush(settle_order_award_key, self.id)

    def send_award(self):
        if self.status in TicketOrder.award_status_list():
            try:
                share = self.share_award()
                team = self.team_award(self.agent, lv=1)
                if share or team:
                    self.change_agent_amount()
            except Exception as e:
                log.error('发奖励失败,{}'.format(e))

    @atomic
    def set_code(self):
        # 付款后创建二维码
        session_seat_qs = SessionSeat.objects.filter(order_no=self.order_no)
        for session_seat in session_seat_qs:
            TicketUserCode.create_record(order=self, session_seat=session_seat)
            # 付款后要把这里重新制成已售
            session_seat.change_pika_mz_seat(True)
            session_seat.set_pika_buy(True)

    def get_snapshot(self, dd):
        import json
        """
        商品快照
        :return:
        """
        session = self.session
        from common.utils import get_config
        config = get_config()
        data = dict(show_name=self.title, show_id=session.show.id, venue_name=session.show.venues.name, price_list=dd,
                    logo='{}{}'.format(config['template_url'], session.show.logo_mobile.url),
                    start_at=session.start_at.strftime('%Y-%m-%d %H:%M'),
                    has_seat=session.has_seat,
                    end_at=session.end_at.strftime('%Y-%m-%d %H:%M'))
        return json.dumps(data)

    def cancel(self):
        """
        取消订单
        :return:
        """
        if self.status != self.STATUS_CANCELED:
            receipt = self.receipt
            if receipt.pay_type in (
                    Receipt.PAY_WeiXin_LP, Receipt.PAY_WeiXin_MP, Receipt.PAY_WeiXin_APP, Receipt.PAY_TikTok_LP):
                receipt.query_status(self.order_no)
            if receipt.paid:
                receipt.biz_paid()
                return False, '取消失败订单已付款'
            self.status = self.STATUS_CANCELED
            self.save(update_fields=['status'])
            if self.order_type in [TicketOrder.TY_NO_SEAT, TicketOrder.TY_HAS_SEAT]:
                self.release_seat(is_cancel=True)
            if self.card_jc_amount > 0 and self.pay_type == Receipt.PAY_CARD_JC:
                TheaterCardChangeRecord.add_record(user=self.user,
                                                   source_type=TheaterCardChangeRecord.SOURCE_TYPE_CANCEL,
                                                   amount=self.card_jc_amount, ticket_order=self)
            # 其他渠道取消，不调也无错误。会自动取消
            # if hasattr(self, 'cy_order'):
            #     self.cy_order.cancel_order()
        return True, ''

    def release_seat(self, is_cancel=False):
        if self.session.has_seat == SessionInfo.SEAT_NO:
            # 无座位的只需要退库存
            snapshot = json.loads(self.snapshot)
            price_list = snapshot['price_list']
            for ll in price_list:
                tf = TicketFile.objects.filter(id=int(ll['level_id']), session_id=self.session.id).first()
                if tf:
                    tf.change_stock(int(ll['multiply']))
                    if not is_cancel:
                        # 取消订单不用减销量
                        tf.update_sales(-int(ll['multiply']))
        else:
            qs = SessionSeat.objects.filter(order_no=self.order_no)
            from caches import get_redis, session_seat_key
            redis = get_redis()
            for inst in qs:
                key = session_seat_key.format(inst.ticket_level_id, inst.seats_id)
                # 下单的锁
                redis.delete(key)
                if not self.need_refund_mz:
                    inst.change_pika_redis(is_buy=False, can_buy=(not inst.is_reserve), order_no=None)
                    inst.change_pika_mz_seat(False)
                    inst.set_pika_buy(False)
            qs.update(order_no=None, is_buy=False, buy_desc=None)
        if self.id_card:
            # 减去已买的
            TicketOrder.get_or_set_real_name_buy_num(self.session.id, self.id_card, -self.multiply, is_get=False)

    def cancel_code(self):
        TicketUserCode.objects.filter(order_id=self.id).update(status=TicketUserCode.STATUS_CANCEL, msg='退款作废')

    @classmethod
    def auth_check_over_time_code(cls):
        # 微信小程序支付的自动核销
        close_old_connections()
        qs = cls.objects.filter(pay_type=Receipt.PAY_WeiXin_LP, status=cls.STATUS_PAID,
                                session__end_at__lt=timezone.now())
        for order in qs:
            tu_all = TicketUserCode.objects.filter(order_id=order.id, status=TicketUserCode.STATUS_DEFAULT)
            tu_all.update(status=TicketUserCode.STATUS_OVER_TIME)
            order.set_finish()

    @classmethod
    def auth_check_over_time_code_tiktok(cls):
        # 抖音小程序支付的自动核销
        from caches import with_redis, auth_check_code_tiktok_key
        with with_redis() as redis:
            if redis.setnx(auth_check_code_tiktok_key, 1):
                redis.expire(auth_check_code_tiktok_key, 5 * 60)
                try:
                    close_old_connections()
                    from douyin import get_dou_yin
                    dy = get_dou_yin()
                    qs = cls.objects.filter(pay_type=Receipt.PAY_TikTok_LP,
                                            session__end_at__lt=timezone.now() - timedelta(hours=25),
                                            status__in=[cls.STATUS_PAID, cls.STATUS_FINISH],
                                            auto_check=cls.CHECK_DEFAULT).filter(
                        models.Q(booking_at__isnull=True) | models.Q(
                            booking_at__lt=timezone.now() - timedelta(hours=28)))
                    for order in qs:
                        order.auto_check = cls.CHECK_FAIL
                        tu_check = TicketUserCode.objects.filter(order_id=order.id, status=TicketUserCode.STATUS_CHECK,
                                                                 tiktok_check=False)
                        tu_uncheck = TicketUserCode.objects.filter(order_id=order.id,
                                                                   status=TicketUserCode.STATUS_DEFAULT,
                                                                   tiktok_check=False)
                        if tu_check or tu_uncheck:
                            # 推抖音
                            tiktok_check = False
                            # order.item_order_info_list 取单号
                            # item_order_list = [{"item_order_id": inst.order.tiktok_order_id}]
                            poi_info = {"shop_name": order.session.tiktok_store.name,
                                        "ext_valid_shop_id": str(order.session.tiktok_store.id),
                                        "valid_poi_id_str": order.session.tiktok_store.supplier_ext_id}
                            params = dict(out_order_no=order.order_no, use_all=True, poi_info=json.dumps(poi_info))
                            try:
                                ret = dy.push_delivery_new(params=params)
                                if ret['error_code'] != 0:
                                    log.error('自动核销失败,{},{}'.format(ret['sub_description'], order.order_no))
                                    msg = ret['sub_description']
                                    # tu_list.update(msg='抖音自动核销失败')
                                else:
                                    # tu_list.update(msg='抖音自动核销成功', push_at=timezone.now())
                                    order.auto_check = cls.CHECK_SUCCESS
                                    tiktok_check = True
                                    msg = '抖音自动核销成功'
                            except Exception as e:
                                log.error(e)
                                msg = e
                            if tu_check:
                                tu_check.update(msg=msg, push_at=timezone.now(), tiktok_check=tiktok_check)
                            if tu_uncheck:
                                # 核销成功才变成已过期
                                if tiktok_check:
                                    tu_uncheck.update(msg=msg, push_at=timezone.now(), tiktok_check=tiktok_check,
                                                      status=TicketUserCode.STATUS_OVER_TIME)
                                else:
                                    tu_uncheck.update(msg=msg, push_at=timezone.now(), tiktok_check=tiktok_check)
                        order.save(update_fields=['auto_check'])
                        order.set_finish()
                finally:
                    redis.delete(auth_check_code_tiktok_key)

    def push_delivery_order(self):
        if self.status not in [TicketOrder.STATUS_PAID, TicketOrder.STATUS_FINISH, TicketOrder.STATUS_OVER_TIME]:
            return
        from douyin import get_dou_yin
        dy = get_dou_yin()
        order = self
        order.auto_check = TicketOrder.CHECK_FAIL
        tu_check = None
        tu_uncheck = None
        if self.status in [TicketOrder.STATUS_PAID, TicketOrder.STATUS_FINISH]:
            tu_check = TicketUserCode.objects.filter(order_id=order.id, status=TicketUserCode.STATUS_CHECK,
                                                     tiktok_check=False)
            tu_uncheck = TicketUserCode.objects.filter(order_id=order.id,
                                                       status=TicketUserCode.STATUS_DEFAULT,
                                                       tiktok_check=False)
        # 推抖音
        tiktok_check = False
        poi_info = {"shop_name": order.session.tiktok_store.name,
                    "ext_valid_shop_id": str(order.session.tiktok_store.id),
                    "valid_poi_id_str": order.session.tiktok_store.supplier_ext_id}
        params = dict(out_order_no=order.order_no, use_all=True, poi_info=json.dumps(poi_info))
        try:
            ret = dy.push_delivery_new(params=params)
            if ret['error_code'] != 0:
                log.error('自动核销失败,{},{}'.format(ret['sub_description'], order.order_no))
                msg = ret['sub_description']
                # tu_list.update(msg='抖音自动核销失败')
            else:
                # tu_list.update(msg='抖音自动核销成功', push_at=timezone.now())
                order.auto_check = TicketOrder.CHECK_SUCCESS
                tiktok_check = True
                msg = '抖音自动核销成功'
        except Exception as e:
            log.error(e)
            msg = e
        if tu_check:
            tu_check.update(msg=msg, push_at=timezone.now(), tiktok_check=tiktok_check)
        if tu_uncheck:
            # 核销成功才变成已过期
            if tiktok_check:
                tu_uncheck.update(msg=msg, push_at=timezone.now(), tiktok_check=tiktok_check,
                                  status=TicketUserCode.STATUS_OVER_TIME)
            else:
                tu_uncheck.update(msg=msg, push_at=timezone.now(), tiktok_check=tiktok_check)
        order.save(update_fields=['auto_check'])
        order.set_finish()

    @classmethod
    def check_auto_expire(cls):
        close_old_connections()
        now = timezone.now()
        from mp.models import BasicConfig
        bc = BasicConfig.get()
        auto_cancel_minutes = bc.auto_cancel_minutes if bc else 10
        expire_at = now + timedelta(minutes=-auto_cancel_minutes)
        for order in cls.objects.filter(create_at__lte=expire_at, status=cls.STATUS_UNPAID):
            order.cancel()

    def set_finish(self):
        self.status = self.STATUS_FINISH
        self.save(update_fields=['status'])

    def push_tiktok(self):
        # https://open.douyin.com/api/apps/trade/v2/order/create_order 发起下单
        from django.db.models import Count
        from mp.models import BasicConfig
        bc = BasicConfig.get()
        config = get_config()
        show = self.session.show
        uc_list = TicketUserCode.objects.filter(order=self).values('level_id').annotate(
            quantity=Count("level_id")).order_by('level_id')
        goods_image = '{}{}'.format(config['template_url'], show.logo_mobile.url)
        labels = '|'.join([str(flag) for flag in show.flag.all()])
        start_at = self.session.start_at - timedelta(
            hours=2) if not self.session.valid_start_time else self.session.valid_start_time
        sku_list = []
        # 回调地址
        pay_notify_url = '{}/api/receipts/tiktok_notify/'.format(config['template_url'])
        params = {"id": str(show.id), "session_id": str(self.session.id)}
        product_name = self.session.get_dy_product_name()
        for sku in uc_list:
            tf = TicketFile.objects.get(id=sku['level_id'])
            sku_list.append(
                {
                    "quantity": sku['quantity'],
                    "sku_id": 'tl'.format(sku['level_id']),
                    "sku_id_type": 2,
                    "price": tf.price,
                    "discount_amount": 0,
                    "goods_info": {
                        "goods_image": goods_image,
                        "goods_title": product_name,
                        "labels": labels,
                        # "date_rule": "周一至周日可用",
                        "goods_id": 'ss'.format(self.session.id),
                        "goods_id_type": 2,
                        # "goods_page": {
                        #     "path": "goods/infoxxxx",
                        #     "params":"{\"id\":\"xxxxxx\"}"
                        # },
                        "order_valid_time": {
                            "valid_start_time": get_timestamp(start_at),
                            "valid_end_time": get_timestamp(self.session.end_at),
                        },
                        # 预售券商品必传,门票商品不用传入
                        # "goods_book_info": {
                        #     "book_type": 2,
                        #     "cancel_policy": 1
                        # },
                    }
                }
            )
        data = {
            "sku_list": sku_list,
            # "cp_book_info":'预约信息',
            "total_amount": self.actual_amount,
            "phone_num": self.mobile,
            "contact_name": self.name,
            # "extra": "",
            "open_id": self.user.openid_tiktok,
            "pay_notify_url": pay_notify_url,
            "out_order_no": self.order_no,
            "pay_expire_seconds": bc.auto_cancel_minutes * 60,
            "order_entry_schema": {
                "path": "pages/pagesKage/showDetail/showDetail",
                "params": json.dumps(params)
            },
            # "cp_extra": "开发者自定义透传字段，不支持二进制，长度 <= 2048 byte",
            "discount_amount": self.amount - self.actual_amount,
            # "price_calculation_detail":{}
        }

    def tiktok_prepare_order(self, session_seat_list):
        # 服务器推的
        if not self.user.openid_tiktok:
            raise CustomAPIException('请先授权抖音登录后再支付')
        from mp.models import SystemDouYinMP
        dy_mp = SystemDouYinMP.get()
        config = get_config()
        show = self.session.show
        goods_image = '{}{}'.format(config['template_url'], show.logo_mobile.url)
        labels = '|'.join([str(flag) for flag in show.flag.all()])
        start_at = self.session.start_at - timedelta(
            hours=2) if not self.session.valid_start_time else self.session.valid_start_time
        item_order_info_list = sku_list = []
        # 回调地址,配置在支付上面
        pay_notify_url = '{}/api/receipts/tiktok_notify/'.format(config['template_url'])
        account = self.user.account
        seat_dict = dict()
        for seat in session_seat_list:
            if seat_dict.get(str(seat.ticket_level.id)):
                seat_dict[str(seat.ticket_level.id)] += 1
            else:
                seat_dict[str(seat.ticket_level.id)] = 1
        for level_id, quantity in seat_dict.items():
            tf = TicketFile.objects.get(id=int(level_id))
            sku_list.append(
                {
                    "quantity": quantity,
                    "sku_id": 'tl'.format(level_id),
                    "sku_id_type": 2,
                    "price": tf.price * account.get_discount(),
                    "origin_price": tf.price,
                    # "atts":"{}",
                    "goods_info": {
                        "img_url": goods_image,
                        "title": show.title,
                        "sub_title": '{}-{}'.format(self.session.start_at.strftime('%Y-%m-%d %H:%M'),
                                                    self.session.end_at.strftime('%Y-%m-%d %H:%M')),
                        "labels": labels,
                        "date_rule": '{}-{}可用'.format(start_at.strftime('%Y-%m-%d %H:%M'),
                                                      self.session.end_at.strftime('%Y-%m-%d %H:%M')),
                        # "poi_id":"对应门店的Poi ID",
                        "goods_id": 'ss'.format(self.session.id),
                        "goods_id_type": 2,
                        # "goods_book_info":{} # 预约信息
                    }
                }
            )
            item_order_info_list.append(
                {
                    "sku_id": 'tl'.format(level_id),
                    "sku_id_type": 2,
                    "goods_id": 'ss'.format(self.session.id),
                    "goods_id_type": 2,
                    "item_order_id": self.order_no,
                    "price": tf.price * account.get_discount(),
                }
            )
        data = {
            "order_id": self.order_no,
            "sku_list": sku_list,
            # "cp_book_info":'预约信息',
            "total_amount": self.actual_amount,
            "discount": self.amount - self.actual_amount,
            # "cp_extra": "开发者自定义透传字段，不支持二进制，长度 <= 2048 byte",
            "create_order_time": get_timestamp(self.create_at),
            "open_id": self.user.openid_tiktok,
            "phone_num": self.mobile,
            "contact_name": self.name,
            "app_id": dy_mp.app_id,
            "union_id": self.user.unionid_tiktok,
            "delivery_type": 0,
            # "address":"有物流配送需求的时候提单页会让用户填入配送地址",
            "item_order_info_list": item_order_info_list
        }
        return data

    def tiktok_client_prepare_order(self, session_seat_list=None, seat_dict=None):
        # 服务器推的
        if not self.user.openid_tiktok:
            raise CustomAPIException('请先授权抖音登录后再支付')
        config = get_config()
        show = self.session.show
        # goods_image = '{}{}'.format(config['template_url'], show.logo_mobile.url)
        # labels = list(show.flag.all().values_list('title', flat=True))
        # start_at = self.session.start_at - timedelta(
        #     hours=2) if not self.session.valid_start_time else self.session.valid_start_time
        sku_list = []
        # 回调地址,配置在支付上面
        pay_notify_url = '{}/api/receipts/tiktok_notify/'.format(config['template_url'])
        # account = self.user.account
        if not seat_dict:
            seat_dict = dict()
            if session_seat_list:
                for seat in session_seat_list:
                    if seat_dict.get(str(seat.ticket_level.id)):
                        seat_dict[str(seat.ticket_level.id)] += 1
                    else:
                        seat_dict[str(seat.ticket_level.id)] = 1
        for level_id, quantity in seat_dict.items():
            tf = TicketFile.objects.get(id=int(level_id))
            # log.error(tf.product_id)
            sku_list.append(
                {
                    "skuId": tf.product_id,
                    "skuType": 1,
                    "quantity": quantity,
                    "price": int(tf.price * 100),
                    # "discountAmount": int((tf.price-tf.price * self.user.account.get_discount()) * 100)
                    # 分
                    # "price": int(tf.price * account.get_discount() * 100),
                    # "goodsInfo": {
                    #     "goodsPhoto": goods_image,
                    #     "goodsName": show.title,
                    #     "goodsId": tf.product_id,
                    #     "goodsType": 2,
                    #     "goodsLabels": labels,
                    #     "dateRule": '{}-{}可用'.format(start_at.strftime('%Y-%m-%d %H:%M'),
                    #                                  self.session.end_at.strftime('%Y-%m-%d %H:%M')),
                    # },
                    # "extraInfo": {
                    #     "ticketName": show.show_type.name,
                    #     "date": self.session.start_at.strftime('%Y-%m-%d'),
                    # }
                }
            )
        data = {
            "skuList": sku_list,
            # 分
            "payment": {"totalAmount": int(self.amount * 100), },
            #  "totalDiscountAmount": int((self.amount - self.actual_amount) * 100)
            "contactInfo": {"phoneNumber": self.mobile, "contactName": self.name},
            "callbackUrl": pay_notify_url,
            "callbackData": {"my_order_id": self.id}
        }
        log.debug(data)
        return data

    def tiktok_client_prepare_order_new(self, session_seat_list=None, seat_dict=None):
        # 服务器推的
        if not self.user.openid_tiktok:
            raise CustomAPIException('请先授权抖音登录后再支付')
        config = get_config()
        show = self.session.show
        goods_image = '{}{}'.format(config['template_url'], show.logo_mobile.url)
        labels = list(show.flag.all().values_list('title', flat=True))
        start_at = self.session.start_at - timedelta(
            hours=2) if not self.session.valid_start_time else self.session.valid_start_time
        sku_list = []
        # 回调地址,配置在支付上面
        pay_notify_url = '{}{}'.format(config['template_url'], tiktok_notify_url)
        # account = self.user.account
        if not seat_dict:
            seat_dict = dict()
            if session_seat_list:
                for seat in session_seat_list:
                    if seat_dict.get(str(seat.ticket_level.id)):
                        seat_dict[str(seat.ticket_level.id)] += 1
                    else:
                        seat_dict[str(seat.ticket_level.id)] = 1
        product_id = self.session.product_id
        goodsName = self.session.get_dy_product_name()
        is_cal_express = False
        for level_id, quantity in seat_dict.items():
            tf = TicketFile.objects.get(id=int(level_id))
            # log.error(tf.product_id)
            # skuType = 1 if tf.product_id and tf.push_status == TicketFile.PUSH_SUCCESS else 2
            if not product_id:
                # product_id = tf.product_id
                # if not product_id:
                product_id = tf.get_product_id()
            price = tf.price
            if not is_cal_express and self.express_fee > 0:
                is_cal_express = True
                price = price + self.express_fee / quantity
            price = int(price * 100)
            sku_list.append(
                {
                    "skuId": tf.get_out_id(),
                    "skuType": 2,
                    "quantity": quantity,
                    "price": price,
                    # "discountAmount": int((tf.price-tf.price * self.user.account.get_discount()) * 100)
                    # 分
                    # "price": int(tf.price * account.get_discount() * 100),
                    "goodsInfo": {
                        "goodsPhoto": goods_image,
                        "goodsName": goodsName,
                        "goodsId": product_id,
                        "goodsType": 1,
                        "goodsLabels": labels,
                        "dateRule": '{}-{}可用'.format(start_at.strftime('%Y-%m-%d %H:%M'),
                                                     self.session.end_at.strftime('%Y-%m-%d %H:%M')),
                    },
                    # "extraInfo": {
                    #     "ticketName": show.show_type.name,
                    #     "date": self.session.start_at.strftime('%Y-%m-%d'),
                    # }
                }
            )
        data = {
            "skuList": sku_list,
            # 分
            "payment": {"totalAmount": int(self.amount * 100), },
            #  "totalDiscountAmount": int((self.amount - self.actual_amount) * 100)
            "contactInfo": {"phoneNumber": self.mobile, "contactName": self.name},
            "callbackUrl": pay_notify_url,
            "callbackData": {"my_order_id": self.id}
        }
        log.debug(data)
        return data

    @classmethod
    def down_to_excel(cls, ids=list):
        queryset = cls.objects.filter(id__in=ids)
        if not queryset:
            return None
        import xlwt
        from ticket.utils import excel_dir, _write_row_by_xlwt
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('演出订单')
        row_index = 1
        _write_row_by_xlwt(ws, cls.export_fields(), row_index)
        row_index += 1
        log.warning('演出订单导出开始')
        dir, rel_url, xlsx_dir = excel_dir()
        filename = '{}{}.xls'.format(timezone.now().strftime('%Y%m%d%H%M%S'), random_str(10))
        filepath = os.path.join(dir, filename)
        for record in queryset:
            create_at = record.create_at.strftime('%Y-%m-%d %H:%M:%S')
            pay_at = record.pay_at.strftime('%Y-%m-%d %H:%M:%S') if record.pay_at else None
            start_at = record.start_at.strftime('%Y-%m-%d %H:%M:%S') if record.start_at else None
            seat_desc = ''
            level_desc = ''
            pay_desc = ''
            if record.wx_pay_config:
                pay_desc = record.wx_pay_config.title
            if record.dy_pay_config:
                pay_desc = record.dy_pay_config.title
            tu_qs = TicketUserCode.objects.filter(order=record)
            for tu in tu_qs:
                if tu.session_seat:
                    ss = tu.session_seat.seat_desc(record.venue)
                    if not seat_desc:
                        seat_desc = ss
                    else:
                        seat_desc += ',{}'.format(ss)
                else:
                    if not seat_desc:
                        seat_desc = '无座'
                    else:
                        seat_desc += ',无座'
                snapshot = json.loads(tu.snapshot)
                if not level_desc:
                    level_desc = snapshot['desc']
                else:
                    level_desc += snapshot['desc']
            data = [str(record.user), record.name, record.mobile, record.show_express_address,
                    str(record.agent) if record.agent else None,
                    record.get_pay_type_display(), pay_desc, seat_desc, level_desc,
                    record.order_no, record.receipt.payno, record.receipt.transaction_id, record.multiply,
                    record.amount, record.card_jc_amount, record.actual_amount, record.express_fee,
                    record.get_status_display(), record.title, create_at, pay_at, start_at,
                    record.get_source_type_display(), record.tiktok_nickname, record.tiktok_douyinid, record.plan_id,
                    record.get_plan_name(), str(record.venue)]
            _write_row_by_xlwt(ws, data, row_index)
            row_index += 1
        _write_row_by_xlwt(ws, ['END'], row_index)
        wb.save(filepath)
        return '{}/{}'.format(xlsx_dir, filename)


class TicketOrderRealName(models.Model):
    order = models.ForeignKey(TicketOrder, verbose_name='订单', on_delete=models.CASCADE, related_name='real_name_order')
    name = models.CharField('姓名', max_length=30)
    mobile = models.CharField('手机号', max_length=20)
    id_card = models.CharField('身份证号', max_length=20, null=True, blank=True)

    class Meta:
        verbose_name_plural = verbose_name = '订单实名信息'
        ordering = ['-pk']

    def __str__(self):
        return self.name

    @classmethod
    def create_record(cls, order: TicketOrder, name, mobile, id_card=None):
        inst = cls.objects.create(order=order, name=name, mobile=mobile, id_card=id_card)
        return inst


class TicketOrderExpress(models.Model):
    file = models.FileField(u'发货回执单', upload_to=f'{FILE_FIELD_PREFIX}/ticket/order_express/%Y/%m/%d',
                            help_text=u'上传回传物流单号的发货单文件，表格字段要求和导出的发货单保持一致！注意!仅支持xlsx格式!',
                            validators=[FileExtensionValidator(allowed_extensions=['xlsx'])])
    create_at = models.DateTimeField(u'创建时间', auto_now_add=True)
    desp = models.CharField(u'备注', max_length=100, null=True, blank=True)
    error_msg = models.TextField('导入错误信息', null=True, blank=True)
    user = models.ForeignKey(User, verbose_name='操作人', null=True, blank=True, editable=False,
                             on_delete=models.SET_NULL)

    class Meta:
        verbose_name_plural = verbose_name = u'发货回执记录'
        ordering = ['-create_at']

    def __str__(self):
        return str(self.create_at)

    def save(self, force_insert=False, force_update=False, using=None,
             update_fields=None):
        ret = super(TicketOrderExpress, self).save(force_insert, force_update, using, update_fields)
        from openpyxl import load_workbook
        wb = load_workbook(self.file.path)
        ws = wb.active
        begin = 2
        error_msg = ''
        for line, row in enumerate(ws.rows, 1):
            if line < begin:
                title_list = [t.value for t in row]
                order_index = title_list.index('订单号')
                comp_code_index = title_list.index('快递公司编码')
                comp_index = title_list.index('快递公司')
                express_no_index = title_list.index('快递单号')
                continue
            dd = list(map(lambda cell: cell.value, row))
            if not dd[order_index]:
                continue
            res = TicketOrderExpress.update_express(dd, order_index, comp_code_index, comp_index, express_no_index)
            if not res:
                msg = ",第{}行出现错误!数据: {}".format(line, row[0].value)
                error_msg += msg
                log.warning(msg)
        if error_msg:
            self.error_msg = error_msg
            if not update_fields:
                update_fields = []
            update_fields.append('error_msg')
            ret = super(TicketOrderExpress, self).save(force_insert, force_update, using, update_fields)
        return ret

    @classmethod
    def _is_invalid_row(cls, row):
        return not row[0]

    @classmethod
    def update_express(cls, row, order_index, comp_code_index, comp_index, express_no_index):
        if cls._is_invalid_row(row):
            return False
        orderno = row[order_index]
        st = False
        try:
            order = TicketOrder.objects.get(order_no=orderno)
            if row[express_no_index]:
                if order.express_status in TicketOrder.can_express_status() and not order.express_comp_no:
                    order.express_comp_no = row[comp_code_index]
                    order.express_name = row[comp_index]
                    order.express_no = row[express_no_index]
                    order.express_status = TicketOrder.EXPRESS_DELIVER
                    order.deliver_at = timezone.now()
                    order.save(
                        update_fields=['express_comp_no', 'express_name', 'express_no', 'express_status', 'deliver_at'])
                    st = True
            else:
                log.warning("快递号为空")
        except TicketOrder.DoesNotExist:
            log.warning("订单找不到: {}".format(orderno))
        return st


class TicketOrderChangePrice(models.Model):
    order = models.ForeignKey(TicketOrder, verbose_name='订单', on_delete=models.CASCADE)
    user = models.ForeignKey(User, verbose_name='修改用户', on_delete=models.SET_NULL, null=True)
    before_amount = models.DecimalField('修改前实付金额', max_digits=13, decimal_places=2, default=0)
    after_amount = models.DecimalField('修改后实付金额', max_digits=13, decimal_places=2, default=0)
    create_at = models.DateTimeField('修改时间', auto_now_add=True)

    class Meta:
        verbose_name_plural = verbose_name = '订单改价记录'
        ordering = ['-pk']

    def __str__(self):
        return self.order.order_no

    @classmethod
    def create_record(cls, user, order, after_amount):
        inst = cls.objects.create(order=order, user=user, before_amount=order.receipt.amount, after_amount=after_amount)
        order.change_actual_amount(after_amount)
        return inst


class ShowComment(models.Model):
    ST_DEFAULT = 1
    ST_FINISH = 2
    ST_FAIL = 3
    ST_CHOICES = ((ST_DEFAULT, '待审核'), (ST_FINISH, '已审核'), (ST_FAIL, '未通过'))
    show = models.ForeignKey(ShowProject, verbose_name='演出项目', on_delete=models.CASCADE)
    title = models.CharField('演出名称', max_length=60, null=True)
    start_at = models.DateTimeField('演出开演时间', null=True)
    order = models.ForeignKey(TicketOrder, verbose_name='订单', on_delete=models.CASCADE)
    order_no = models.CharField(u'订单号', max_length=128, null=True)
    user = models.ForeignKey(User, verbose_name='用户', on_delete=models.CASCADE)
    mobile = models.CharField('手机', max_length=20, null=True)
    content = models.TextField('内容', max_length=200)
    status = models.IntegerField('状态', choices=ST_CHOICES, default=ST_DEFAULT)
    is_display = models.BooleanField('是否展示', default=False, help_text='展示在演出')
    is_quality = models.BooleanField('是否优质评论', default=False)
    create_at = models.DateTimeField('创建时间', auto_now_add=True, editable=True)
    approve_at = models.DateTimeField('审核时间', null=True, blank=True)

    class Meta:
        verbose_name_plural = verbose_name = '演出评论'
        ordering = ['-pk']

    def __str__(self):
        return '{}: {}'.format(self.title, self.content[0:10])

    @classmethod
    def create_record(cls, show, order, content, user):
        inst = cls.objects.create(show=show, title=show.title, order=order, order_no=order.order_no, user=user,
                                  mobile=user.mobile, content=content, start_at=order.session.start_at)
        order.is_comment = True
        order.save(update_fields=['is_comment'])
        return inst


class ShowCommentImage(models.Model):
    record = models.ForeignKey(ShowComment, verbose_name='演出评论', on_delete=models.CASCADE, null=True)
    user_id = models.PositiveIntegerField('用户ID')
    image = models.ImageField('附件图片', upload_to=f'{IMAGE_FIELD_PREFIX}/ticket/comment',
                              validators=[validate_image_file_extension])

    class Meta:
        verbose_name_plural = verbose_name = '演出评论图片'

    def __str__(self):
        return str(self.id)


class TicketUserCode(models.Model):
    order = models.ForeignKey(TicketOrder, verbose_name='订单', on_delete=models.CASCADE, null=True,
                              related_name='user_code')
    level_id = models.IntegerField('票档ID', default=0, editable=False)
    product_id = models.CharField('抖音商品ID', null=True, blank=True, max_length=50)
    price = models.DecimalField('售价', max_digits=13, decimal_places=2, default=0)
    session_seat = models.ForeignKey(SessionSeat, verbose_name='演出场次座位', null=True, blank=True,
                                     on_delete=models.CASCADE)
    session_id = models.IntegerField('场次ID', default=0, editable=False, db_index=True)
    STATUS_DEFAULT = 1
    STATUS_CHECK = 2
    STATUS_OVER_TIME = 3
    STATUS_CANCEL = 4
    STATUS_CHOICES = (
        (STATUS_DEFAULT, u'未检票'), (STATUS_CHECK, '已检票'), (STATUS_OVER_TIME, '已过期'), (STATUS_CANCEL, '已作废'))
    status = models.IntegerField(u'检票状态', choices=STATUS_CHOICES, default=STATUS_DEFAULT)
    source_type = models.IntegerField(u'带货场景', choices=TicketOrder.SOURCE_CHOICES, default=TicketOrder.SOURCE_DEFAULT)
    check_user = models.ForeignKey('shopping_points.UserAccount', verbose_name='验票员', null=True, blank=True,
                                   on_delete=models.SET_NULL)
    code = models.CharField('检票码', max_length=30, null=True, db_index=True)
    code_img = models.ImageField('检票二维码', null=True, blank=True, upload_to=f'{IMAGE_FIELD_PREFIX}/ticket/code',
                                 validators=[validate_image_file_extension])
    check_at = models.DateTimeField('检票时间', null=True, blank=True)
    snapshot = models.TextField('场次座位快照', null=True, blank=True, help_text='下单时保存的场次座位快照', editable=False,
                                max_length=2048)
    tiktok_check = models.BooleanField('是否推送抖音核销', default=False)
    push_at = models.DateTimeField('核销时间', null=True, blank=True)
    msg = models.CharField('核销返回', max_length=100, null=True, blank=True)
    create_at = models.DateTimeField('创建时间', auto_now_add=True, null=True)
    GIVE_DEFAULT = 0
    GIVE_UNCLAIMED = 1
    GIVE_FINISH = 2
    GIVE_CHOICES = [(GIVE_DEFAULT, '未赠送'), (GIVE_UNCLAIMED, '待领取'), (GIVE_FINISH, '已领取')]
    give_status = models.SmallIntegerField('赠送状态', choices=GIVE_CHOICES, default=GIVE_DEFAULT)
    give_mobile = models.CharField('受赠人手机号', max_length=20, null=True, blank=True)
    give_id = models.IntegerField('门票赠送记录ID', default=0, editable=False)

    def __str__(self):
        return str(self.id)

    class Meta:
        verbose_name_plural = verbose_name = '演出票(座位)信息'
        ordering = ['-pk']

    random_code_len = 4

    def clear_give(self):
        self.give_status = self.GIVE_DEFAULT
        self.give_mobile = None
        self.give_id = 0
        self.save(update_fields=['give_status', 'give_mobile', 'give_id'])

    def confirm_give(self):
        self.give_status = self.GIVE_FINISH
        self.save(update_fields=['give_status'])

    @classmethod
    def check_refund_mz(cls, session_seat_id: int):
        tu = cls.objects.filter(session_seat_id=session_seat_id,
                                status__in=[cls.STATUS_DEFAULT,
                                            cls.STATUS_CHECK]).first()
        if tu:
            tu.order.set_need_refund_mz(True)

    @classmethod
    def check_cps_source(cls):
        close_old_connections()
        from douyin import get_dou_yin
        dy = get_dou_yin()
        now = timezone.now()
        start_at = now - timedelta(days=2)
        for code in cls.objects.filter(source_type=TicketOrder.SOURCE_DEFAULT, product_id__isnull=False,
                                       status__in=[cls.STATUS_CHECK, cls.STATUS_DEFAULT],
                                       order__source_type=TicketOrder.SOURCE_DEFAULT,
                                       order__status__in=[TicketOrder.STATUS_PAID, TicketOrder.STATUS_FINISH],
                                       order__pay_type=Receipt.PAY_TikTok_LP,
                                       order__pay_at__lt=now, order__pay_at__gt=start_at):
            try:
                st, ret = dy.query_cps(code.order.order_no)
                if st and ret.get('data') and ret['data'].get('cps_info') and \
                        ret['data']['cps_info'][
                            'cps_item_list']:
                    code.source_type = ret['data']['cps_info']['cps_item_list'][0]['source_type']
                    code.order.tiktok_douyinid = ret['data']['cps_info']['cps_item_list'][0][
                        'commission_user_douyinid']
                    code.order.tiktok_nickname = ret['data']['cps_info']['cps_item_list'][0][
                        'commission_user_nickname']
                    code.order.plan_id = ret['data']['cps_info']['cps_item_list'][0]['task_id']
                    code.order.source_type = code.source_type
                    code.order.save(update_fields=['source_type', 'tiktok_douyinid', 'tiktok_nickname', 'plan_id',
                                                   'source_type'])

                else:
                    # 非CPS订单或带货场景为空的订单按照目前的逻辑正常计算
                    code.source_type = TicketOrder.SOURCE_NO
                    code.order.source_type = TicketOrder.SOURCE_NO
                    code.order.save(update_fields=['source_type'])
                code.save(update_fields=['source_type'])
            except Exception as e:
                continue

    @classmethod
    def set_overtime(cls):
        close_old_connections()
        qs = cls.objects.filter(deadline_at__isnull=False, deadline_at__lte=timezone.now(), status=cls.STATUS_DEFAULT)
        for inst in qs:
            inst.status = cls.STATUS_OVER_TIME
            inst.save(update_fields=['status'])

    @property
    def share_img_key(self):
        # code->动态码 key
        return get_redis_name('simg' + str(self.code))

    def change_share_code_img(self, share_img_path: str):
        # log.error(share_img_path)
        with get_pika_redis() as redis:
            key = self.share_img_key
            img_path = redis.get(key)
            # log.error(img_path)
            try:
                if img_path and os.path.isfile(img_path):
                    # log.error(img_path)
                    os.unlink(img_path)
            except Exception as e:
                log.error(e)
            redis.set(key, share_img_path)

    @property
    def pika_code_key(self):
        # code->动态码 key
        return get_redis_name('pck' + str(self.code))

    @classmethod
    def pika_t_code_key(cls, t_code):
        # 动态码->code key
        return get_redis_name('ptck' + str(t_code))

    def refresh_random_digits(self, len: int) -> str:
        session = self.order.session
        random_dg = random_new_digits(len)
        t_code = '{}{}'.format(self.code, random_dg)
        t_code_key = TicketUserCode.pika_t_code_key(t_code)
        with get_pika_redis() as redis:
            redis.set(self.pika_code_key, json.dumps(dict(t_code=t_code, random_dg=random_dg)))
            redis.set(t_code_key, self.code)
            if session.check_is_dy_code:
                dc_expires_in = session.dc_expires_in * 60
                redis.expire(self.pika_code_key, dc_expires_in)
                redis.expire(t_code_key, dc_expires_in)
        return str(random_dg)

    def code_get_t_code(self) -> dict:
        # 不存在则重新生成
        # expires_in -1 是永久 -2是没有
        # dict(t_code=t_code, random_dg=random_dg)
        with get_pika_redis() as redis:
            data = redis.get(self.pika_code_key)
            expires_in = redis.ttl(self.pika_code_key)
            data = json.loads(data) if data else dict()
            data['expires_in'] = expires_in
        return data

    @classmethod
    def t_code_get_code(cls, t_code: str):
        # t_code 获取code
        t_code_key = cls.pika_t_code_key(t_code)
        with get_pika_redis() as redis:
            code = redis.get(t_code_key)
        return code

    @classmethod
    def check_t_code(cls, t_code: str):
        # log.debug(t_code)
        if len(t_code) == 12:
            # 兼容旧的二维码code 12位
            return t_code
        code = cls.t_code_get_code(t_code)
        return code

    def del_random_digits(self):
        # 清除随机数
        with get_pika_redis() as redis:
            t_code_dict = self.code_get_t_code()
            redis.delete(self.pika_code_key)
            if t_code_dict.get('t_code'):
                t_code_key = TicketUserCode.pika_t_code_key(t_code_dict.get('t_code'))
                redis.delete(t_code_key)

    @classmethod
    def create_record(cls, order, session_seat=None, ticket_level=None):
        # from common import qrutils
        if not ticket_level:
            ticket_level = session_seat.ticket_level
        inst = cls.objects.create(order=order, level_id=ticket_level.id, session_seat=session_seat,
                                  price=ticket_level.price,
                                  session_id=ticket_level.session.id, product_id=ticket_level.product_id)
        inst.code = inst.get_code()
        # dir, rel_url, img_dir = qrcode_dir_tk()
        # filename = 'code_{}_v{}.jpg'.format(inst.id, 1)
        # file_path = os.path.join(dir, filename)
        # qrutils.generate(inst.code, size=(410, 410), save_path=file_path)
        # inst.code_img = '{}/{}'.format(img_dir, filename)
        inst.snapshot = inst.get_snapshot(session_seat, ticket_level)
        inst.save(update_fields=['code', 'snapshot'])

    def del_code_img(self):
        try:
            if self.code_img and os.path.isfile(self.code_img.path):
                os.unlink(self.code_img.path)
                self.code_img = None
                self.save(update_fields=['code_img'])
            self.del_random_digits()
        except Exception as e:
            log.error(e)

    def new_code_path(self, random_digits):
        dir, rel_url, img_dir = qrcode_dir_tk()
        code = self.code + str(random_digits)
        name = hash_ids(int(code))
        filename = '{}.jpg'.format(name)
        file_path = os.path.join(dir, filename)
        if os.path.isfile(file_path):
            filename = '{}{}.jpg'.format(name, random_str(4))
        return img_dir, file_path, filename

    def create_qr_code_img(self):
        # 生成码
        self.del_code_img()
        random_digits = self.refresh_random_digits(self.random_code_len)
        code = self.code + str(random_digits)
        # from common import qrutils
        # img_dir, file_path, filename = self.new_code_path(random_digits)
        # qrutils.generate(code, size=(410, 410), save_path=file_path)
        img_dir, file_path, filename = self.create_qr_code(str(random_digits))
        self.code_img = '{}/{}'.format(img_dir, filename)
        self.save(update_fields=['code_img'])
        return self.code_img.url, code

    def create_qr_code(self, random_digits: str):
        code = self.code + str(random_digits)
        from common import qrutils
        img_dir, file_path, filename = self.new_code_path(random_digits)
        qrutils.generate(code, size=(410, 410), save_path=file_path)
        return img_dir, file_path, filename

    def get_code_img_new(self, is_refresh=False):
        """
        动态二维码重新生成t_code，静态直接获取
        """
        session = self.order.session
        url = None
        # random_digits = None
        create_qr_code = False
        code = None
        t_code_dict = self.code_get_t_code()
        # log.debug(t_code_dict)
        if session.check_is_dy_code:
            # config = get_config()
            # dynamics_share_minute = config.get('dynamics_share_minute') or 120
            create_qr_code = is_refresh
            if not create_qr_code and t_code_dict['expires_in'] <= 10:
                """
                已经过期，需要刷新动态码
                """
                create_qr_code = True
        if self.code_img:
            """ 旧的二维码没有t_code，新的有"""
            url = self.code_img.url
            code = t_code_dict.get('t_code') if t_code_dict else None
        else:
            # 新逻辑生成二维码，和固定随机数
            create_qr_code = True
        if create_qr_code:
            url, code = self.create_qr_code_img()
        if not code:
            code = self.code
        return url, code

    def check_can_share(self):
        """
        重新获取判断过期时间，时候满足分享码要求
        """
        session = self.order.session
        can_share = True
        expires_in = -1
        if session.check_is_dy_code:
            t_code_dict = self.code_get_t_code()
            config = get_config()
            dynamics_share_minute = config.get('dynamics_share_minute') or 120
            expires_in = t_code_dict['expires_in']
            if expires_in < dynamics_share_minute * 60:
                can_share = False
        deadline_at = timezone.now() + timedelta(seconds=expires_in)
        deadline_at_str = deadline_at.strftime('%Y-%m-%d %H:%M') if expires_in > 0 else None
        deadline_timestamp = get_timestamp(deadline_at)
        return can_share, deadline_at_str, deadline_timestamp

    def get_code(self):
        import random
        num = 100000000000
        ss = str(random.randint(100, 999))
        if self.id < 100000000:
            i = 11
        else:
            i = len(str(self.id)) + 3
        code = num + int(ss.ljust(i, '0')) + self.id
        return str(code)

    def get_snapshot(self, session_seat=None, ticket_level=None):
        import json
        """
        商品快照
        :return:
        """
        seat = ''
        if session_seat:
            seat = session_seat.seat_desc(self.order.venue)
        data = dict(color=ticket_level.color.name,
                    origin_price=float(ticket_level.origin_price), desc=ticket_level.desc,
                    seat=seat,
                    price=float(ticket_level.price))
        return json.dumps(data)

    @classmethod
    def check_code_order(cls, user, session_id, code):
        # session_id is no
        # session_id = int(session_id)
        inst = cls.objects.filter(code=code).first()
        if not inst or (not inst.order):
            raise CustomAPIException('二维码无效')
        else:
            if inst.status != cls.STATUS_DEFAULT:
                raise CustomAPIException('二维码已核销')
        if inst.give_id > 0:
            # 已经赠送的二维码走单个扫码
            status, msg, snapshot = cls.check_code(user, session_id, code)
            success = 0
            fail = 0
            if status:
                success = 1
            else:
                fail = 1
            return dict(status=status, msg=msg, snapshot=snapshot, success=success, fail=fail)
        else:
            try:
                session = SessionInfo.objects.get(no=session_id)
                xunyan = ShowType.xunyan()
                if session.show.show_type_id == xunyan.id:
                    raise CustomAPIException('巡演类型，扫码失败')
            except SessionInfo.DoesNotExist:
                raise CustomAPIException('不是本场次门票，请核对后重试！')
            qs = cls.objects.filter(order_id=inst.order.id, status=cls.STATUS_DEFAULT,
                                    check_at__isnull=True, give_id=0)
            ret = dict(status=False, msg='', snapshot=None, success=0, fail=0)
            for inst in qs:
                status, msg, snapshot = cls.check_code(user, session_id, inst.code, inst)
                if status:
                    # 一条成功就算成功
                    ret['status'] = status
                    ret['msg'] = ''
                    ret['success'] += 1
                else:
                    ret['fail'] += 1
                    ret['msg'] = msg
                if not ret['snapshot']:
                    ret['snapshot'] = snapshot
            return ret

    @classmethod
    @atomic
    def check_code(cls, user, session_id, code, inst=None):
        # 防止同时检验时两边都成功
        # log.debug('开始验票')
        # session_id = int(session_id)
        from caches import get_redis, check_code_key
        key = check_code_key.format(session_id, code)
        redis = get_redis()
        snapshot = None
        try:
            if redis.setnx(key, 1):
                redis.expire(key, 5)
                if not inst:
                    inst = cls.objects.filter(code=code).get()
                if inst.status != cls.STATUS_DEFAULT:
                    status = False
                    msg = '该二维码已经核销过了'
                else:
                    snapshot = json.loads(inst.snapshot)
                    snapshot['title'] = inst.order.title
                    snapshot['start_at'] = inst.order.start_at.strftime('%Y-%m-%d %H:%M')
                    snapshot['layer'] = inst.session_seat.layers if inst.session_seat else 0
                    if not inst.session_seat:
                        snapshot['seat'] = '无座'
                    session = SessionInfo.objects.get(id=inst.session_id)
                    if session.end_at <= timezone.now():
                        status = False
                        msg = '该场次已结束了'
                    else:
                        if inst.session.no != session_id and (
                                not session.main_session or session.main_session.no != session_id):
                            status = False
                            msg = '不是本场次门票，请核对后重试！'
                        else:
                            st = True
                            # 抖音改了逻辑要演出结束后一天才可以推
                            # if inst.order.pay_type == Receipt.PAY_TikTok_LP:
                            #     order = inst.order
                            #     from douyin import get_dou_yin
                            #     dy = get_dou_yin()
                            #     # order.item_order_info_list 取单号
                            #     # item_order_list = [{"item_order_id": inst.order.tiktok_order_id}]
                            #     poi_info = {"shop_name": order.session.tiktok_store.name,
                            #                 "ext_valid_shop_id": str(order.session.tiktok_store.id),
                            #                 "valid_poi_id_str": order.session.tiktok_store.supplier_ext_id}
                            #     params = dict(out_order_no=order.order_no, use_all=True, poi_info=json.dumps(poi_info))
                            #     ret = dy.push_delivery(params=params)
                            #     log.error(ret)
                            #     if ret['error_code'] == 0:
                            #         st = True
                            if st:
                                inst.status = cls.STATUS_CHECK
                                inst.check_at = timezone.now()
                                inst.check_user = user.account
                                inst.save(update_fields=['status', 'check_at', 'check_user'])
                                inst.order.is_check_num = inst.order.is_check_num + 1
                                if inst.order.is_check_num >= inst.order.multiply:
                                    inst.order.set_finish()
                                inst.order.save(update_fields=['is_check_num'])
                                # qs = cls.objects.filter(order=inst.order)
                                # total = qs.count()
                                # check_total = qs.filter(status=cls.STATUS_CHECK).count()
                                # if total == check_total:
                                #     inst.order.set_finish()
                                status = True
                                msg = ''
                            else:
                                status = False
                                msg = '核销失败'
            else:
                status = False
                msg = '该二维码已经核销过了'
        except Exception as e:
            redis.delete(key)
            status = False
            msg = '二维码无效'
        log.debug('结束验票，{},{}'.format(code, msg))
        return status, msg, snapshot


class TicketCheckRecordReceipt(Manager):
    """
    filter by role implicitly.
    """

    def get_queryset(self):
        return super().get_queryset().filter(status=TicketUserCode.STATUS_CHECK)


class TicketCheckRecord(TicketUserCode):
    class Meta:
        verbose_name_plural = verbose_name = '验票记录'
        proxy = True


class ShowCollectRecord(models.Model):
    user = models.ForeignKey(User, verbose_name='用户', on_delete=models.CASCADE)
    show = models.ForeignKey(ShowProject, verbose_name='演出项目', on_delete=models.CASCADE)
    is_collect = models.BooleanField('是否收藏', default=True)
    create_at = models.DateTimeField('收藏时间', null=True, blank=True)

    class Meta:
        verbose_name_plural = verbose_name = '用户演出收藏记录'
        unique_together = ['user', 'show']
        ordering = ['-pk']

    def __str__(self):
        return self.show.title[0:10]

    @classmethod
    def create_record(cls, user, show):
        inst, create = cls.objects.get_or_create(user=user, show=show)
        inst.create_at = timezone.now()
        fields = ['create_at']
        if not create:
            inst.is_collect = not inst.is_collect
            fields.append('is_collect')
        inst.save(update_fields=fields)

    def show_collect_copy_to_pika(self):
        from caches import get_pika_redis, show_collect_copy_key
        with get_pika_redis() as pika:
            name = '{}_{}'.format(self.user_id, self.show_id)
            if not self.is_collect:
                pika.delete(show_collect_copy_key, name)
            else:
                pika.hset(show_collect_copy_key, name, 1)


class PerformerFocusRecord(models.Model):
    user = models.ForeignKey(User, verbose_name='用户', on_delete=models.CASCADE)
    performer = models.ForeignKey(ShowPerformer, verbose_name='演员', on_delete=models.CASCADE)
    is_collect = models.BooleanField('是否收藏', default=True)
    create_at = models.DateTimeField('收藏时间', null=True, blank=True)

    class Meta:
        verbose_name_plural = verbose_name = '用户演员关注记录'
        unique_together = ['user', 'performer']
        ordering = ['-pk']

    def __str__(self):
        return self.performer.name

    @classmethod
    def create_record(cls, user, performer):
        inst, create = cls.objects.get_or_create(user=user, performer=performer)
        inst.create_at = timezone.now()
        fields = ['create_at']
        if not create:
            inst.is_collect = not inst.is_collect
            fields.append('is_collect')
        inst.save(update_fields=fields)
        performer.set_focus_num()


class SessionPushTiktokTask(models.Model):
    session = models.ForeignKey(SessionInfo, verbose_name='用户', on_delete=models.CASCADE)
    PUSH_DEFAULT = 1
    PUSH_APPROVE = 2
    PUSH_SUCCESS = 3
    PUSH_FAIL = 4
    PUSH_CHOICES = (
        (PUSH_DEFAULT, u'未推送'), (PUSH_APPROVE, u'推送中'), (PUSH_SUCCESS, u'已完成'), (PUSH_FAIL, u'推送失败'))
    status = models.IntegerField(u'状态', choices=PUSH_CHOICES, default=PUSH_DEFAULT)
    error_msg = models.CharField('错误信息', max_length=200, null=True, blank=True)
    create_at = models.DateTimeField('创建时间', auto_now_add=True)
    msg = models.CharField('推送描述', max_length=200, null=True, blank=True)

    class Meta:
        verbose_name_plural = verbose_name = '演出推送抖音商品记录'

    def __str__(self):
        return str(self.id)

    @classmethod
    def create_record(cls, session, msg):
        TicketFile.objects.filter(session_id=session.id, is_tiktok=True).exclude(
            push_status=TicketFile.PUSH_SUCCESS).update(push_status=TicketFile.PUSH_DEFAULT)
        qs = cls.objects.filter(session=session, status__in=[cls.PUSH_DEFAULT, cls.PUSH_APPROVE])
        if qs:
            return False, '该演出正在推送中，请推送完后再试'
        cls.objects.create(session=session, msg=msg)
        return True, None

    @classmethod
    def push_to_dou_yin(cls):
        close_old_connections()
        qs = cls.objects.filter(status=cls.PUSH_DEFAULT, create_at__lt=timezone.now() - timedelta(minutes=1))
        for inst in qs:
            inst.status = cls.PUSH_APPROVE
            inst.save(update_fields=['status'])
            session = inst.session
            try:
                st, msg = session.goods_push_dou_yin_new()
                if not st:
                    log.error(msg)
                    inst.status = inst.PUSH_FAIL
                    inst.error_msg = msg
                else:
                    inst.status = inst.PUSH_SUCCESS
                    inst.error_msg = None
            except Exception as e:
                log.error(e)
                inst.status = inst.PUSH_FAIL
                inst.error_msg = str(e)
            inst.save(update_fields=['status', 'error_msg'])
            if inst.status == inst.PUSH_FAIL:
                session.push_status = session.PUSH_FAIL
                session.save(update_fields=['push_status'])


class TicketOrderRefund(models.Model):
    STATUS_DEFAULT = 1
    STATUS_NEED_CONFIRM = 2
    STATUS_PAYING = 3
    STATUS_PAY_FAILED = 4
    STATUS_FINISHED = 5
    STATUS_CANCELED = 6
    STATUS_CHOICES = (
        (STATUS_DEFAULT, '待退款'), (STATUS_NEED_CONFIRM, '待第三方确认'),
        (STATUS_PAYING, '退款支付中'), (STATUS_PAY_FAILED, '退款支付失败'), (STATUS_FINISHED, '已完成'),
        (STATUS_CANCELED, '已取消'))
    status = models.IntegerField('状态', choices=STATUS_CHOICES, default=STATUS_DEFAULT)
    ST_WX = 1
    ST_TIKTOK = 2
    ST_CARD = 3
    ST_KS = 4
    ST_XHS = 5
    ST_CHOICES = ((ST_WX, '微信退款'), (ST_CARD, '剧场卡订单退款'))
    source_type = models.IntegerField('退款类型', choices=ST_CHOICES, default=ST_WX)
    user = models.ForeignKey(User, verbose_name='用户', on_delete=models.CASCADE)
    order = models.ForeignKey(TicketOrder, related_name='ticket_refund_apply', verbose_name='退款订单',
                              on_delete=models.CASCADE)
    out_refund_no = models.CharField(u'退款单号', max_length=100, default=randomstrwithdatetime_refund, unique=True,
                                     db_index=True)
    refund_amount = models.DecimalField(u'退款金额', max_digits=13, decimal_places=2, default=0)
    theater_amount = models.DecimalField(u'剧场卡订单退款数额', max_digits=13, decimal_places=2, default=0,
                                         help_text='仅用剧场卡支付有效')
    amount = models.DecimalField(u'实退金额', max_digits=13, decimal_places=2, default=0)
    return_reason = models.CharField('退款原因', max_length=100)
    error_msg = models.CharField('退款返回信息', max_length=1000, null=True, blank=True)
    transaction_id = models.CharField('支付订单号', max_length=32, null=True, blank=True, help_text='收据里的')
    refund_id = models.CharField('退款方退款单号', max_length=32, null=True, blank=True)
    return_code = models.CharField('微信通信结果', max_length=20, null=True, blank=True)
    result_code = models.CharField('微信返回结果', max_length=20, null=True, blank=True)
    create_at = models.DateTimeField('创建时间', auto_now_add=True)
    confirm_at = models.DateTimeField('确认时间', null=True, blank=True)
    finish_at = models.DateTimeField('完成时间', null=True, blank=True)
    op_user = models.ForeignKey(User, verbose_name='操作人员', on_delete=models.SET_NULL, blank=True, null=True,
                                related_name='op_user')

    class Meta:
        verbose_name_plural = verbose_name = '退款记录'
        ordering = ['-pk']

    def get_refund_notify_url(self):
        return wx_refund_notify_url

    @classmethod
    def create_record(cls, order, refund_amount, return_reason, source_type):
        can_refund = False
        receipt = order.receipt
        theater_amount = 0
        if source_type == cls.ST_CARD:
            can_refund = receipt.amount > 0 and receipt.status == Receipt.STATUS_FINISHED and receipt.transaction_id or receipt.amount == 0
            theater_amount = order.card_jc_amount
        elif order.receipt.status == Receipt.STATUS_FINISHED and order.receipt.transaction_id:
            can_refund = True
        if can_refund:
            return True, cls.objects.create(user=order.user, order=order, refund_amount=refund_amount,
                                            return_reason=return_reason, theater_amount=theater_amount,
                                            source_type=source_type, transaction_id=order.receipt.transaction_id)
        else:
            return False, '该订单未付款不能退款'

    def refund_stl(self):
        try:
            order = self.order
            order.add_stl(is_refund=True)
            order.cps_stl(-1)
            from statistical.models import TotalStatistical
            TotalStatistical.change_ticket_refund_stl(refund_num=1, refund_amount=self.refund_amount)
        except Exception as e:
            log.error('refund_stl {}'.format(str(e)))

    def ticket_refund_back(self):
        from shopping_points.models import UserCommissionChangeRecord
        source_type = UserCommissionChangeRecord.SOURCE_TYPE_REFUND
        desc = '订单号:{},退款单号:{},退款扣除'.format(self.order.order_no, self.out_refund_no)
        if self.order:
            order = self.order
            self.refund_stl()
            if order.agent:
                qs = UserCommissionChangeRecord.objects.filter(order_id=self.order_id)
                for inst in qs:
                    amount = self.refund_amount * inst.amount / order.award_amount
                    if amount > 0.01:
                        account = inst.account
                        UserCommissionChangeRecord.add_record(account, -amount, source_type, desc,
                                                              status=UserCommissionChangeRecord.STATUS_CAN_WITHDRAW,
                                                              order=order)
                if qs:
                    # 每日代理销售记录退款扣除
                    order.change_agent_amount(is_refund=True)
            if order.tiktok_douyinid and order.tiktok_commission_amount > 0:
                # 每日达人销售记录退款扣除
                c_amount = -order.refund_amount * order.tiktok_commission_amount / order.actual_amount
                self.order.change_cps_agent_amount(self.order.source_type, c_amount, -order.refund_amount,
                                                   TiktokUser.ST_DY)
            if order.card_jc_amount > 0 and order.pay_type == Receipt.PAY_CARD_JC:
                TheaterCardChangeRecord.add_record(user=order.user,
                                                   source_type=TheaterCardChangeRecord.SOURCE_TYPE_REFUND,
                                                   amount=order.card_jc_amount, ticket_order=order)
            if order.pay_type == Receipt.PAY_CARD_JC and order.card_jc_amount > 0:
                # 退款扣减优惠金额
                user_card = TheaterCardUserRecord.get_inst(self.user)
                amount = order.amount - order.actual_amount
                if amount > 0:
                    user_card.add_discount_total(-amount)

    @atomic
    def set_confirm(self, op_user=None):
        st = False
        msg = '退款类型不对'
        is_finish = False
        if self.source_type == self.ST_TIKTOK:
            st = self.tiktok_refund()
            msg = self.error_msg
        elif self.source_type == self.ST_KS:
            from kuaishou_wxa.models import KsOrderSettleRecord
            st = KsOrderSettleRecord.ks_refund(self)
            msg = self.error_msg
        elif self.source_type == self.ST_WX:
            st = self.wx_refund()
            msg = self.error_msg
        elif self.source_type == self.ST_XHS:
            from xiaohongshu.models import XhsOrder
            st = XhsOrder.xhs_refund(self)
            msg = self.error_msg
        elif self.source_type == self.ST_CARD:
            if self.order.pay_type == Receipt.PAY_CARD_JC:
                if self.order.receipt.pay_type == Receipt.PAY_WeiXin_LP:
                    st = self.wx_refund()
                    msg = self.error_msg
                elif self.order.receipt.pay_type == Receipt.PAY_CARD_JC:
                    st = True
                    msg = None
                    is_finish = True
        if st:
            fields = ['status', 'confirm_at']
            self.status = self.STATUS_PAYING
            self.confirm_at = timezone.now()
            if op_user:
                self.op_user = op_user
                fields.append('op_user')
            self.save(update_fields=fields)
            self.order_refund_back()
            if is_finish:
                self.set_finished(0)
        return st, msg

    def order_refund_back(self):
        self.order.session.update_actual_amount(-self.refund_amount)
        if self.order.order_type in [TicketOrder.TY_HAS_SEAT, TicketOrder.TY_NO_SEAT]:
            self.order.release_seat()
            self.order.cancel_code()
        if self.order.is_lock_seat:
            # 手动出票的
            self.order.cancel_lock_seats()

    def check_refund(self):
        from douyin import get_dou_yin
        dy = get_dou_yin()
        ret = dy.query_refund(self.refund_id)
        if ret['refund_status'] == 'SUCCESS':
            self.set_finished(ret['refund_total_amount'])
        elif ret['refund_status'] == 'PROCESSING':
            pass
        elif ret['refund_status'] == 'FAIL':
            self.set_fail(ret['message'])

    def tiktok_refund(self):
        # booking = TicketBooking.objects.filter(order_id=self.order_id, status=TicketBooking.STATUS_SUCCESS).first()
        # if booking:
        #     ret, msg = booking.set_cancel(self.return_reason)
        #     if not ret:
        #         log.error(msg)
        #         raise CustomAPIException(msg)
        config = get_config()
        from douyin import get_dou_yin
        dy = get_dou_yin()
        params = {"id": self.order.id}
        item_order_info_list = self.order.item_order_info_list
        if not item_order_info_list:
            return False
        item_order_info_list = json.loads(item_order_info_list)
        item_order_detail = []
        rest_amount = int(self.refund_amount * 100)
        is_end = False
        for item_order_info in item_order_info_list:
            # 增加退款金额，抖音只能按明细退
            has_refund_amount = item_order_info.get('has_refund_amount') or 0
            rest_refund_amount = int(item_order_info['price']) - has_refund_amount
            if rest_refund_amount >= rest_amount:
                items = {"item_order_id": item_order_info['item_order_id'],
                         "refund_amount": rest_amount}
                if has_refund_amount > 0:
                    item_order_info['has_refund_amount'] = has_refund_amount + rest_amount
                else:
                    item_order_info['has_refund_amount'] = rest_amount
                is_end = True
            else:
                rest_amount = rest_amount - rest_refund_amount
                items = {"item_order_id": item_order_info['item_order_id'],
                         "refund_amount": rest_refund_amount}
                item_order_info['has_refund_amount'] = item_order_info['price']
            item_order_detail.append(items)
            if is_end:
                break
        self.order.item_order_info_list = json.dumps(item_order_info_list)
        self.order.save(update_fields=['item_order_info_list'])
        data = {
            "out_order_no": self.order.order_no,
            "out_refund_no": self.out_refund_no,
            "cp_extra": self.out_refund_no,
            "order_entry_schema": {
                "path": tiktok_order_detail_url,
                "params": json.dumps(params)
            },
            "notify_url": '{}{}'.format(config['template_url'], tiktok_refund_notify_url),
            "item_order_detail": item_order_detail
        }
        ret = dy.create_refund(data)
        self.result_code = ret.get('error_code') or 0
        if self.result_code == 0:
            self.refund_id = ret['refund_id']
            self.save(update_fields=['refund_id'])
            return True
        else:
            self.status = self.STATUS_PAY_FAILED
            self.error_msg = ret['description']
            self.save(update_fields=['status', 'error_msg'])
        return False

    def set_cancel(self, op_user):
        self.status = self.STATUS_CANCELED
        self.confirm_at = timezone.now()
        self.op_user = op_user
        self.save(update_fields=['status', 'confirm_at', 'op_user'])
        self.return_order_status()

    def return_order_status(self, is_fail=False):
        if self.order.refund_amount > 0:
            self.order.status = self.order.status_before_refund if not is_fail else TicketOrder.STATUS_REFUNDED_FAIL
            self.order.refund_amount -= self.refund_amount
            self.order.save(update_fields=['status', 'refund_amount'])
            item_order_info_list = self.order.item_order_info_list
            if item_order_info_list:
                item_order_info_list = json.loads(item_order_info_list)
                for item_order_info in item_order_info_list:
                    item_order_info['has_refund_amount'] = 0
                self.order.item_order_info_list = json.dumps(item_order_info_list)
                self.order.save(update_fields=['item_order_info_list'])

    def wx_refund(self):
        if not self.order.receipt.transaction_id:
            self.order.receipt.update_info()
        from mall.pay_service import get_mp_pay_client
        mp_pay_client = get_mp_pay_client(self.order.receipt.pay_type, self.order.receipt.wx_pay_config)
        config = get_config()
        refund_notify_url = config['uri'] + self.get_refund_notify_url()
        result = mp_pay_client.ticket_refund(self, refund_notify_url=refund_notify_url)
        self.return_code = result.get('return_code')
        self.result_code = result.get('result_code')
        self.error_msg = '{}, {}'.format(result.get('return_msg'), result.get('err_code_des'))
        self.refund_id = result.get('refund_id', None)
        self.save(update_fields=['return_code', 'result_code', 'error_msg', 'refund_id'])
        if self.return_code == self.result_code == 'SUCCESS':
            return True
        return False

    def set_fail(self, msg=None):
        self.status = self.STATUS_PAY_FAILED
        self.finish_at = timezone.now()
        self.error_msg = msg
        self.save(update_fields=['status', 'error_msg', 'finish_at'])
        self.return_order_status(True)

    def set_finished(self, amount=0):
        """
        设置为完成:
        1.在退款支付通知回调中
        2.后台手动
        :return:
        """
        from decimal import Decimal
        self.status = self.STATUS_FINISHED
        self.finish_at = timezone.now()
        amount = Decimal(amount) / 100
        self.amount = amount if amount > 0 else self.refund_amount
        self.save(update_fields=['status', 'finish_at', 'amount'])
        self.order.status = TicketOrder.STATUS_REFUNDED
        self.order.save(update_fields=['status'])

    def biz_refund(self, op_user):
        if hasattr(self.order, 'cy_order'):
            from caiyicloud.models import CyOrderRefund
            st, msg = CyOrderRefund.confirm_refund(self)
            if st:
                self.status = self.STATUS_NEED_CONFIRM
                self.op_user = op_user
                self.save(update_fields=['status', 'op_user'])
        else:
            st, msg = self.set_confirm(op_user)
        return st, msg


def validate_commission_rate(min=0, max=2900):
    """
    检查整数范围
    :param min:
    :param max:
    :return:
    """

    def validate_positive_rate(value):
        """
        :return:
        """
        if not (min <= value and (max is None or (max and value <= max))):
            raise ValidationError('必须大于%s%s' % (min, (('大于%s' % max) if max else '')))
        return value

    return validate_positive_rate


class CommonPlanCps(models.Model):
    title = models.CharField('计划名称', null=True, max_length=20)
    commission_rate = models.IntegerField(u'商品的抽佣率(万分数)', help_text='小程序商品可选值范围：100-2900', default=100,
                                          validators=[validate_commission_rate])
    ST_DEFAULT = 0
    ST_USING = 1
    ST_PAUSED = 2
    ST_CANCEL = 3
    ST_APPROVE = 5
    ST_FAIL = 4
    ST_CHOICES = ((ST_DEFAULT, '待推送'), (ST_APPROVE, '推送中'), (ST_USING, '进行中'), (ST_PAUSED, '暂停中'), (ST_CANCEL, '已关闭'),
                  (ST_FAIL, '推送失败'))
    status = models.IntegerField('状态', choices=ST_CHOICES, default=ST_DEFAULT,
                                 help_text='状态为已关闭，不支持修改为其他值')
    CT_VIDEO = 1
    CT_LIVE = 2
    CT_ALL = 3
    CT_CHOICES = ((CT_VIDEO, '仅短视频'), (CT_LIVE, '仅直播间'), (CT_ALL, '短视频和直播间'))
    content_type = models.IntegerField('商品支持的达人带货场景', choices=CT_CHOICES, default=CT_ALL)
    goods = models.ManyToManyField(SessionInfo, verbose_name='抖音内部商品',
                                   limit_choices_to=models.Q(push_status=SessionInfo.PUSH_SUCCESS,
                                                             status=SessionInfo.STATUS_ON),
                                   help_text='每个商品只能仅有一个计划')
    goods_add = models.ManyToManyField(SessionInfo, related_name='+', verbose_name='增加抖音内部商品', blank=True)
    create_at = models.DateTimeField('创建时间', auto_now_add=True)
    error_msg = models.TextField('推送信息', null=True, blank=True)

    class Meta:
        verbose_name_plural = verbose_name = '通用佣金计划'

    def __str__(self):
        return str(self.id)

    def clean(self):
        if self.commission_rate < 100:
            raise ValidationError('商品的抽佣率须大于100')

    @classmethod
    def common_plan_to_dou_yin(cls):
        close_old_connections()
        qs = cls.objects.filter(status=cls.ST_DEFAULT)
        from douyin import get_dou_yin
        dy = get_dou_yin()
        for inst in qs:
            inst.status = cls.ST_APPROVE
            inst.save(update_fields=['status'])
            error_msg = ''
            status = inst.ST_FAIL
            for good in inst.goods.all():
                session = good
                if session.product_id and session.status == SessionInfo.STATUS_ON:
                    try:
                        st = dy.save_common_plan(inst.commission_rate, inst.content_type, session.product_id,
                                                 plan_id=session.plan_id if session.plan_id else 0)
                        if st.get('plan_id'):
                            status = inst.ST_USING
                            if not session.plan_id:
                                session.plan_id = st.get('plan_id')
                                session.save(update_fields=['plan_id'])
                        else:
                            error_msg += '{}推送失败,'.format(str(session))
                            continue
                    except Exception as e:
                        log.error(e)
                        # inst.status = inst.ST_FAIL
                        # inst.error_msg = str(e)
                        error_msg += '{}推送失败{},'.format(str(good.session), str(e))
                        continue
            inst.error_msg = error_msg
            inst.status = status
            inst.save(update_fields=['status', 'error_msg'])

    def update_common_plan_status(self, status):
        from douyin import get_dou_yin
        dy = get_dou_yin()
        plan_update_list = []
        for good in self.goods.all():
            plan_update_list.append(dict(plan_id=good.plan_id, status=status))
        try:
            ret = dy.update_common_plan_status(plan_update_list)
            desc = ''
            st = True
            log.debug(ret)
            if ret['fail_plan_id_list']:
                st = False
                desc = '执行成功，其中修改失败的商品:{}'.format(','.join(ret['fail_plan_id_list']))
            self.status = status
            self.save(update_fields=['status'])
        except Exception as e:
            st = False
            desc = str(e)
        return st, desc


class TiktokUser(models.Model):
    name = models.CharField('账号昵称', max_length=50)
    ST_DY = 1
    ST_KS = 2
    ST_WX = 3
    ST_CHOICES = ((ST_DY, '抖音'), (ST_KS, '快手'), (ST_WX, '微信视频号'))
    source_type = models.IntegerField('平台', choices=ST_CHOICES, default=ST_DY)
    tiktok_no = models.CharField('平台账号ID', max_length=50, help_text='团购带货达人的平台账号号ID')
    create_at = models.DateTimeField('创建时间', auto_now_add=True)

    class Meta:
        verbose_name_plural = verbose_name = '团购带货达人管理'

    def __str__(self):
        return '{}({})'.format(self.name, self.tiktok_no)


class CpsDirectional(models.Model):
    plan_name = models.CharField('计划名称', max_length=20)
    plan_id = models.CharField('计划ID', max_length=50, editable=False, null=True, blank=True)
    merchant_phone = models.CharField('计划联系人手机号码', max_length=11)
    ST_DEFAULT = 0
    ST_USING = 1
    ST_FINISH = 2
    ST_CANCEL = 3
    ST_FAIL = 4
    ST_APPROVE = 5
    ST_CHOICES = ((ST_DEFAULT, '待推送'), (ST_APPROVE, '推送中'), (ST_USING, '进行中'), (ST_FINISH, '已完成'), (ST_CANCEL, '已取消'),
                  (ST_FAIL, '推送失败'))
    status = models.IntegerField('状态', choices=ST_CHOICES, default=ST_DEFAULT,
                                 help_text='当短视频定向佣金计划开始时间小于当前时间时，不支持取消')
    tiktok_users = models.ManyToManyField(TiktokUser, verbose_name='已选择定向达人', related_name='+',
                                          help_text='出于安全考虑，请求参数中含有“非团购带货达人的抖音号ID”会被视为异常，异常请求当天超过50次，'
                                                    '则禁止应用当日再调用本接口。最多200个', blank=True)
    tiktok_users_add = models.ManyToManyField(TiktokUser, related_name='+', verbose_name='增加定向达人', blank=True,
                                              help_text='出于安全考虑，请求参数中含有“非团购带货达人的抖音号ID”会被视为异常，异常请求当天超过50次，则禁止应用当日再调用本接口')
    tiktok_users_delete = models.ManyToManyField(TiktokUser, related_name='+', verbose_name='取消定向达人', blank=True,
                                                 help_text='仅当满足以下两个条件时，开发者才能取消达人：达人未发布视频当前时间 < 视频发布时间-1day。')
    error_msg = models.CharField('推送信息', null=True, blank=True, editable=False, max_length=300)
    create_at = models.DateTimeField('创建时间', auto_now_add=True, null=True)

    class Meta:
        abstract = True

    def update_oriented_plan_status(self, status):
        from douyin import get_dou_yin
        dy = get_dou_yin()
        plan_update_list = []
        plan_update_list.append(dict(plan_id=self.plan_id, status=status))
        try:
            st, desc = dy.update_oriented_plan_status(plan_update_list)
        except Exception as e:
            st = False
            desc = str(e)
        if st:
            self.status = status
            self.save(update_fields=['status'])
        return st, desc

    def set_cancel(self):
        self.status = self.ST_CANCEL
        self.save(update_fields=['status'])


class LiveRoomCps(CpsDirectional):
    class Meta:
        verbose_name_plural = verbose_name = '直播间定向佣金计划'
        ordering = ['-pk']

    @classmethod
    def live_room_plan_to_dou_yin(cls):
        close_old_connections()
        qs = cls.objects.filter(status=cls.ST_DEFAULT)
        from douyin import get_dou_yin
        dy = get_dou_yin()
        for inst in qs:
            fields = ['status', 'error_msg']
            douyin_id_list = list(inst.tiktok_users.all().values_list('tiktok_no', flat=True))
            product_list = []
            inst.status = cls.ST_APPROVE
            inst.save(update_fields=['status'])
            items = LiveRoomCpsItem.objects.filter(plan=inst)
            for item in items:
                session = item.good
                if session.product_id and session.status == SessionInfo.STATUS_ON:
                    product_list.append(
                        dict(product_id=session.product_id, commission_rate=item.commission_rate))
            if not product_list or not douyin_id_list:
                continue
            try:
                st = dy.save_live_oriented_plan(inst.plan_name, inst.merchant_phone, douyin_id_list[:200],
                                                product_list[:50],
                                                plan_id=inst.plan_id if inst.plan_id else 0)
                if st.get('plan_id'):
                    inst.status = inst.ST_USING
                    inst.plan_id = st.get('plan_id')
                    inst.error_msg = '推送成功'
                    fields.append('plan_id')
                    items.update(is_push=True)
                else:
                    inst.status = inst.ST_FAIL
                    inst.error_msg = '推送失败'
            except Exception as e:
                log.error(e)
                inst.status = inst.ST_FAIL
                inst.error_msg = str(e)
            inst.save(update_fields=fields)


class LiveRoomCpsItem(models.Model):
    plan = models.ForeignKey(LiveRoomCps, verbose_name='直播间定向佣金计划', on_delete=models.CASCADE)
    good = models.ForeignKey(SessionInfo, verbose_name='商品', on_delete=models.CASCADE, help_text='仅支持上架且审核通过的正式品',
                             limit_choices_to=models.Q(push_status=SessionInfo.PUSH_SUCCESS,
                                                       status=SessionInfo.STATUS_ON))
    commission_rate = models.IntegerField(u'定向抽佣率(万分数)', help_text='可设置为0-2900', default=0,
                                          validators=[validate_commission_rate])
    is_push = models.BooleanField('是否已推送', default=False)

    class Meta:
        verbose_name_plural = verbose_name = '直播间佣金计划商品配置'
        unique_together = ['plan', 'good']

    def __str__(self):
        return str(self.good)


class ShortVideoCps(CpsDirectional):
    start_time = models.DateTimeField('可上传投稿视频的开始时间', help_text='必须大于当前时间，如果修改必须小于旧的开始时间')
    end_time = models.DateTimeField('可上传投稿视频的结束时间', help_text='必须大于开始时间，如果修改必须大于旧的结束时间')
    commission_duration = models.IntegerField('佣金有效期', help_text='单位是天,不可修改', default=1)
    change_start = models.BooleanField('是否修改开始时间', default=False, editable=False)
    change_end = models.BooleanField('是否修改结束时间', default=False, editable=False)

    class Meta:
        verbose_name_plural = verbose_name = '短视频定向佣金计划'
        ordering = ['-pk']

    @classmethod
    def short_video_plan_to_dou_yin(cls):
        close_old_connections()
        qs = cls.objects.filter(status=cls.ST_DEFAULT)
        from douyin import get_dou_yin
        from common.utils import get_timestamp
        dy = get_dou_yin()
        for inst in qs:
            fields = ['status', 'error_msg']
            douyin_id_list = list(inst.tiktok_users.all().values_list('tiktok_no', flat=True))
            product_list = []
            items = ShortVideoCpsItem.objects.filter(plan=inst)
            for item in items:
                session = item.good
                if session.product_id and session.status == SessionInfo.STATUS_ON:
                    product_list.append(
                        dict(product_id=session.product_id, commission_rate=item.commission_rate))
            if not product_list or not douyin_id_list:
                continue
            inst.status = cls.ST_APPROVE
            inst.save(update_fields=['status'])
            try:
                st = dy.save_video_oriented_plan(inst.plan_name, inst.commission_duration,
                                                 inst.merchant_phone, douyin_id_list[:200],
                                                 product_list[:50],
                                                 int(get_timestamp(
                                                     inst.start_time) / 1000) if inst.change_start else None,
                                                 int(get_timestamp(inst.end_time) / 1000) if inst.change_end else None,
                                                 plan_id=inst.plan_id if inst.plan_id else 0)
                if st.get('plan_id'):
                    inst.status = inst.ST_USING
                    inst.plan_id = st.get('plan_id')
                    inst.error_msg = '推送成功'
                    fields.append('plan_id')
                    items.update(is_push=True)
                else:
                    inst.status = inst.ST_FAIL
                    inst.error_msg = '推送失败'
            except Exception as e:
                log.error(e)
                inst.status = inst.ST_FAIL
                inst.error_msg = str(e)
            inst.save(update_fields=fields)


class ShortVideoCpsItem(models.Model):
    plan = models.ForeignKey(ShortVideoCps, verbose_name='短视频定向佣金计划', on_delete=models.CASCADE)
    good = models.ForeignKey(SessionInfo, verbose_name='商品', on_delete=models.CASCADE, help_text='仅支持上架且审核通过的正式品',
                             limit_choices_to=models.Q(push_status=SessionInfo.PUSH_SUCCESS,
                                                       status=SessionInfo.STATUS_ON))
    commission_rate = models.IntegerField(u'定向抽佣率(万分数)', help_text='可设置为0-2900', default=0,
                                          validators=[validate_commission_rate])
    is_push = models.BooleanField('是否已推送', default=False)

    class Meta:
        verbose_name_plural = verbose_name = '直播间佣金计划商品配置'
        unique_together = ['plan', 'good']

    def __str__(self):
        return str(self.good)


class TicketBooking(models.Model):
    STATUS_DEFAULT = 1
    STATUS_CANCELED = 2
    STATUS_SUCCESS = 3
    STATUS_FAIL = 4
    STATUS_CHOICES = ((STATUS_DEFAULT, '未预约'), (STATUS_SUCCESS, '已预约'), (STATUS_CANCELED, '已取消'), (STATUS_FAIL, '预约失败'))
    status = models.IntegerField('状态', choices=STATUS_CHOICES, default=STATUS_DEFAULT)
    user = models.ForeignKey(User, verbose_name='用户', on_delete=models.CASCADE)
    order = models.ForeignKey(TicketOrder, verbose_name='关联订单', on_delete=models.SET_NULL, null=True, blank=True)
    order_no = models.CharField(u'关联订单号', max_length=128)
    tiktok_order_id = models.CharField(u'抖音订单号', max_length=128)
    out_book_no = models.CharField(u'预约单号', max_length=100, default=randomstrwithdatetime_booking, unique=True,
                                   db_index=True)
    book_id = models.CharField(u'抖音预约单号', max_length=64, null=True, blank=True)
    create_at = models.DateTimeField('创建时间', auto_now_add=True)
    cancel_at = models.DateTimeField('取消时间', null=True, blank=True)
    cancel_reason = models.CharField('取消理由', null=True, blank=True, max_length=100)
    err_msg = models.CharField(u'预约返回信息', max_length=100, null=True, blank=True)
    err_logid = models.CharField(u'抖音订单号', max_length=50, null=True, blank=True)

    class Meta:
        verbose_name_plural = verbose_name = '抖音预约单'
        ordering = ['-pk']

    def set_book_id(self):
        from douyin import get_dou_yin
        dy = get_dou_yin()
        st, ret = dy.query_book(self.order.tiktok_order_id)
        log.error(ret)
        if st and ret:
            self.book_id = ret[0]['book_id']
            self.save(update_fields=['book_id'])

    def set_status(self, status, err_msg, err_logid, book_id=None):
        self.status = status
        self.err_msg = err_msg
        self.err_logid = err_logid
        self.book_id = book_id
        self.save(update_fields=['status', 'err_msg', 'err_logid', 'book_id'])

    def set_cancel(self, reason):
        # 先推送抖音取消
        from douyin import get_dou_yin
        dy = get_dou_yin()
        try:
            ret, msg = dy.merchant_cancel_book(self.book_id, reason)
        except Exception as e:
            ret = False
            msg = e
        if ret:
            self.status = self.STATUS_CANCELED
            self.cancel_at = timezone.now()
            self.cancel_reason = reason
            self.save(update_fields=['status', 'cancel_at', 'cancel_reason'])
        return ret, msg

    @classmethod
    @atomic
    def create_booking_new(cls, order):
        if not order.tiktok_order_id:
            raise CustomAPIException('抖音下单失败')
        session = order.session
        if session.tiktok_store:
            book_info = []
            is_create = False
            snapshot = json.loads(order.snapshot)
            from common.utils import get_timestamp
            tiktok_store = session.tiktok_store
            booking = cls.objects.filter(tiktok_order_id=order.tiktok_order_id).first()
            if booking and booking.status == cls.STATUS_SUCCESS:
                raise CustomAPIException('该订单已经下过抖音预约单了,{}'.format(order.order_no))
            if not booking:
                is_create = True
                booking = cls.objects.create(user=order.user, order=order, order_no=order.order_no,
                                             tiktok_order_id=order.tiktok_order_id)
            user_info = [{"name": order.name, "phone": order.mobile}]
            item_data = dict(booking=booking, goods_id=None,
                             ext_shopid=tiktok_store.id,
                             poi_id=tiktok_store.supplier_ext_id,
                             shop_name=tiktok_store.name,
                             book_start_time=session.start_at,
                             book_end_time=session.end_at, sku_info=None,
                             user_info=json.dumps(user_info))
            for price_list in snapshot['price_list']:
                multiply = price_list['multiply']
                ticket_level = TicketFile.objects.get(id=int(price_list['level_id']))
                # sku_info = {
                #     "price": float(ticket_level.price),
                #     "skuId": ticket_level.get_out_id(),
                #     "skuType": 2
                # }
                sku_info = {
                    "price": 0,
                    "skuId": "",
                    "skuType": 0
                }
                goods_id = ticket_level.get_product_id()
                if not goods_id:
                    raise CustomAPIException('未找到抖音商品ID')
                item_data['goods_id'] = goods_id
                item_data['sku_info'] = json.dumps(sku_info)
                # if is_create:
                #     booking_item = TicketBookingItem.objects.create(**item_data)
                # else:
                #     booking_item = TicketBookingItem.objects.filter(booking_id=booking.id, goods_id=goods_id).first()
                dd = {
                    "bookStartTime": get_timestamp(session.start_at),
                    "bookEndTime": get_timestamp(session.end_at),
                    "bookRangeType": 0,  # 非日历房不传，或者传0
                    "outShopId": str(tiktok_store.id),
                    "goodsId": goods_id,
                    "poiId": tiktok_store.supplier_ext_id,
                    "shopName": tiktok_store.name,
                    "userInfo": user_info,
                    "skuInfo": sku_info
                }
                if order.order_type == order.TY_HAS_SEAT:
                    book_info.append(dd)
                    if is_create:
                        TicketBookingItem.objects.create(**item_data)
                elif order.order_type == order.TY_NO_SEAT:
                    for i in list(range(int(multiply))):
                        book_info.append(dd)
                        if is_create:
                            TicketBookingItem.objects.create(**item_data)
            data = {
                "orderId": booking.tiktok_order_id,
                "outBookNo": booking.out_book_no,
                "bookInfo": book_info,
                # 加价信息
                # "markupInfo": {
                #     "totalAmount": 2,
                # },
            }
            return data
        else:
            raise CustomAPIException('未配置抖音店铺')

    @classmethod
    def create_book(cls, order, data=None):
        st = False
        if order.user.openid_tiktok:
            if not data:
                data = cls.create_booking_new(order)
            from douyin import get_dou_yin
            dy = get_dou_yin()
            item_book_info_list = []
            for dd in data['bookInfo']:
                item_book_info_list.append(
                    {
                        "book_end_time": get_timestamp(timezone.now() + timedelta(hours=2)),
                        "book_start_time": get_timestamp(timezone.now() + timedelta(hours=1)),
                        "ext_shop_id": dd['outShopId'],
                        "goods_id": dd['goodsId'],
                        "poi_id": dd['poiId'],
                        "shop_name": dd['shopName'],
                        "user_info_list": dd['userInfo'],
                        # "book_sku_info":{
                        #     "price": dd['skuInfo']['price'],
                        #     "sku_id": dd['skuInfo']['skuId'],
                        #     "sku_id_type": dd['skuInfo']['skuType']
                        # }
                    })
            status, ret = dy.create_book(data['orderId'], data['outBookNo'], order.user.openid_tiktok,
                                         item_book_info_list)
            booking = cls.objects.get(out_book_no=data['outBookNo'])
            if status:
                booking.set_status(cls.STATUS_SUCCESS, "success", "0", ret['book_id'])
                st = True
            else:
                log.error(ret)
                booking.set_status(cls.STATUS_FAIL, ret['sub_description'], ret['logid'])
        else:
            log.error('找不到openid')
        return st


class TicketBookingItem(models.Model):
    booking = models.ForeignKey(TicketBooking, verbose_name='抖音预约单', on_delete=models.CASCADE)
    goods_id = models.CharField('商品id', max_length=50)
    ext_shopid = models.IntegerField('预约门店的外部店铺id')
    poi_id = models.CharField('预约门店的poiId', max_length=50)
    shop_name = models.CharField('预约门店的名称', max_length=100)
    book_start_time = models.DateTimeField('预订的开始时间')
    book_end_time = models.DateTimeField('预订结束时间')
    sku_info = models.TextField('留资信息', null=True, blank=True)
    user_info = models.TextField('预约商品信息', null=True, blank=True)

    class Meta:
        verbose_name_plural = verbose_name = '预约信息'
        ordering = ['-pk']


class MaiZuoTask(models.Model):
    session = models.ForeignKey(SessionInfo, verbose_name='演出场次', on_delete=models.CASCADE)
    PUSH_DEFAULT = 1
    PUSH_APPROVE = 2
    PUSH_SUCCESS = 3
    PUSH_FAIL = 4
    PUSH_CHOICES = (
        (PUSH_DEFAULT, u'未拉取'), (PUSH_APPROVE, u'更新中'), (PUSH_SUCCESS, u'已完成'), (PUSH_FAIL, u'更新失败'))
    status = models.IntegerField(u'状态', choices=PUSH_CHOICES, default=PUSH_DEFAULT)
    error_msg = models.CharField('错误信息', max_length=200, null=True, blank=True)
    create_at = models.DateTimeField('创建时间', auto_now_add=True)

    class Meta:
        verbose_name_plural = verbose_name = '麦座拉取记录'

    def __str__(self):
        return str(self.id)

    @classmethod
    def create_record(cls, session_id):
        inst = cls.objects.filter(session_id=session_id).first()
        if inst and inst.status in [cls.PUSH_DEFAULT, cls.PUSH_APPROVE]:
            return inst
        else:
            return cls.objects.create(session_id=session_id)

    @classmethod
    def login_task(cls):
        from maizuo import get_mai_zuo
        from mp.models import MaiZuoAccount
        qs = MaiZuoAccount.objects.all()
        for inst in qs:
            mz = get_mai_zuo(inst.name, inst.password)
            mz.login_task()

    @classmethod
    def pull_record(cls):
        close_old_connections()
        qs = cls.objects.filter(status=cls.PUSH_DEFAULT)
        error_msg = ''
        for inst in qs:
            inst.status = cls.PUSH_APPROVE
            inst.save(update_fields=['status'])
            inst.session.set_mz_status(SessionInfo.PULL_APPROVE)
            status = cls.PUSH_FAIL
            mz_status = SessionInfo.PULL_FAIL
            try:
                seat_list = SessionSeat.objects.filter(session_id=inst.session_id)
                seat_dict = dict()
                for kk in seat_list:
                    row, box_no_special = kk.get_row_and_box_no_special()
                    name, key = SessionSeat.get_pika_mz_seat_key(inst.session_id, kk.layers, row,
                                                                 kk.showCol, box_no_special)
                    seat_dict[key] = kk
                st, error_msg, _ = inst.session.get_mai_zuo_data_new(is_init=True, seat_dict=seat_dict)
                if st:
                    status = cls.PUSH_SUCCESS
                    mz_status = SessionInfo.PULL_SUCCESS
            except Exception as e:
                log.error(e)
            inst.status = status
            inst.error_msg = error_msg
            inst.save(update_fields=['status'])
            inst.session.set_mz_status(mz_status)


class DownLoadTask(models.Model):
    name = models.CharField('任务名称', max_length=20)
    ST_DEFAULT = 0
    ST_DO = 1
    ST_FAIL = 2
    ST_SUCCESS = 3
    ST_CHOICES = ((ST_DEFAULT, '未执行'), (ST_DO, '导出中'), (ST_FAIL, '导出失败'), (ST_SUCCESS, '导出完成'))
    status = models.IntegerField('状态', choices=ST_CHOICES, default=ST_DEFAULT)
    TY_ORDER = 1
    TY_CHOICES = ((TY_ORDER, '演出订单'),)
    source_type = models.IntegerField('类型', choices=TY_CHOICES, default=TY_ORDER)
    export_file = models.FileField('文件', null=True, blank=True)
    create_at = models.DateTimeField('导出时间', auto_now_add=True)

    class Meta:
        verbose_name_plural = verbose_name = '导出记录'
        ordering = ['-pk']

    @classmethod
    def create_record(cls, name):
        return cls.objects.create(name=name)

    def pika_down_key(self):
        from caches import pika_down_key
        return pika_down_key.format(self.id)

    @classmethod
    def do_task(cls):
        close_old_connections()
        inst = cls.objects.filter(status=cls.ST_DEFAULT).first()
        if inst:
            inst.status = cls.ST_DO
            inst.save(update_fields=['status'])
            if inst.source_type == cls.TY_ORDER:
                from caches import get_pika_redis
                with get_pika_redis() as redis:
                    key = inst.pika_down_key()
                    ids = redis.get(key)
                    export_file = None
                    if ids:
                        ids = json.loads(ids)
                        export_file = TicketOrder.down_to_excel(ids)
                    if export_file:
                        inst.export_file = export_file
                        inst.status = cls.ST_SUCCESS
                        inst.save(update_fields=['export_file', 'status'])
                        # 执行完删除
                        redis.delete(key)
                    else:
                        inst.status = cls.ST_FAIL
                        inst.save(update_fields=['status'])
                    log.warning('导出完成')


class MaiZuoLoginLog(models.Model):
    msg = models.CharField('内容', max_length=200)
    ST_DEFAULT = 1
    ST_DONE = 2
    ST_CHOICES = ((ST_DEFAULT, '未处理'), (ST_DONE, '已处理'))
    status = models.IntegerField('状态', choices=ST_CHOICES, default=ST_DEFAULT)
    TY_LOGIN = 1
    TY_SESSION = 2
    TY_CHOICES = ((TY_LOGIN, '登录失败'), (TY_SESSION, '同步状态失败'))
    source_type = models.IntegerField('类型', choices=TY_CHOICES, default=TY_LOGIN)
    create_at = models.DateTimeField('时间', auto_now_add=True)

    class Meta:
        verbose_name_plural = verbose_name = '麦座登录日志'
        ordering = ['-pk']

    @classmethod
    def create_record(cls, msg, source_type, session_id=0):
        from caches import with_redis, mz_error_log_key, mz_error_sms_key
        if not cls.objects.filter(source_type=source_type, status=cls.ST_DEFAULT):
            cls.objects.create(source_type=source_type, status=cls.ST_DEFAULT, msg=msg)
        key = mz_error_log_key.format(source_type, session_id)
        with with_redis() as redis:
            num = redis.incr(key)
            if num == 1:
                redis.expire(key, 30 * 60)
            if num > 15:
                # 30分钟大于20次发一次通知，30分钟后再发
                try:
                    if not redis.get(mz_error_sms_key):
                        from mp.models import BasicConfig
                        from qcloud.sms import get_sms
                        sms = get_sms()
                        config = BasicConfig.get()
                        data = dict(mobile=config.maizuo_mobile, biz='mz', name='多个场次存在问题', reason='麦座拉取记录出现更新失败')
                        sms.smsvrcode(data)
                        redis.set(mz_error_sms_key, 1)
                        redis.expire(mz_error_sms_key, 30 * 60)
                    redis.delete(key)
                except Exception as e:
                    log.error('发送麦座同步短息消息失败')


class TicketGiveRecord(UseNoAbstract):
    user = models.ForeignKey(User, verbose_name='操作用户', on_delete=models.SET_NULL, null=True)
    order = models.ForeignKey(TicketOrder, verbose_name=u'订单', on_delete=models.SET_NULL, null=True,
                              related_name='give_order')
    session = models.ForeignKey(SessionInfo, verbose_name=u'场次', null=True, on_delete=models.CASCADE)
    mobile = models.CharField('赠送人手机号', max_length=20)
    give_mobile = models.CharField('受赠人手机号', max_length=20)
    STAT_DEFAULT = 1
    STAT_FINISH = 2
    STAT_CANCEL = 3
    STAT_CHOICES = [(STAT_DEFAULT, '待领取'), (STAT_FINISH, '已领取'), (STAT_CANCEL, '已失效')]
    status = models.SmallIntegerField('状态', choices=STAT_CHOICES, default=STAT_DEFAULT)
    create_at = models.DateTimeField('赠送时间', auto_now_add=True)
    cancel_at = models.DateTimeField('取消时间', null=True, blank=True)
    receive_user = models.ForeignKey(User, verbose_name='领取用户', on_delete=models.SET_NULL, null=True,
                                     related_name='receive_user')
    receive_at = models.DateTimeField('领取时间', null=True, blank=True)
    push_message = models.BooleanField('是否已推送开场2小时前短信', default=False, editable=False)
    push_message_day = models.BooleanField('是否已推送开场24小时前短信', default=False, editable=False)

    class Meta:
        verbose_name_plural = verbose_name = '门票赠送记录'
        ordering = ['-pk']

    @property
    def order_no(self):
        return self.order.order_no if self.order else None

    @classmethod
    def create_record(cls, user, give_mobile: str, ticket_code_qs):
        order = ticket_code_qs.first().order
        inst = cls.objects.create(user=user, order=order, give_mobile=give_mobile, mobile=user.mobile,
                                  session=order.session)
        tg = []
        for ticket_code in ticket_code_qs:
            tg.append(TicketGiveDetail(record=inst, ticket_code=ticket_code))
        if tg:
            TicketGiveDetail.objects.bulk_create(tg)
        return inst

    @atomic
    def set_cancel(self):
        self.status = self.STAT_CANCEL
        self.cancel_at = timezone.now()
        self.save(update_fields=['status', 'cancel_at'])
        qs = TicketGiveDetail.objects.filter(record_id=self.id)
        for inst in qs:
            inst.ticket_code.clear_give()

    @atomic
    def set_receive(self, receive_user):
        self.receive_user = receive_user
        self.receive_at = timezone.now()
        self.status = self.STAT_FINISH
        self.save(update_fields=['status', 'receive_user', 'receive_at'])
        qs = TicketGiveDetail.objects.filter(record_id=self.id)
        for inst in qs:
            inst.ticket_code.confirm_give()

    @classmethod
    def send_show_start_notice_give(cls):
        close_old_connections()
        hours = 2
        start_at = timezone.now() + timedelta(hours=hours)
        start_at_day = timezone.now() + timedelta(hours=24)
        from qcloud.sms import get_sms
        sms = get_sms()
        # log.error(start_at)
        two_qs = cls.objects.filter(push_message=False, session__start_at__lte=start_at, status=cls.STAT_FINISH)
        has_send_dict = dict()
        for give in two_qs:
            if not has_send_dict.get(give.mobile, None):
                has_send_dict[give.mobile] = 1
                order = give.order
                try:
                    data = dict(name=order.title, mobile=give.give_mobile, number=2,
                                time=give.session.start_at.strftime("%Y-%m-%d %H:%M"))
                    sms.smsvrcode(data)
                except Exception as e:
                    log.error('发送短息消息失败')
                give.push_message = True
                give.push_message_day = True
                give.save(update_fields=['push_message', 'push_message_day'])
        has_send_dict = dict()
        day_qs = cls.objects.filter(push_message_day=False, session__start_at__lte=start_at_day, status=cls.STAT_FINISH)
        for give in day_qs:
            if not has_send_dict.get(give.mobile, None):
                has_send_dict[give.mobile] = 1
                order = give.order
                try:
                    data = dict(name=order.title, mobile=give.give_mobile, number=24,
                                time=give.session.start_at.strftime("%Y-%m-%d %H:%M"))
                    sms.smsvrcode(data)
                except Exception as e:
                    log.error('发送短息消息失败')
                give.push_message_day = True
                give.save(update_fields=['push_message_day'])


class TicketGiveDetail(models.Model):
    record = models.ForeignKey(TicketGiveRecord, verbose_name='门票赠送记录', on_delete=models.CASCADE,
                               related_name='give_codes')
    ticket_code = models.ForeignKey(TicketUserCode, verbose_name='演出票(座位)信息', on_delete=models.CASCADE)

    class Meta:
        verbose_name_plural = verbose_name = '赠送门票'
        unique_together = ['record', 'ticket_code']

    def __str__(self):
        return str(self.ticket_code.session_seat)
